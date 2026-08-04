from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from datetime import date, timedelta, time
from django.db.models import Q
import calendar

from temple_project.apps.reservations.models import (
    Reservation, ReservationSalle, SalleReunion, BlocageCreneaux
)
from temple_project.apps.loges.models import Loge
from .forms import (
    ReservationDirecteForm, TraiteurReservationDirecteForm,
    BlocageCreneauxForm, NotificationCouvertsForm,
)
from .models import NotificationCouverts

# Horaire à partir duquel une tenue sans agapes confirmées est considérée "probable"
HEURE_SOIR = time(18, 0)


# ── Décorateurs d'accès ───────────────────────────────────────────────────────

def traiteur_required(view_func):
    from functools import wraps

    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if request.user.is_authenticated and (
            request.user.is_staff
            or request.user.groups.filter(name="Traiteur").exists()
        ):
            return view_func(request, *args, **kwargs)
        return redirect(f"/auth/traiteur/?next={request.path}")

    return wrapper


def membre_ou_traiteur_required(view_func):
    from functools import wraps

    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        est_staff    = request.user.is_authenticated and request.user.is_staff
        est_traiteur = request.user.is_authenticated and request.user.groups.filter(name="Traiteur").exists()
        est_membre   = request.COOKIES.get("kellermann_membre") == "1"
        if est_staff or est_traiteur or est_membre:
            return view_func(request, *args, **kwargs)
        return redirect(f"/auth/login/?next={request.path}")

    return wrapper


# ── Helpers ───────────────────────────────────────────────────────────────────

def _nav_mois(annee, mois):
    if mois == 1:
        mp, ap = 12, annee - 1
    else:
        mp, ap = mois - 1, annee
    if mois == 12:
        ms, as_ = 1, annee + 1
    else:
        ms, as_ = mois + 1, annee
    return mp, ap, ms, as_


def _couverts_effectifs(r):
    """Retourne (couverts_ou_None, est_estimation).
    couverts=None signifie "inconnu" — le template affiche '?'.
    Hiérarchie : nombre_repas/participants > couverts_habituels > effectif_moyen_agapes.
    """
    if hasattr(r, "nombre_repas"):
        couverts = r.nombre_repas
        loge_obj = r.loge
    else:
        couverts = r.nombre_participants
        loge_obj = r.loge if hasattr(r, "loge") else None

    if couverts and couverts > 0:
        return couverts, False

    if loge_obj:
        if getattr(loge_obj, 'couverts_habituels', None):
            return loge_obj.couverts_habituels, True
        if getattr(loge_obj, 'effectif_moyen_agapes', None):
            return loge_obj.effectif_moyen_agapes, True
    return None, False


def _contact_loge(loge):
    if not loge:
        return "", "", ""
    return loge.nom_contact or "", loge.email or "", loge.telephone or ""


def _agapes_status(r):
    """'confirme' | 'probable_classique' | 'probable' | 'aucune' pour une Reservation temple."""
    if r.besoin_agapes:
        return "confirme"
    if r.heure_debut >= HEURE_SOIR:
        # Tenue récurrente (issue d'une règle) = agapes classiques probables
        if getattr(r, 'regle_source_id', None):
            return "probable_classique"
        return "probable"
    return "aucune"


def _build_repas(r, type_label):
    """Construit un dict unifié pour la vue planning (temple ou salle)."""
    couverts, estimation = _couverts_effectifs(r)
    if type_label == "Temple":
        loge = r.loge
        lieu = str(r.temple) if r.temple else "—"
        org  = loge.nom if loge else (r.nom_organisation or r.nom_demandeur or "—")
        status = _agapes_status(r)
    else:
        loge = r.loge
        lieu = str(r.salle) if r.salle else "—"
        org  = loge.nom if loge else (r.organisation or r.nom_demandeur or "—")
        # Banquet d'ordre en salle = agapes confirmées ; salle agapes standard aussi
        tr = getattr(r, 'type_reunion', '')
        status = "confirme" if (getattr(r.salle, 'type_salle', '') == 'agapes' or tr == 'banquet') else "probable"

    nom_c, email_c, tel_c = _contact_loge(loge)
    return {
        "date":          r.date,
        "heure_debut":   r.heure_debut,
        "heure_fin":     r.heure_fin,
        "lieu":          lieu,
        "organisation":  org,
        "couverts":      couverts,
        "estimation":    estimation,
        "agapes_status": status,
        "type":          type_label,
        "contact_nom":   nom_c,
        "contact_email": email_c,
        "contact_tel":   tel_c,
        "statut":        getattr(r, "statut", "validee"),
    }


# ── Tableau de bord ───────────────────────────────────────────────────────────

@traiteur_required
def tableau_de_bord(request):
    today = date.today()

    # Navigation mensuelle : mois affiché dans le planning
    mois  = int(request.GET.get("mois",  today.month))
    annee = int(request.GET.get("annee", today.year))

    premier_jour = date(annee, mois, 1)
    dernier_jour = date(annee, mois, calendar.monthrange(annee, mois)[1])

    mois_prec, annee_prec, mois_suiv, annee_suiv = _nav_mois(annee, mois)
    nom_mois = premier_jour.strftime("%B %Y").capitalize()
    est_mois_courant = (mois == today.month and annee == today.year)

    # Tenues temple du soir + agapes confirmées pour le mois affiché
    tenues_temple = (
        Reservation.objects.filter(
            statut="validee",
            date__gte=premier_jour,
            date__lte=dernier_jour,
        )
        .filter(Q(besoin_agapes=True) | Q(heure_debut__gte=HEURE_SOIR))
        .select_related("loge", "temple")
        .order_by("date", "heure_debut")
    )

    # Banquets/agapes salles pour le mois affiché
    tenues_salle = (
        ReservationSalle.objects.filter(
            date__gte=premier_jour,
            date__lte=dernier_jour,
            statut__in=["attente", "validee"],
        )
        .filter(Q(salle__type_salle="agapes") | Q(type_reunion="banquet"))
        .select_related("loge", "salle")
        .order_by("date", "heure_debut")
    )

    tous_repas = [_build_repas(r, "Temple") for r in tenues_temple]
    tous_repas += [_build_repas(r, "Salle") for r in tenues_salle]
    tous_repas.sort(key=lambda x: (x["date"], x["heure_debut"]))

    # Séparer aujourd'hui / reste du mois (uniquement si mois courant)
    repas_aujourd_hui  = [r for r in tous_repas if r["date"] == today] if est_mois_courant else []
    repas_mois         = [r for r in tous_repas if r["date"] != today or not est_mois_courant]
    if est_mois_courant:
        repas_mois = [r for r in tous_repas if r["date"] > today]

    nb_confirmes = sum(1 for r in tous_repas if r["agapes_status"] == "confirme")
    nb_probables = sum(1 for r in tous_repas if r["agapes_status"] in ("probable", "probable_classique"))

    blocages = (
        BlocageCreneaux.objects
        .filter(Q(date_fin__gte=today) | Q(date_fin__isnull=True, date__gte=today))
        .prefetch_related("salles")
        .order_by("date")[:5]
    )

    notifications = (
        NotificationCouverts.objects.filter(statut="non_lu")
        .select_related("loge")
        .order_by("-created_at")
    )

    return render(request, "traiteur/tableau_de_bord.html", {
        "repas_aujourd_hui": repas_aujourd_hui,
        "repas_mois":        repas_mois,
        "nb_confirmes":      nb_confirmes,
        "nb_probables":      nb_probables,
        "blocages":          blocages,
        "notifications":     notifications,
        "nb_notifications":  notifications.count(),
        "today":             today,
        "nom_mois":          nom_mois,
        "mois":              mois,
        "annee":             annee,
        "mois_prec":         mois_prec,
        "annee_prec":        annee_prec,
        "mois_suiv":         mois_suiv,
        "annee_suiv":        annee_suiv,
        "est_mois_courant":  est_mois_courant,
        "premier_jour":      premier_jour,
        "dernier_jour":      dernier_jour,
    })


@traiteur_required
def marquer_notification_lue(request, pk):
    notif = get_object_or_404(NotificationCouverts, pk=pk)
    if request.method == "POST":
        notif.statut = "lu"
        notif.save()
    return redirect("traiteur:tableau_de_bord")


# ── Calendrier ────────────────────────────────────────────────────────────────

@traiteur_required
def calendrier(request):
    today  = date.today()
    annee  = int(request.GET.get("annee", today.year))
    mois   = int(request.GET.get("mois",  today.month))
    filtre = request.GET.get("filtre", "tout")   # tout | soir | agapes | blocages

    mois_prec, annee_prec, mois_suiv, annee_suiv = _nav_mois(annee, mois)
    premier_jour = date(annee, mois, 1)
    dernier_jour = date(annee, mois, calendar.monthrange(annee, mois)[1])

    reservations = (
        Reservation.objects.filter(statut="validee", date__gte=premier_jour, date__lte=dernier_jour)
        .select_related("loge", "temple")
        .order_by("date", "heure_debut")
    )
    reservations_salles = (
        ReservationSalle.objects.filter(
            date__gte=premier_jour, date__lte=dernier_jour,
            statut__in=["attente", "validee"],
        )
        .filter(Q(salle__type_salle="agapes") | Q(type_reunion="banquet"))
        .select_related("salle", "loge")
        .order_by("date", "heure_debut")
    )
    blocages = (
        BlocageCreneaux.objects.filter(date__lte=dernier_jour)
        .filter(Q(date_fin__gte=premier_jour) | Q(date_fin__isnull=True, date__gte=premier_jour))
        .prefetch_related("salles")
        .order_by("date", "heure_debut")
    )

    if filtre == "agapes":
        # Tout ce qui implique de la nourriture : agapes confirmées + tenues de soir + banquets
        reservations = reservations.filter(Q(besoin_agapes=True) | Q(heure_debut__gte=HEURE_SOIR))
        # reservations_salles reste (banquets + salles agapes)
    elif filtre == "classiques":
        # Loges récurrentes du soir (règle de récurrence) — agapes classiques
        reservations = reservations.filter(
            heure_debut__gte=HEURE_SOIR,
            regle_source__isnull=False,
            besoin_agapes=False,
        )
        reservations_salles = reservations_salles.none()
    elif filtre == "banquets":
        # Uniquement les réservations de salles agapes / banquets d'ordre
        reservations = reservations.none()
        # reservations_salles reste
    elif filtre == "soiree":
        # Réunions exceptionnelles en soirée, hors récurrences et hors agapes conf.
        # (ex. CAALA, congrès, assemblées exceptionnelles)
        reservations = reservations.filter(
            heure_debut__gte=HEURE_SOIR,
            regle_source__isnull=True,
            besoin_agapes=False,
        )
        reservations_salles = reservations_salles.none()
    elif filtre == "blocages":
        reservations        = reservations.none()
        reservations_salles = reservations_salles.none()

    events_by_date = {}
    for r in reservations:
        couverts, est = _couverts_effectifs(r)
        status = _agapes_status(r)
        nom_c, email_c, tel_c = _contact_loge(r.loge)
        events_by_date.setdefault(r.date, []).append({
            "type": "reservation", "obj": r,
            "agapes": r.besoin_agapes, "agapes_status": status,
            "statut": r.statut,
            "couverts": couverts, "estimation": est,
            "contact_nom": nom_c, "contact_email": email_c, "contact_tel": tel_c,
        })
    for r in reservations_salles:
        couverts, est = _couverts_effectifs(r)
        nom_c, email_c, tel_c = _contact_loge(r.loge)
        tr = getattr(r, 'type_reunion', '')
        salle_type = getattr(r.salle, 'type_salle', '') if r.salle else ''
        is_banquet = (salle_type == 'agapes' or tr == 'banquet')
        salle_status = "confirme" if is_banquet else "probable"
        events_by_date.setdefault(r.date, []).append({
            "type": "salle", "obj": r,
            "is_banquet": is_banquet,
            "agapes": True, "agapes_status": salle_status,
            "statut": r.statut,
            "couverts": couverts, "estimation": est,
            "contact_nom": nom_c, "contact_email": email_c, "contact_tel": tel_c,
        })
    for b in blocages:
        b_fin = b.date_fin or b.date
        cur = b.date
        while cur <= b_fin:
            if cur.month == mois and cur.year == annee:
                events_by_date.setdefault(cur, []).append({
                    "type": "blocage", "obj": b, "agapes": False,
                    "agapes_status": "aucune", "couverts": None, "estimation": False,
                    "contact_nom": "", "contact_email": "", "contact_tel": "",
                })
            from datetime import timedelta
            cur += timedelta(days=1)

    mois_choices = []
    for delta in range(-3, 15):
        m = today.month + delta
        y = today.year + (m - 1) // 12
        m = ((m - 1) % 12) + 1
        mois_choices.append({
            "mois": m, "annee": y,
            "label": date(y, m, 1).strftime("%B %Y").capitalize(),
            "actif": m == mois and y == annee,
        })

    # Liste triée (date, evts) pour la vue agenda mobile
    events_sorted = sorted(events_by_date.items())

    return render(request, "traiteur/calendrier.html", {
        "annee": annee, "mois": mois,
        "nom_mois":       premier_jour.strftime("%B %Y").capitalize(),
        "cal":            calendar.monthcalendar(annee, mois),
        "events_by_date": events_by_date,
        "events_sorted":  events_sorted,
        "today":          today,
        "mois_prec": mois_prec, "annee_prec": annee_prec,
        "mois_suiv": mois_suiv, "annee_suiv": annee_suiv,
        "filtre":    filtre,
        "mois_choices": mois_choices,
        "reservations":        reservations,
        "reservations_salles": reservations_salles,
        "blocages":            blocages,
    })


# ── Planning mensuel ──────────────────────────────────────────────────────────

@traiteur_required
def planning(request):
    today  = date.today()
    annee  = int(request.GET.get("annee", today.year))
    mois   = int(request.GET.get("mois",  today.month))
    filtre = request.GET.get("filtre", "soir")  # agapes | soir | tout

    mois_prec, annee_prec, mois_suiv, annee_suiv = _nav_mois(annee, mois)
    premier_jour = date(annee, mois, 1)
    dernier_jour = date(annee, mois, calendar.monthrange(annee, mois)[1])

    qs_temple = (
        Reservation.objects.filter(statut="validee", date__gte=premier_jour, date__lte=dernier_jour)
        .select_related("loge", "temple")
        .order_by("date", "heure_debut")
    )
    if filtre == "agapes":
        qs_temple = qs_temple.filter(besoin_agapes=True)
    elif filtre == "soir":
        qs_temple = qs_temple.filter(Q(besoin_agapes=True) | Q(heure_debut__gte=HEURE_SOIR))
    # filtre == "tout" : pas de filtre supplémentaire

    qs_salle = (
        ReservationSalle.objects.filter(
            date__gte=premier_jour, date__lte=dernier_jour,
            statut__in=["attente", "validee"],
        )
        .filter(Q(salle__type_salle="agapes") | Q(type_reunion="banquet"))
        .select_related("loge", "salle")
        .order_by("date", "heure_debut")
    )

    repas = [_build_repas(r, "Temple") for r in qs_temple]
    repas += [_build_repas(r, "Salle")  for r in qs_salle]
    repas.sort(key=lambda x: (x["date"], x["heure_debut"]))

    total_confirmes   = sum(
        r["couverts"] for r in repas
        if r["agapes_status"] == "confirme" and not r["estimation"] and r["couverts"]
    )
    total_estimations = sum(r["couverts"] for r in repas if r["estimation"] and r["couverts"])
    nb_probables      = sum(1 for r in repas if r["agapes_status"] in ("probable", "probable_classique"))
    nb_inconnus       = sum(1 for r in repas if r["couverts"] is None)

    return render(request, "traiteur/planning.html", {
        "repas":             repas,
        "total_confirmes":   total_confirmes,
        "total_estimations": total_estimations,
        "nb_probables":      nb_probables,
        "nb_inconnus":       nb_inconnus,
        "filtre":            filtre,
        "annee": annee, "mois": mois,
        "nom_mois": premier_jour.strftime("%B %Y").capitalize(),
        "mois_prec": mois_prec, "annee_prec": annee_prec,
        "mois_suiv": mois_suiv, "annee_suiv": annee_suiv,
        "today": today,
    })


# ── Réservation directe (traiteur) ────────────────────────────────────────────

@traiteur_required
def reserver(request):
    form = TraiteurReservationDirecteForm(request.POST or None)

    if request.method == "POST" and form.is_valid():
        cd       = form.cleaned_data
        loge     = cd.get("loge")
        org      = cd.get("organisation") or ""
        salle    = cd["salle"]
        date_r   = cd["date"]
        hd, hf   = cd["heure_debut"], cd["heure_fin"]
        couverts = cd.get("nombre_repas") or 0
        note     = cd.get("commentaire") or ""

        ReservationSalle.objects.create(
            loge=loge, salle=salle, date=date_r,
            heure_debut=hd, heure_fin=hf, statut="validee",
            nom_demandeur=loge.nom if loge else org,
            email_demandeur="traiteur@kellermann.local",
            organisation=loge.nom if loge else org,
            objet=note or "Agapes",
            nombre_participants=couverts,
            commentaire=note,
        )
        messages.success(request, f"Réservation créée sur {salle.nom} le {date_r:%d/%m/%Y}.")
        return redirect("traiteur:planning")

    return render(request, "traiteur/reserver.html", {"form": form})


# ── Couverts habituels par loge ───────────────────────────────────────────────

@traiteur_required
def loges_couverts(request):
    """Liste des loges avec couverts habituels éditables par le traiteur."""
    if request.method == "POST":
        loge_pk  = request.POST.get("loge_pk", "").strip()
        val      = request.POST.get("couverts", "").strip()
        if loge_pk:
            try:
                loge = Loge.objects.get(pk=int(loge_pk), actif=True)
                loge.couverts_habituels = int(val) if val else None
                loge.save(update_fields=["couverts_habituels"])
                messages.success(request, f"Mis à jour : {loge.nom}.")
            except (Loge.DoesNotExist, ValueError):
                messages.error(request, "Valeur invalide.")
        return redirect("traiteur:loges_couverts")

    loges = (
        Loge.objects.filter(actif=True)
        .select_related("obedience")
        .order_by("nom")
    )
    return render(request, "traiteur/loges_couverts.html", {"loges": loges})


# ── Blocages ──────────────────────────────────────────────────────────────────

@traiteur_required
def bloquer(request):
    form = BlocageCreneauxForm(request.POST or None)

    if request.method == "POST" and form.is_valid():
        blocage = form.save(commit=False)
        blocage.created_by = request.user
        blocage.save()
        form.save_m2m()
        periode = f"{blocage.date:%d/%m/%Y}"
        if blocage.date_fin and blocage.date_fin != blocage.date:
            periode += f" → {blocage.date_fin:%d/%m/%Y}"
        messages.success(request, f"Créneau bloqué : {periode} {blocage.heure_debut:%H:%M}–{blocage.heure_fin:%H:%M}.")
        return redirect("traiteur:calendrier")

    today_d = date.today()
    blocages = (
        BlocageCreneaux.objects
        .filter(Q(date_fin__gte=today_d) | Q(date_fin__isnull=True, date__gte=today_d))
        .prefetch_related("salles")
        .order_by("date")
    )
    return render(request, "traiteur/bloquer.html", {"form": form, "blocages": blocages})


@traiteur_required
def etat_des_lieux(request):
    """Récapitulatif de toutes les tenues + banquets jusqu'en fin d'année courante."""
    today_d = date.today()
    fin_annee = date(today_d.year, 12, 31)

    tenues_temple = (
        Reservation.objects.filter(
            statut="validee",
            date__gte=today_d,
            date__lte=fin_annee,
        )
        .filter(Q(besoin_agapes=True) | Q(heure_debut__gte=HEURE_SOIR))
        .select_related("loge", "temple")
        .order_by("date", "heure_debut")
    )
    banquets_salles = (
        ReservationSalle.objects.filter(
            date__gte=today_d,
            date__lte=fin_annee,
            statut__in=["attente", "validee"],
        )
        .filter(Q(salle__type_salle="agapes") | Q(type_reunion="banquet"))
        .select_related("loge", "salle")
        .order_by("date", "heure_debut")
    )

    repas = [_build_repas(r, "Temple") for r in tenues_temple]
    repas += [_build_repas(r, "Salle") for r in banquets_salles]
    repas.sort(key=lambda x: (x["date"], x["heure_debut"]))

    total_agapes  = sum(1 for r in repas if r["agapes_status"] == "confirme")
    total_probables = sum(1 for r in repas if r["agapes_status"] in ("probable", "probable_classique"))
    total_couverts = sum(r["couverts"] for r in repas if r["couverts"] and not r["estimation"])
    total_couverts_est = sum(r["couverts"] for r in repas if r["couverts"] and r["estimation"])

    # Grouper par mois
    from itertools import groupby
    def mois_key(r): return (r["date"].year, r["date"].month)
    par_mois = []
    for (y, m), items in groupby(repas, mois_key):
        lst = list(items)
        par_mois.append({
            "label": date(y, m, 1).strftime("%B %Y").capitalize(),
            "repas": lst,
            "nb_agapes": sum(1 for r in lst if r["agapes_status"] == "confirme"),
            "nb_couverts": sum(r["couverts"] for r in lst if r["couverts"] and not r["estimation"]),
        })

    return render(request, "traiteur/etat_des_lieux.html", {
        "par_mois":         par_mois,
        "total_agapes":     total_agapes,
        "total_probables":  total_probables,
        "total_couverts":   total_couverts,
        "total_couverts_est": total_couverts_est,
        "fin_annee":        fin_annee,
        "today":            today_d,
        "export_url":       f"/traiteur/export-agapes/?date_debut={today_d}&date_fin={fin_annee}&type_export=soir",
    })


@traiteur_required
def guide_traiteur(request):
    return render(request, "traiteur/guide.html", {})


@traiteur_required
def contact_traiteur(request):
    """Page contacts : coordonnées admin + loges avec tenues à venir."""
    from django.contrib.auth import get_user_model
    User = get_user_model()

    today   = date.today()
    horizon = today + timedelta(days=60)

    # Contacts admin (staff)
    admins = list(
        User.objects.filter(is_staff=True)
        .exclude(email="")
        .values("first_name", "last_name", "email")
        .order_by("first_name")
    )

    # Loges avec tenues de soir ou agapes dans les 60 prochains jours
    tenues = (
        Reservation.objects.filter(
            statut="validee",
            date__gte=today,
            date__lte=horizon,
        )
        .filter(Q(besoin_agapes=True) | Q(heure_debut__gte=HEURE_SOIR))
        .select_related("loge", "temple")
        .order_by("loge__nom", "date")
    )

    # Dédupliquer par loge, garder la date de prochaine tenue
    loges_vues = {}
    for t in tenues:
        if not t.loge:
            continue
        loge = t.loge
        if loge.pk not in loges_vues:
            loges_vues[loge.pk] = {
                "nom":         loge.nom,
                "contact_nom": loge.nom_contact or "",
                "email":       loge.email or "",
                "telephone":   loge.telephone or "",
                "prochaine":   t.date,
                "agapes":      _agapes_status(t),
            }

    contacts_loges = sorted(loges_vues.values(), key=lambda x: x["prochaine"])

    return render(request, "traiteur/contact.html", {
        "admins":         admins,
        "contacts_loges": contacts_loges,
        "horizon":        horizon,
        "today":          today,
    })


@traiteur_required
def supprimer_blocage(request, pk):
    blocage = get_object_or_404(BlocageCreneaux, pk=pk)
    if request.method == "POST":
        blocage.delete()
        messages.success(request, "Blocage supprimé.")
    return redirect("traiteur:bloquer")


# ── Export Excel ──────────────────────────────────────────────────────────────

@traiteur_required
def export_agapes_excel(request):
    """Export Excel agapes / tenues soir / banquets — période et type filtrables."""
    from datetime import datetime as dt
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse

    today          = date.today()
    annee_courante = today.year if today.month >= 9 else today.year - 1
    debut_s        = request.GET.get('date_debut', '')
    fin_s          = request.GET.get('date_fin', '')
    type_export    = request.GET.get('type_export', 'soir')  # tout|agapes|banquet|soir

    try:
        debut = dt.strptime(debut_s, '%Y-%m-%d').date() if debut_s else date(annee_courante, 9, 1)
        fin   = dt.strptime(fin_s,   '%Y-%m-%d').date() if fin_s   else date(annee_courante + 1, 6, 30)
    except ValueError:
        debut = date(annee_courante, 9, 1)
        fin   = date(annee_courante + 1, 6, 30)

    lignes = []

    # Tenues temple
    if type_export in ('tout', 'agapes', 'soir'):
        qs_t = (
            Reservation.objects.filter(statut='validee', date__gte=debut, date__lte=fin)
            .select_related('loge', 'temple')
            .order_by('date')
        )
        if type_export == 'agapes':
            qs_t = qs_t.filter(besoin_agapes=True)
        elif type_export == 'soir':
            qs_t = qs_t.filter(Q(besoin_agapes=True) | Q(heure_debut__gte=HEURE_SOIR))

        for r in qs_t:
            couverts, est = _couverts_effectifs(r)
            loge = r.loge
            lignes.append({
                "date":         r.date,
                "organisation": loge.nom if loge else (r.nom_organisation or r.nom_demandeur or ''),
                "type_label":   "✓ Agapes conf." if r.besoin_agapes else "~ Soir (probable)",
                "couverts":     couverts,
                "estimation":   est,
                "lieu":         str(r.temple) if r.temple else '',
                "horaires":     f"{r.heure_debut:%H:%M} – {r.heure_fin:%H:%M}",
                "commentaire":  r.commentaire or '',
                "contact_nom":  loge.nom_contact if loge else '',
                "contact_email": loge.email if loge else '',
                "contact_tel":  loge.telephone if loge else '',
            })

    # Banquets / salles agapes
    if type_export in ('tout', 'banquet', 'soir'):
        for b in (
            ReservationSalle.objects.filter(
                salle__type_salle='agapes', statut='validee', date__gte=debut, date__lte=fin
            )
            .select_related('loge', 'salle')
            .order_by('date')
        ):
            couverts, est = _couverts_effectifs(b)
            loge = b.loge
            lignes.append({
                "date":         b.date,
                "organisation": loge.nom if loge else (b.organisation or b.nom_demandeur or ''),
                "type_label":   "Banquet d'ordre",
                "couverts":     couverts,
                "estimation":   est,
                "lieu":         str(b.salle),
                "horaires":     f"{b.heure_debut:%H:%M} – {b.heure_fin:%H:%M}",
                "commentaire":  b.commentaire or '',
                "contact_nom":  loge.nom_contact if loge else '',
                "contact_email": loge.email if loge else '',
                "contact_tel":  loge.telephone if loge else '',
            })

    lignes.sort(key=lambda x: x["date"])

    # ── Construction Excel ──────────────────────────────────────────────────
    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = f"Agapes {debut:%d%m%Y}-{fin:%d%m%Y}"

    BLEU = "0F2137"; OR = "C8A84B"; GRIS = "F1F5F9"
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin'),
    )
    ctr = Alignment(horizontal="center", vertical="center")

    headers    = ["Date", "Loge / Organisation", "Type", "Couverts", "Est.",
                  "Lieu", "Horaires", "Commentaire", "Contact", "Email", "Téléphone"]
    col_widths = [14, 36, 22, 10, 6, 22, 18, 30, 22, 28, 16]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font      = Font(bold=True, color=OR)
        c.fill      = PatternFill("solid", fgColor=BLEU)
        c.alignment = ctr
        c.border    = thin
        ws.column_dimensions[get_column_letter(col)].width = w

    for ri, l in enumerate(lignes, 2):
        bg = GRIS if ri % 2 == 0 else "FFFFFF"
        vals = [
            l["date"].strftime('%d/%m/%Y'),
            l["organisation"],
            l["type_label"],
            l["couverts"] if l["couverts"] else "?",
            "estim." if l["estimation"] else ("?" if l["couverts"] is None else ""),
            l["lieu"],
            l["horaires"],
            l["commentaire"],
            l["contact_nom"],
            l["contact_email"],
            l["contact_tel"],
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=ri, column=col, value=val)
            c.border = thin
            c.fill   = PatternFill("solid", fgColor=bg)
            if col in (4, 5):
                c.alignment = ctr

    # Ligne TOTAL
    row_tot = len(lignes) + 3
    c = ws.cell(row=row_tot, column=1, value="TOTAL")
    c.font = Font(bold=True)
    total_cov = sum(l["couverts"] for l in lignes if l["couverts"])
    c = ws.cell(row=row_tot, column=4, value=total_cov)
    c.font = Font(bold=True)
    inconnus = sum(1 for l in lignes if not l["couverts"])
    if inconnus:
        ws.cell(row=row_tot, column=5, value=f"({inconnus} sans données)")

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="agapes_{debut:%d%m%Y}-{fin:%d%m%Y}.xlsx"'
    wb.save(response)
    return response


# ── Notification couverts (accessible aux membres) ────────────────────────────

@membre_ou_traiteur_required
def notification(request):
    form = NotificationCouvertsForm(request.POST or None)

    if request.method == "POST" and form.is_valid():
        cd = form.cleaned_data
        notif = NotificationCouverts.objects.create(
            loge=cd["loge"],
            date_tenue=cd["date_tenue"],
            nombre_couverts=cd["nombre_couverts"],
            commentaire=cd.get("commentaire") or "",
            email_contact=cd["email_contact"],
        )
        _envoyer_email_notification_traiteur(notif)
        _envoyer_email_confirmation_demandeur(notif)
        messages.success(
            request,
            f"Notification transmise. Confirmation envoyée à {notif.email_contact}."
        )
        return redirect("traiteur:notification_confirmee")

    return render(request, "traiteur/notification.html", {"form": form})


@membre_ou_traiteur_required
def notification_confirmee(request):
    return render(request, "traiteur/notification_confirmee.html")


# ── Emails ────────────────────────────────────────────────────────────────────

def _envoyer_email_notification_traiteur(notif):
    try:
        from temple_project.apps.administration.email_utils import (
            send_mail_kellermann, get_email_admin,
        )
        sujet = f"[Traiteur] Notification couverts — {notif.loge} — {notif.date_tenue:%d/%m/%Y}"
        corps = (
            f"Notification de couverts.\n\n"
            f"Loge       : {notif.loge}\n"
            f"Date tenue : {notif.date_tenue:%d/%m/%Y}\n"
            f"Couverts   : {notif.nombre_couverts}\n"
            f"Commentaire: {notif.commentaire or '—'}\n"
            f"Contact    : {notif.email_contact}\n"
        )
        send_mail_kellermann(sujet, corps, [get_email_admin()], fail_silently=True)
    except Exception:
        pass


def _envoyer_email_confirmation_demandeur(notif):
    try:
        from temple_project.apps.administration.email_utils import send_mail_kellermann
        sujet = f"[Kellermann] Notification reçue — {notif.date_tenue:%d/%m/%Y}"
        corps = (
            f"Bonjour,\n\n"
            f"Votre notification a bien été transmise au traiteur.\n\n"
            f"Loge       : {notif.loge}\n"
            f"Date tenue : {notif.date_tenue:%d/%m/%Y}\n"
            f"Couverts   : {notif.nombre_couverts}\n"
            f"Commentaire: {notif.commentaire or '—'}\n\n"
            f"Cordialement,\nKellermann Réservations"
        )
        send_mail_kellermann(sujet, corps, [notif.email_contact], fail_silently=True)
    except Exception:
        pass
