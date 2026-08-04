from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.conf import settings
from temple_project.apps.administration.email_utils import send_mail_kellermann, get_email_admin, get_email_traiteur
from django.http import HttpResponse
from datetime import date, timedelta
import calendar
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from temple_project.apps.reservations.models import (
    Reservation, RegleRecurrence, Temple, SalleReunion, ReservationSalle,
    DemandeAccesPortail, ValidationSaison, ValidationSaisonLigne,
    Indisponibilite, BlocageCreneaux, RegleRecurrenceSalle,
    DemandeRegleRecurrenceSalle,
)
from temple_project.apps.loges.models import Loge, Obedience
from .models import Parametres, JournalEvenement, Annonce
from .journal import log_evenement
from django.utils import timezone
from django.utils.dateparse import parse_datetime
from functools import wraps
from django.core.exceptions import PermissionDenied
from django.contrib.auth.views import redirect_to_login


def staff_required(view_func):
    """Réserve la vue aux administrateurs (is_staff) : les membres connectés non
    staff reçoivent un 403, les visiteurs sont redirigés vers la connexion.
    Protège notamment les coordonnées des loges (annuaire, fiches)."""
    @wraps(view_func)
    def _wrapped(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect_to_login(request.get_full_path())
        if not request.user.is_staff:
            raise PermissionDenied
        return view_func(request, *args, **kwargs)
    return _wrapped


# ── Tableau de bord ───────────────────────────────────────────────────────────

@staff_required
def tableau_de_bord(request):
    reservations_attente  = Reservation.objects.filter(statut='attente').select_related('loge', 'temple').order_by('date')
    reservations_recentes = Reservation.objects.order_by('-created_at')[:10]
    reservations_salle_attente = ReservationSalle.objects.filter(
        statut='attente'
    ).select_related('salle').order_by('date')
    demandes_portail_attente = DemandeAccesPortail.objects.filter(statut='attente').order_by('created_at')
    demandes_recsalle_attente = DemandeRegleRecurrenceSalle.objects.filter(statut='attente').select_related('loge').order_by('date_demande')
    from temple_project.apps.reservations.models import MessageContact
    messages_nouveaux = MessageContact.objects.filter(statut='nouveau').order_by('-created_at')
    context = {
        'attente':                  reservations_attente,
        'recentes':                 reservations_recentes,
        'nb_attente':               reservations_attente.count(),
        'nb_loges':                 Loge.objects.filter(actif=True).count(),
        'nb_reservations':          Reservation.objects.count(),
        'nb_regles':                RegleRecurrence.objects.filter(actif=True).count(),
        'attente_salles':           reservations_salle_attente,
        'nb_attente_salles':        reservations_salle_attente.count(),
        'demandes_portail':         demandes_portail_attente,
        'nb_demandes_portail':      demandes_portail_attente.count(),
        'demandes_recsalle':        demandes_recsalle_attente,
        'nb_demandes_recsalle':     demandes_recsalle_attente.count(),
        'messages_nouveaux':        messages_nouveaux,
        'nb_messages_nx':           messages_nouveaux.count(),
    }
    return render(request, 'administration/tableau_de_bord.html', context)


# ── Tarification ──────────────────────────────────────────────────────────────

def _exceptionnelle_facturable(date_):
    """Une occupation exceptionnelle n'est facturée que le week-end (samedi/dimanche)
    ou hors période d'ouverture (juillet/août). Les tenues de semaine en saison ne
    sont pas facturées."""
    return date_.weekday() >= 5 or date_.month in (7, 8)


def _reservation_facturable(resa, params):
    """Détermine si une réservation est facturable selon toutes les règles."""
    # Jamais les réservations récurrentes (générées par une règle)
    if resa.regle_source_id:
        return False
    # Tarifs non rétroactifs : rien avant leur date d'entrée en vigueur (vote AG)
    if params.tarif_date_effet and resa.date < params.tarif_date_effet:
        return False
    if resa.type_reservation == 'congres':
        return True
    if resa.type_reservation == 'exceptionnelle':
        return _exceptionnelle_facturable(resa.date)
    return False


def tarif_reservation(resa, params=None):
    """Tarif d'une réservation (0 si non facturable)."""
    from decimal import Decimal
    if params is None:
        params = Parametres.get_instance()
    if not _reservation_facturable(resa, params):
        return Decimal('0')
    if resa.type_reservation == 'congres':
        jours = 1
        if resa.date_fin and resa.date_fin > resa.date:
            jours = (resa.date_fin - resa.date).days
        return params.tarif_congres_jour * jours
    # Tenue funèbre : tarif spécifique (week-end / vacances)
    if resa.sous_type == 'funebre':
        return params.tarif_funebre
    return params.tarif_exc_avec_agapes if resa.besoin_agapes else params.tarif_exc_sans_agapes


# ── Validation réservations ───────────────────────────────────────────────────

@staff_required
def valider_reservation(request, pk):
    resa = get_object_or_404(Reservation, pk=pk)

    # Détection conflits (réservations validées qui chevauchent, même temple)
    conflits = Reservation.objects.filter(
        temple=resa.temple,
        date=resa.date,
        statut='validee',
        heure_debut__lt=resa.heure_fin,
        heure_fin__gt=resa.heure_debut,
    ).exclude(pk=pk).select_related('loge')

    # Autres demandes EN ATTENTE sur le même créneau (priorité au premier)
    demandes_attente = Reservation.objects.filter(
        temple=resa.temple, date=resa.date,
        heure_debut__lt=resa.heure_fin, heure_fin__gt=resa.heure_debut,
        statut='attente',
    ).exclude(pk=pk).select_related('loge')

    # Blocages / indisponibilités sur le temple
    blocages = BlocageCreneaux.objects.filter(
        temples=resa.temple, date=resa.date,
        heure_debut__lt=resa.heure_fin, heure_fin__gt=resa.heure_debut,
    )
    indisponibilites = Indisponibilite.objects.filter(
        temples=resa.temple, date_debut__lte=resa.date, date_fin__gte=resa.date,
    )

    # Temples alternatifs libres si le créneau est occupé
    temples_alternatives = []
    if conflits.exists() or blocages.exists() or indisponibilites.exists():
        for tp in Temple.objects.exclude(pk=resa.temple_id).order_by('nom'):
            if not _temple_occupe(tp, resa.date, resa.heure_debut, resa.heure_fin):
                temples_alternatives.append(tp)

    if request.method == 'POST':
        action            = request.POST.get('action')
        commentaire_admin = request.POST.get('commentaire_admin', '').strip()

        if action not in ('valider', 'refuser'):
            messages.error(request, "Action invalide.")
            return redirect('administration:tableau_de_bord')

        resa.statut = 'validee' if action == 'valider' else 'refusee'
        if action == 'valider':
            # L'admin peut corriger le type avant que le tarif soit figé
            type_corrige = request.POST.get('type_reservation', '').strip()
            if type_corrige in ('reguliere', 'exceptionnelle', 'congres'):
                resa.type_reservation = type_corrige
            # Fige le tarif en vigueur au moment de la validation
            resa.tarif = tarif_reservation(resa)
        resa.save()

        _envoyer_email_decision(resa, action, commentaire_admin)

        if action == 'valider':
            messages.success(request, f"Demande de {resa.loge} validée — email envoyé à {resa.email_demandeur}.")
            log_evenement('validation_reservation',
                f"Réservation validée : {resa.loge} — {resa.date:%d/%m/%Y} {resa.heure_debut:%H:%M}–{resa.heure_fin:%H:%M} ({resa.temple})",
                request=request, objet=resa)
        else:
            messages.warning(request, f"Demande de {resa.loge} refusée — email envoyé à {resa.email_demandeur}.")
            log_evenement('refus_reservation',
                f"Réservation refusée : {resa.loge} — {resa.date:%d/%m/%Y} {resa.heure_debut:%H:%M}–{resa.heure_fin:%H:%M} ({resa.temple})",
                request=request, objet=resa)

        return redirect('administration:tableau_de_bord')

    occupants_rec = _occupants_recurrents(
        resa.temple, resa.date, resa.heure_debut, resa.heure_fin, exclure_pk=resa.pk
    ) if conflits.exists() else []

    return render(request, 'administration/valider_reservation.html', {
        'reservation':         resa,
        'conflits':            conflits,
        'demandes_attente':    demandes_attente,
        'blocages':            blocages,
        'indisponibilites':    indisponibilites,
        'temples_alternatives': temples_alternatives,
        'occupants_recurrents': occupants_rec,
        'echange_next':         request.get_full_path(),
    })


@staff_required
def valider_reservation_salle(request, pk):
    resa = get_object_or_404(ReservationSalle, pk=pk)
    is_cabinet = resa.salle.type_salle == 'cabinet_reflexion'

    # Détection conflits (non applicable aux cabinets — chacun a sa propre salle)
    conflits = ReservationSalle.objects.filter(
        salle=resa.salle,
        date=resa.date,
        statut='validee',
        heure_debut__lt=resa.heure_fin,
        heure_fin__gt=resa.heure_debut,
    ).exclude(pk=pk)

    # Disponibilité des cabinets (en excluant la demande en cours)
    cabinets_dispo = []
    if is_cabinet:
        for cabinet in SalleReunion.objects.filter(type_salle='cabinet_reflexion', actif=True).order_by('nom'):
            occupe = ReservationSalle.objects.filter(
                salle=cabinet,
                date=resa.date,
                heure_debut__lt=resa.heure_fin,
                heure_fin__gt=resa.heure_debut,
                statut__in=['attente', 'validee'],
            ).exclude(pk=pk).exists()
            cabinets_dispo.append({
                'cabinet': cabinet,
                'libre': not occupe,
                'prefere': resa.cabinet_prefere_id == cabinet.pk,
            })

    # Autres demandes EN ATTENTE sur le même créneau (priorité au premier)
    demandes_attente = ReservationSalle.objects.filter(
        salle=resa.salle, date=resa.date,
        heure_debut__lt=resa.heure_fin, heure_fin__gt=resa.heure_debut,
        statut='attente',
    ).exclude(pk=pk).select_related('loge')

    # Blocages / indisponibilités sur la salle
    blocages = BlocageCreneaux.objects.filter(
        salles=resa.salle, date=resa.date,
        heure_debut__lt=resa.heure_fin, heure_fin__gt=resa.heure_debut,
    )
    indisponibilites = Indisponibilite.objects.filter(
        salles=resa.salle, date_debut__lte=resa.date, date_fin__gte=resa.date,
    )

    # Salles alternatives libres (même type) si le créneau est occupé
    salles_alternatives = []
    if not is_cabinet and (conflits.exists() or blocages.exists() or indisponibilites.exists()):
        for s in SalleReunion.objects.filter(
            type_salle=resa.salle.type_salle, actif=True,
        ).exclude(pk=resa.salle_id).order_by('nom'):
            if not _salle_occupee(s, resa.date, resa.heure_debut, resa.heure_fin):
                salles_alternatives.append(s)

    if request.method == 'POST':
        action            = request.POST.get('action')
        commentaire_admin = request.POST.get('commentaire_admin', '').strip()

        if action not in ('valider', 'refuser'):
            messages.error(request, "Action invalide.")
            return redirect('administration:tableau_de_bord')

        if action == 'valider' and is_cabinet:
            cabinet_attribue_id = request.POST.get('cabinet_attribue_id')
            if cabinet_attribue_id:
                try:
                    resa.salle = SalleReunion.objects.get(pk=cabinet_attribue_id, type_salle='cabinet_reflexion')
                except SalleReunion.DoesNotExist:
                    pass

        resa.statut = 'validee' if action == 'valider' else 'refusee'
        if 'facturable' in request.POST:
            resa.facturable = request.POST.get('facturable') == 'on'
        resa.save()

        _envoyer_email_decision_salle(resa, action, commentaire_admin)

        if action == 'valider':
            messages.success(request, f"Demande de salle pour {resa.organisation} validée — email envoyé à {resa.email_demandeur}.")
            log_evenement('validation_reservation',
                f"Réservation salle validée : {resa.organisation} — {resa.date:%d/%m/%Y} {resa.heure_debut:%H:%M}–{resa.heure_fin:%H:%M} ({resa.salle})",
                request=request, objet=resa)
        else:
            messages.warning(request, f"Demande de salle pour {resa.organisation} refusée — email envoyé à {resa.email_demandeur}.")
            log_evenement('refus_reservation',
                f"Réservation salle refusée : {resa.organisation} — {resa.date:%d/%m/%Y} {resa.heure_debut:%H:%M}–{resa.heure_fin:%H:%M} ({resa.salle})",
                request=request, objet=resa)

        return redirect('administration:tableau_de_bord')

    return render(request, 'administration/valider_reservation_salle.html', {
        'reservation':      resa,
        'conflits':         conflits,
        'is_cabinet':       is_cabinet,
        'cabinets_dispo':   cabinets_dispo,
        'demandes_attente': demandes_attente,
        'blocages':         blocages,
        'indisponibilites': indisponibilites,
        'salles_alternatives': salles_alternatives,
    })


def _envoyer_email_decision(resa, action, commentaire_admin=''):
    validee = (action == 'valider')
    sujet = (
        f"[Kellermann] Votre demande du {resa.date:%d/%m/%Y} a été validée"
        if validee else
        f"[Kellermann] Votre demande du {resa.date:%d/%m/%Y} n'a pas pu être accordée"
    )
    corps = f"""Bonjour {resa.nom_demandeur},

{"Votre demande de réservation a été validée." if validee else "Votre demande de réservation n'a pas pu être acceptée."}

Détails :
  Temple    : {resa.temple}
  Date      : {resa.date:%d/%m/%Y}
  Horaires  : {resa.heure_debut:%H:%M} - {resa.heure_fin:%H:%M}
  Type      : {resa.get_type_reservation_display()}
"""
    if resa.besoin_agapes:
        corps += f"  Agapes    : {resa.nombre_repas} couverts\n"
    if commentaire_admin:
        corps += f"\nMessage de l'administrateur :\n{commentaire_admin}\n"
    corps += "\nFraternellement,\nL'administration des Temples Kellermann\n"

    try:
        destinataires = [resa.email_demandeur]
        # CC traiteur si agapes validée
        if resa.besoin_agapes and action == 'valider':
            email_t = get_email_traiteur()
            if email_t:
                destinataires.append(email_t)
        send_mail_kellermann(sujet, corps, destinataires, fail_silently=False)
    except Exception as e:
        print(f"Erreur email décision : {e}")


@staff_required
def valider_acces_portail(request, pk):
    demande = get_object_or_404(DemandeAccesPortail, pk=pk)

    if request.method == 'POST':
        action = request.POST.get('action')
        if action not in ('valider', 'refuser'):
            messages.error(request, "Action invalide.")
            return redirect('administration:tableau_de_bord')

        demande.statut = 'validee' if action == 'valider' else 'refusee'
        demande.save()

        if action == 'valider':
            lien = request.build_absolute_uri(f'/reservations/portail/{demande.token}/')
            send_mail_kellermann(
                subject="[Kellermann] Votre accès au portail loge a été validé",
                message=(
                    f"Bonjour {demande.nom_venerable},\n\n"
                    f"Votre demande d'accès au portail loge a été validée.\n\n"
                    f"Vous pouvez accéder à votre espace loge via le lien suivant :\n"
                    f"{lien}\n\n"
                    f"Ce lien est personnel — ne le partagez pas.\n\n"
                    f"Fraternellement,\nL'administration des Temples Kellermann"
                ),
                recipient_list=[demande.email],
            )
            messages.success(request, f"Accès validé pour {demande.nom_loge_display()} — lien envoyé à {demande.email}.")
        else:
            send_mail_kellermann(
                subject="[Kellermann] Votre demande d'accès portail",
                message=(
                    f"Bonjour {demande.nom_venerable},\n\n"
                    f"Votre demande d'accès au portail loge n'a pas pu être accordée.\n\n"
                    f"Pour toute question, contactez l'administration.\n\n"
                    f"Fraternellement,\nL'administration des Temples Kellermann"
                ),
                recipient_list=[demande.email],
            )
            messages.warning(request, f"Demande refusée pour {demande.nom_loge_display()}.")

        return redirect('administration:tableau_de_bord')

    return render(request, 'administration/valider_acces_portail.html', {'demande': demande})


def _envoyer_email_decision_salle(resa, action, commentaire_admin=''):
    validee = (action == 'valider')
    is_cabinet = resa.salle.type_salle == 'cabinet_reflexion'

    if is_cabinet:
        sujet = (
            f"[Kellermann] Votre demande de cabinet du {resa.date:%d/%m/%Y} a été validée"
            if validee else
            f"[Kellermann] Votre demande de cabinet du {resa.date:%d/%m/%Y} n'a pas pu être accordée"
        )
        if validee:
            corps = f"""Bonjour {resa.nom_demandeur},

Votre demande de cabinet de réflexion a été validée.

Cabinet attribué : {resa.salle.nom}
Date             : {resa.date:%d/%m/%Y}
Horaires         : {resa.heure_debut:%H:%M} - {resa.heure_fin:%H:%M}
Objet            : {resa.objet}
"""
        else:
            corps = f"""Bonjour {resa.nom_demandeur},

Votre demande de cabinet de réflexion du {resa.date:%d/%m/%Y} n'a pas pu être accordée.

Date     : {resa.date:%d/%m/%Y}
Horaires : {resa.heure_debut:%H:%M} - {resa.heure_fin:%H:%M}
Objet    : {resa.objet}
"""
    else:
        sujet = (
            f"[Kellermann] Votre demande de salle du {resa.date:%d/%m/%Y} a été validée"
            if validee else
            f"[Kellermann] Votre demande de salle du {resa.date:%d/%m/%Y} n'a pas pu être accordée"
        )
        corps = f"""Bonjour {resa.nom_demandeur},

{"Votre demande de réservation de salle a été validée." if validee else "Votre demande de réservation de salle n'a pas pu être acceptée."}

Détails :
  Salle     : {resa.salle}
  Date      : {resa.date:%d/%m/%Y}
  Horaires  : {resa.heure_debut:%H:%M} - {resa.heure_fin:%H:%M}
  Objet     : {resa.objet}
"""

    if commentaire_admin:
        corps += f"\nMessage de l'administrateur :\n{commentaire_admin}\n"
    corps += "\nFraternellement,\nL'administration des Temples Kellermann\n"

    try:
        send_mail_kellermann(sujet, corps, [resa.email_demandeur], fail_silently=False)
    except Exception as e:
        print(f"Erreur email decision salle : {e}")


# ── Règles de récurrence ──────────────────────────────────────────────────────

@staff_required
def regles_liste(request):
    regles = RegleRecurrence.objects.select_related('loge', 'loge__obedience', 'temple').order_by('temple__nom', 'jour_semaine', 'numero_semaine')
    if request.GET.get('temple'):
        regles = regles.filter(temple_id=request.GET['temple'])
    if request.GET.get('loge'):
        regles = regles.filter(loge_id=request.GET['loge'])
    return render(request, 'administration/regles_liste.html', {
        'regles': regles, 'temples': Temple.objects.all(),
        'loges': Loge.objects.filter(actif=True).order_by('nom'),
        'nb_regles': regles.count(),
    })


@staff_required
def regle_form(request, pk=None):
    regle = get_object_or_404(RegleRecurrence, pk=pk) if pk else None
    if request.method == 'POST':
        try:
            mois_actifs = [int(m) for m in request.POST.getlist('mois_actifs') if m.isdigit()]
            data = {
                'loge_id': request.POST['loge'], 'temple_id': request.POST['temple'],
                'jour_semaine': int(request.POST['jour_semaine']),
                'numero_semaine': int(request.POST['numero_semaine']),
                'heure_debut': request.POST['heure_debut'], 'heure_fin': request.POST['heure_fin'],
                'mois_actifs': mois_actifs,
                'actif': request.POST.get('actif') == 'on',
                'date_debut': request.POST.get('date_debut') or None,
                'date_fin': request.POST.get('date_fin') or None,
            }
            if regle:
                for k, v in data.items():
                    setattr(regle, k, v)
                regle.save()
                messages.success(request, "Règle modifiée.")
            else:
                RegleRecurrence.objects.create(**data)
                messages.success(request, "Règle ajoutée.")
            return redirect('administration:regles_liste')
        except Exception as e:
            messages.error(request, f"Erreur : {e}")

    return render(request, 'administration/regle_form.html', {
        'regle': regle, 'temples': Temple.objects.all(),
        'loges': Loge.objects.filter(actif=True).order_by('nom'),
        'jours': RegleRecurrence.JOUR_CHOICES,
        'semaines': RegleRecurrence.SEMAINE_CHOICES,
        'horaires': [
            ('09:00','09h00'),('09:30','09h30'),('10:00','10h00'),('10:30','10h30'),
            ('11:00','11h00'),('11:30','11h30'),('12:00','12h00'),
            ('14:00','14h00'),('14:30','14h30'),('15:00','15h00'),('15:30','15h30'),
            ('16:00','16h00'),('16:30','16h30'),('17:00','17h00'),
            ('19:00','19h00'),('19:30','19h30'),('20:00','20h00'),('20:30','20h30'),
            ('21:00','21h00'),('22:00','22h00'),('22:30','22h30'),('23:00','23h00'),
        ],
        'tranches': [
            ('matin',    'Matin',       '09:00', '12:00'),
            ('apmidi',   'Après-midi',  '14:00', '17:00'),
            ('soir',     'Soir',        '19:00', '22:30'),
            ('journee',  'Journée',     '09:00', '17:00'),
        ],
        'mois_choices': [
            (1,'Janvier'),(2,'Février'),(3,'Mars'),(4,'Avril'),
            (5,'Mai'),(6,'Juin'),(7,'Juillet'),(8,'Août'),
            (9,'Septembre'),(10,'Octobre'),(11,'Novembre'),(12,'Décembre'),
        ],
    })


@staff_required
def regle_supprimer(request, pk):
    regle = get_object_or_404(RegleRecurrence, pk=pk)
    if request.method == 'POST':
        nb = Reservation.objects.filter(regle_source=regle).count()
        regle.delete()
        messages.success(request, f"Règle supprimée ({nb} réservations conservées).")
        return redirect('administration:regles_liste')
    return render(request, 'administration/regle_supprimer.html', {'regle': regle})


# ── Regénération intelligente ─────────────────────────────────────────────────

@staff_required
def regenerer_intelligent(request):
    if request.method == 'POST':
        annee     = int(request.POST.get('annee', date.today().year))
        loge_id   = request.POST.get('loge') or None
        temple_id = request.POST.get('temple') or None
        mode      = request.POST.get('mode', 'ajouter')

        regles = RegleRecurrence.objects.filter(actif=True).select_related('loge', 'temple')
        if loge_id:
            regles = regles.filter(loge_id=loge_id)
        if temple_id:
            regles = regles.filter(temple_id=temple_id)

        # Saison maçonnique : 1er sept. (annee) → 30 juin (annee+1), hors juil/août
        d1, d2 = date(annee, 9, 1), date(annee + 1, 6, 30)

        cree = conflit = 0
        conflits_details = []
        for regle in regles:
            dates_saison = [
                d for d in (
                    _calculer_dates_regle(regle, annee) +
                    _calculer_dates_regle(regle, annee + 1)
                )
                if d1 <= d <= d2 and d.month not in [7, 8]
                and not (regle.date_fin and d > regle.date_fin)
                and not (regle.date_debut and d < regle.date_debut)
            ]

            if mode == 'remplacer':
                Reservation.objects.filter(
                    regle_source=regle,
                    date__gte=d1, date__lte=d2,
                ).delete()

            exclues = set(regle.dates_exclues or [])
            for d in dates_saison:
                if d.isoformat() in exclues:
                    continue
                if Reservation.objects.filter(
                    temple=regle.temple, date=d,
                    statut__in=['validee', 'attente'],
                    heure_debut__lt=regle.heure_fin,
                    heure_fin__gt=regle.heure_debut
                ).exclude(regle_source=regle).exists():
                    conflit += 1
                    conflits_details.append({
                        'loge_id': regle.loge_id,
                        'temple_id': regle.temple_id,
                        'date': d.isoformat(),
                        'hd': regle.heure_debut.strftime('%H:%M'),
                        'hf': regle.heure_fin.strftime('%H:%M'),
                    })
                    continue
                if not Reservation.objects.filter(regle_source=regle, date=d).exists():
                    Reservation.objects.create(
                        loge=regle.loge, temple=regle.temple, date=d,
                        heure_debut=regle.heure_debut, heure_fin=regle.heure_fin,
                        type_reservation='reguliere', statut='validee',
                        nom_demandeur='Generation automatique',
                        email_demandeur=regle.loge.email or settings.DEFAULT_FROM_EMAIL,
                        regle_source=regle,
                    )
                    cree += 1

        if conflits_details:
            request.session['regen_conflits'] = conflits_details
            messages.warning(request, f"{cree} tenues créées. {conflit} conflit(s) : "
                             "voir les suggestions de placement ci-dessous.")
            return redirect('administration:regenerer_conflits')
        messages.success(request, f"{cree} tenues créées pour la saison {annee}/{annee + 1}.")
        return redirect('administration:tableau_de_bord')

    return render(request, 'administration/regenerer.html', {
        'annees': [date.today().year, date.today().year + 1],
        'loges': Loge.objects.filter(actif=True).order_by('nom'),
        'temples': Temple.objects.all(),
    })


@staff_required
def regenerer_conflits(request):
    """Rapport des conflits de régénération : pour chaque tenue récurrente qui
    n'a pas pu être créée (temple occupé), propose les temples LIBRES du créneau
    et permet de placer la tenue en 1 clic (ou d'ignorer)."""
    conflits = request.session.get('regen_conflits', [])

    if request.method == 'POST':
        idx = int(request.POST.get('index', -1))
        action = request.POST.get('action')
        if 0 <= idx < len(conflits):
            item = conflits[idx]
            if action == 'placer':
                loge   = Loge.objects.filter(pk=item['loge_id']).first()
                temple = Temple.objects.filter(pk=request.POST.get('temple_choisi')).first()
                d  = date.fromisoformat(item['date'])
                hd = _to_time(item['hd']); hf = _to_time(item['hf'])
                if loge and temple and not _temple_occupe(temple, d, hd, hf):
                    Reservation.objects.create(
                        loge=loge, temple=temple, date=d,
                        heure_debut=hd, heure_fin=hf,
                        type_reservation='reguliere', statut='validee',
                        nom_demandeur='Placement (conflit régénération)',
                        email_demandeur=loge.email or settings.DEFAULT_FROM_EMAIL,
                        commentaire=(f"Placée sur {temple} le {d:%d/%m/%Y} car le temple "
                                     "d'origine était déjà occupé (retardataire)."),
                    )
                    messages.success(request, f"Tenue de {loge} placée sur {temple} le {d:%d/%m/%Y}.")
                    conflits.pop(idx)
                else:
                    messages.error(request, "Placement impossible : ce temple n'est plus libre.")
            elif action == 'ignorer':
                conflits.pop(idx)
        request.session['regen_conflits'] = conflits
        return redirect('administration:regenerer_conflits')

    lignes = []
    for i, item in enumerate(conflits):
        temple = Temple.objects.filter(pk=item['temple_id']).first()
        d  = date.fromisoformat(item['date'])
        hd = _to_time(item['hd']); hf = _to_time(item['hf'])
        occ = []
        for r in Reservation.objects.filter(
            temple=temple, date=d, heure_debut__lt=hf, heure_fin__gt=hd,
            statut__in=['validee', 'attente'],
        ).select_related('loge'):
            occ.append(str(r.loge or r.nom_organisation or r.nom_demandeur))
        libres = [t for t in Temple.objects.exclude(pk=temple.pk).order_by('nom')
                  if not _temple_occupe(t, d, hd, hf)]
        lignes.append({
            'index': i,
            'loge': Loge.objects.filter(pk=item['loge_id']).first(),
            'temple': temple, 'date': d, 'hd': item['hd'], 'hf': item['hf'],
            'occupant': ', '.join(occ), 'libres': libres,
        })
    return render(request, 'administration/regenerer_conflits.html', {'lignes': lignes})


def _occupants_recurrents(temple, date_r, hd, hf, exclure_pk=None):
    """Loges dont une tenue occupe déjà ce temple/créneau — pour proposer un
    échange bienveillant. Renvoie [{resa, loge, libres:[temples libres]}]."""
    hd_t = _to_time(hd); hf_t = _to_time(hf)
    qs = Reservation.objects.filter(
        temple=temple, date=date_r, heure_debut__lt=hf_t, heure_fin__gt=hd_t,
        statut__in=['validee', 'attente'], loge__isnull=False,
    ).select_related('loge')
    if exclure_pk:
        qs = qs.exclude(pk=exclure_pk)
    libres = [t for t in Temple.objects.exclude(pk=temple.pk).order_by('nom')
              if not _temple_occupe(t, date_r, hd_t, hf_t)]
    return [{'resa': r, 'loge': r.loge, 'libres': libres} for r in qs]


def _exclure_date_regle(resa):
    """Avant de déplacer une tenue issue d'une règle, on marque sa date d'origine
    comme exclue sur la règle, pour que la régénération ne la recrée pas."""
    regle = resa.regle_source
    if not regle:
        return
    iso = resa.date.isoformat()
    exclues = list(regle.dates_exclues or [])
    if iso not in exclues:
        exclues.append(iso)
        regle.dates_exclues = exclues
        regle.save(update_fields=['dates_exclues'])


@staff_required
def echanger_tenue(request):
    """Déplace exceptionnellement UNE tenue vers un temple libre (échange
    bienveillant), libérant ainsi le temple d'origine. La tenue est détachée de
    sa règle (regle_source=NULL) pour ne pas être écrasée à la régénération."""
    if request.method != 'POST':
        return redirect('administration:tableau_de_bord')
    resa   = get_object_or_404(Reservation, pk=request.POST.get('resa_id'))
    temple = get_object_or_404(Temple, pk=request.POST.get('temple_id'))
    nxt    = request.POST.get('next') or 'administration:tableau_de_bord'

    if _temple_occupe(temple, resa.date, resa.heure_debut, resa.heure_fin):
        messages.error(request, f"{temple} n'est plus libre sur ce créneau.")
    else:
        ancien = resa.temple
        note = (f"Déplacée de {ancien} vers {temple} le {date.today():%d/%m/%Y} "
                "(échange bienveillant).")
        _exclure_date_regle(resa)
        resa.temple = temple
        resa.regle_source = None
        resa.commentaire = (resa.commentaire + "\n" if resa.commentaire else "") + note
        resa.save()
        messages.success(request, f"Tenue de {resa.loge} déplacée sur {temple}. "
                         f"« {ancien} » est désormais libre le {resa.date:%d/%m/%Y}.")
        log_evenement('modification_reservation', f"{note} (loge {resa.loge})",
                      request=request, objet=resa)
    return redirect(nxt)


@staff_required
def deplacer_tenue(request):
    """Déplace exceptionnellement UNE tenue vers une autre DATE (échange de dates
    entre loges), avec contrôle du créneau. La tenue est détachée de sa règle
    (regle_source=NULL) pour ne pas être recréée à sa date d'origine lors d'une
    régénération."""
    if request.method != 'POST':
        return redirect('administration:tableau_de_bord')
    resa = get_object_or_404(Reservation, pk=request.POST.get('resa_id'))
    nxt  = request.POST.get('next') or 'administration:tableau_de_bord'
    try:
        nd = date.fromisoformat(request.POST.get('nouvelle_date', ''))
    except ValueError:
        messages.error(request, "Date invalide.")
        return redirect(nxt)

    # Temple cible : celui choisi (si fourni) sinon le temple actuel
    cible = Temple.objects.filter(pk=request.POST.get('nouveau_temple')).first() or resa.temple

    conflit = Reservation.objects.filter(
        temple=cible, date=nd,
        heure_debut__lt=resa.heure_fin, heure_fin__gt=resa.heure_debut,
        statut__in=['validee', 'attente'],
    ).exclude(pk=resa.pk).select_related('loge').first()

    if conflit and request.POST.get('forcer') != 'on':
        qui = conflit.loge or conflit.nom_organisation or conflit.nom_demandeur
        libres = [t for t in Temple.objects.exclude(pk=cible.pk).order_by('nom')
                  if not _temple_occupe(t, nd, resa.heure_debut, resa.heure_fin)]
        suff = (" Temples libres ce jour-là : " + ", ".join(str(t) for t in libres) + "."
                if libres else " Aucun autre temple n'est libre ce jour-là.")
        messages.warning(request, f"{cible} est déjà occupé le {nd:%d/%m/%Y} par {qui}."
                         + suff + " Choisissez un temple libre, ou cochez « Forcer ».")
        return redirect(nxt)

    ancienne = resa.date
    ancien_temple = resa.temple
    note = f"Déplacée du {ancienne:%d/%m/%Y} au {nd:%d/%m/%Y}"
    if cible.pk != ancien_temple.pk:
        note += f", et de « {ancien_temple} » vers « {cible} »"
    note += " (échange)."
    _exclure_date_regle(resa)
    resa.date = nd
    resa.temple = cible
    resa.regle_source = None
    resa.commentaire = (resa.commentaire + "\n" if resa.commentaire else "") + note
    resa.save()
    messages.success(request, f"Tenue de {resa.loge} déplacée au {nd:%d/%m/%Y} ({cible}).")
    log_evenement('modification_reservation', f"{note} (loge {resa.loge})",
                  request=request, objet=resa)
    return redirect(nxt)


@staff_required
def annuler_tenue(request):
    """Supprime UNE tenue (occurrence), en marquant sa date comme exclue sur sa
    règle pour qu'elle ne soit pas recréée à la régénération. Utile pour retirer
    un doublon ou une occurrence annulée exceptionnellement."""
    if request.method != 'POST':
        return redirect('administration:tableau_de_bord')
    resa = get_object_or_404(Reservation, pk=request.POST.get('resa_id'))
    nxt  = request.POST.get('next') or 'administration:tableau_de_bord'
    info = f"{resa.loge} — {resa.date:%d/%m/%Y} {resa.heure_debut:%H:%M} ({resa.temple})"
    _exclure_date_regle(resa)
    log_evenement('modification_reservation', f"Tenue annulée : {info}",
                  request=request, objet=resa)
    resa.delete()
    messages.success(request, f"Tenue annulée : {info}.")
    return redirect(nxt)


def _audit_capacite(annee):
    """Analyse de capacité. Deux vues :
    - HOMES récurrents du soir : (temple × jour lun-sam × position 1-4) → combien
      de créneaux récurrents restent libres ≈ combien de loges en plus.
    - OCCUPATION calendaire du soir sur la saison → taux d'occupation par temple
      et nombre de soirées libres (location ponctuelle « vendable »)."""
    from collections import defaultdict
    from datetime import time, timedelta

    JF = ['Lun', 'Mar', 'Mer', 'Jeu', 'Ven', 'Sam', 'Dim']
    JOURS = [0, 1, 2, 3, 4, 5]        # lundi → samedi (soirs)
    JOURS_LV = [0, 1, 2, 3, 4]        # lundi → vendredi
    POS = [1, 2, 3, 4]
    temples = list(Temple.objects.all().order_by('nom'))

    # ── Homes récurrents du soir ──────────────────────────────────────────────
    taken = defaultdict(set)          # temple_id → {(jour, position)}
    for r in RegleRecurrence.objects.filter(actif=True):
        if r.heure_debut.hour < 17 or r.jour_semaine not in JOURS:
            continue
        if r.numero_semaine == -1:
            positions = [4]
        elif r.numero_semaine == 0:   # règle "toutes les semaines"
            positions = [1, 2, 3, 4]
        elif r.numero_semaine in POS:
            positions = [r.numero_semaine]
        else:
            positions = []
        for p in positions:
            taken[r.temple_id].add((r.jour_semaine, p))

    par_temple = []
    occ_total = occ_lv = 0
    libres_par_jour = {JF[j]: 0 for j in JOURS}
    for t in temples:
        occ = len(taken[t.id])
        par_temple.append({'temple': str(t), 'total': len(JOURS) * len(POS),
                           'occ': occ, 'libres': len(JOURS) * len(POS) - occ})
        occ_total += occ
        occ_lv += len([1 for (j, p) in taken[t.id] if j in JOURS_LV])
        for j in JOURS:
            for p in POS:
                if (j, p) not in taken[t.id]:
                    libres_par_jour[JF[j]] += 1
    total_homes = len(temples) * len(JOURS) * len(POS)
    total_lv = len(temples) * len(JOURS_LV) * len(POS)

    # ── Occupation calendaire du soir sur la saison ──────────────────────────
    d1, d2 = date(annee, 9, 1), date(annee + 1, 6, 30)
    dates_elig = []
    d = d1
    while d <= d2:
        if d.weekday() in JOURS and d.month not in (7, 8):
            dates_elig.append(d)
        d += timedelta(days=1)
    occ_map = defaultdict(set)
    for tid, dd in Reservation.objects.filter(
        date__gte=d1, date__lte=d2, statut__in=['validee', 'attente'],
        heure_fin__gt=time(18, 0), heure_debut__lt=time(23, 0),
    ).values_list('temple_id', 'date'):
        occ_map[tid].add(dd)
    cal = []
    n = len(dates_elig)
    tot_occ = 0
    for t in temples:
        occ = sum(1 for dd in dates_elig if dd in occ_map[t.id])
        tot_occ += occ
        cal.append({'temple': str(t), 'dates': n, 'occ': occ, 'libres': n - occ,
                    'taux': round(100 * occ / n, 1) if n else 0})

    # ── Capacité WEEK-END matin & après-midi (samedi + dimanche) ─────────────
    BANDES = [('Matin', time(8, 0), time(13, 0)),
              ('Après-midi', time(13, 0), time(18, 0))]
    JOURS_WE = [5, 6]                 # samedi, dimanche
    JN = {5: 'Samedi', 6: 'Dimanche'}

    def _positions(num):
        if num == -1:
            return [4]
        if num == 0:
            return [1, 2, 3, 4]
        return [num] if num in POS else []

    def bande_of(h):
        if h < time(13, 0):
            return 'Matin'
        if h < time(18, 0):
            return 'Après-midi'
        return 'Soir'

    # homes week-end : (temple × jour_we × bande × position)
    we_taken = defaultdict(set)       # temple_id → {(jour, bande, pos)}
    for r in RegleRecurrence.objects.filter(actif=True, jour_semaine__in=JOURS_WE):
        b = bande_of(r.heure_debut)
        if b == 'Soir':
            continue
        for p in _positions(r.numero_semaine):
            we_taken[r.temple_id].add((r.jour_semaine, b, p))

    we_par_temple, we_occ = [], 0
    for t in temples:
        occ = len(we_taken[t.id])
        we_par_temple.append({'temple': str(t), 'total': len(JOURS_WE) * len(BANDES) * len(POS),
                              'occ': occ, 'libres': len(JOURS_WE) * len(BANDES) * len(POS) - occ})
        we_occ += occ
    we_par_case = []
    for wd in JOURS_WE:
        for bname, _, _ in BANDES:
            occ = sum(1 for t in temples for (j, b, p) in we_taken[t.id] if j == wd and b == bname)
            tot = len(temples) * len(POS)
            we_par_case.append({'jour': JN[wd], 'partie': bname, 'total': tot,
                                'occ': occ, 'libres': tot - occ})
    we_total = len(temples) * len(JOURS_WE) * len(BANDES) * len(POS)

    # occupation calendaire week-end matin/après-midi
    we_dates = {wd: [] for wd in JOURS_WE}
    dd = d1
    while dd <= d2:
        if dd.weekday() in JOURS_WE and dd.month not in (7, 8):
            we_dates[dd.weekday()].append(dd)
        dd += timedelta(days=1)
    occ_we = defaultdict(set)          # (temple_id, jour, bande) → {dates}
    for tid, ddate, hd_, hf_ in Reservation.objects.filter(
        date__gte=d1, date__lte=d2, statut__in=['validee', 'attente'],
    ).values_list('temple_id', 'date', 'heure_debut', 'heure_fin'):
        wd = ddate.weekday()
        if wd not in JOURS_WE or ddate.month in (7, 8):
            continue
        for bname, bs, be in BANDES:
            if hd_ < be and hf_ > bs:
                occ_we[(tid, wd, bname)].add(ddate)
    we_cal = []
    for wd in JOURS_WE:
        nd = len(we_dates[wd])
        for bname, _, _ in BANDES:
            occ = sum(len(occ_we[(t.id, wd, bname)]) for t in temples)
            slots = nd * len(temples)
            we_cal.append({'jour': JN[wd], 'partie': bname, 'dates': slots, 'occ': occ,
                           'libres': slots - occ,
                           'taux': round(100 * occ / slots, 1) if slots else 0})

    return {
        'annee': annee, 'nb_temples': len(temples),
        'homes': {'total': total_homes, 'occ': occ_total, 'libres': total_homes - occ_total,
                  'total_lv': total_lv, 'occ_lv': occ_lv, 'libres_lv': total_lv - occ_lv,
                  'par_temple': par_temple, 'libres_par_jour': libres_par_jour},
        'calendrier': {'par_temple': cal, 'dates_par_temple': n,
                       'total_dates': n * len(temples), 'total_occ': tot_occ,
                       'total_libres': n * len(temples) - tot_occ,
                       'taux': round(100 * tot_occ / (n * len(temples)), 1) if n and temples else 0},
        'weekend': {'homes': {'total': we_total, 'occ': we_occ, 'libres': we_total - we_occ,
                              'par_case': we_par_case, 'par_temple': we_par_temple},
                    'calendrier': {'par_case': we_cal}},
    }


def _info_resa(r, kind):
    if kind == 'temple':
        who = r.loge.abreviation if r.loge else (r.nom_organisation or r.nom_demandeur or '?')
        nom = r.loge.nom if r.loge else (r.nom_organisation or r.nom_demandeur or '')
        typ = r.get_type_reservation_display()
    else:
        who = r.loge.abreviation if r.loge else (r.organisation or r.nom_demandeur or '?')
        nom = r.loge.nom if r.loge else (r.organisation or '')
        typ = r.objet or 'Réservation'
    return {'pk': r.pk, 'who': who, 'nom': nom, 'hd': r.heure_debut, 'hf': r.heure_fin,
            'typ': typ, 'loge_id': r.loge_id}


def _scan_conflits():
    """Catapultages : deux structures DIFFÉRENTES sur le même lieu/créneau
    (temples et salles), à partir d'aujourd'hui."""
    from collections import defaultdict
    today = date.today()

    def sig(r, orgf):
        if r.loge_id:
            return ('L', r.loge_id)
        return ('O', (getattr(r, orgf, '') or r.nom_demandeur or '').strip().lower())

    confs = []

    def scan(qs, fkid, orgf, kind, lieu_attr):
        par = defaultdict(list)
        for r in qs:
            par[(getattr(r, fkid), r.date)].append(r)
        for (_, dd), items in par.items():
            items.sort(key=lambda x: x.heure_debut)
            for i in range(len(items)):
                for j in range(i + 1, len(items)):
                    a, b = items[i], items[j]
                    if (a.heure_debut < b.heure_fin and a.heure_fin > b.heure_debut
                            and sig(a, orgf) != sig(b, orgf)):
                        confs.append({'kind': kind, 'lieu': str(getattr(a, lieu_attr)),
                                      'date': dd, 'a': _info_resa(a, kind), 'b': _info_resa(b, kind)})

    scan(Reservation.objects.filter(statut__in=['validee', 'attente'], date__gte=today)
         .select_related('loge', 'temple'), 'temple_id', 'nom_organisation', 'temple', 'temple')
    scan(ReservationSalle.objects.filter(statut__in=['validee', 'attente'], date__gte=today)
         .select_related('loge', 'salle'), 'salle_id', 'organisation', 'salle', 'salle')
    confs.sort(key=lambda c: (c['date'], c['lieu']))
    return confs


@staff_required
def conflits(request):
    """Page de gestion des conflits / catapultages (lecture + annulation)."""
    if request.method == 'POST' and request.POST.get('action') == 'annuler':
        kind, pk = request.POST.get('kind'), request.POST.get('pk')
        if kind == 'temple':
            r = Reservation.objects.filter(pk=pk).first()
            if r:
                info = f"{r.temple} — {r.date:%d/%m/%Y}"
                _exclure_date_regle(r)
                r.delete()
                messages.success(request, f"Réservation temple annulée ({info}).")
        elif kind == 'salle':
            r = ReservationSalle.objects.filter(pk=pk).first()
            if r:
                info = f"{r.salle} — {r.date:%d/%m/%Y}"
                r.delete()
                messages.success(request, f"Réservation salle annulée ({info}).")
        return redirect('administration:conflits')
    return render(request, 'administration/conflits.html', {'conflits': _scan_conflits()})


def _occupation_temples(annee):
    """Occupation des temples par créneau (matin/après-midi/soir) sur la saison
    sept→juin, et capacité d'accueil (loge = 2 tenues/mois×10 = 20/an ;
    haut grade = 1/mois×10 = 10/an ; un créneau = 1 daypart × 1 temple × 1 jour)."""
    from datetime import time, timedelta
    from collections import defaultdict
    d1, d2 = date(annee, 9, 1), date(annee + 1, 6, 30)
    DAYPARTS = [('matin', time(8, 0), time(13, 0)),
                ('apres', time(13, 0), time(18, 0)),
                ('soir', time(18, 0), time(23, 59))]
    temples = list(Temple.objects.all().order_by('nom'))
    dates, d = [], d1
    while d <= d2:
        if d.month not in (7, 8):
            dates.append(d)
        d += timedelta(days=1)
    n = len(dates) or 1
    occ = defaultdict(set)
    for tid, dd, hd, hf in Reservation.objects.filter(
        date__gte=d1, date__lte=d2, statut__in=['validee', 'attente']
    ).values_list('temple_id', 'date', 'heure_debut', 'heure_fin'):
        if dd.month in (7, 8):
            continue
        for name, ds, de in DAYPARTS:
            if hd < de and hf > ds:
                occ[tid].add((dd, name))
    lignes, tot = [], 0
    for t in temples:
        o = len(occ[t.id]); tot += o
        p = {nm: sum(1 for (x, y) in occ[t.id] if y == nm) for nm, _, _ in DAYPARTS}
        lignes.append({
            'nom': str(t).replace('Temple ', ''),
            'occ_pct': round(100 * o / (n * 3)), 'libres': n * 3 - o,
            'matin_pct': round(100 * p['matin'] / n),
            'aprem_pct': round(100 * p['apres'] / n),
            'soir_pct': round(100 * p['soir'] / n),
        })
    grand = (len(temples) or 1) * n * 3
    libre = grand - tot
    soir_occ = sum(1 for tid in occ for (x, y) in occ[tid] if y == 'soir')
    soir_tot = (len(temples) or 1) * n
    soir_libre = soir_tot - soir_occ

    # ── Soir EN SEMAINE (lundi→vendredi) : le créneau réellement demandé par les loges
    dates_sem = {dd for dd in dates if dd.weekday() < 5}
    n_sem = len(dates_sem) or 1
    soir_sem_occ = sum(1 for tid in occ for (x, y) in occ[tid] if y == 'soir' and x in dates_sem)
    soir_sem_tot = (len(temples) or 1) * n_sem
    soir_sem_libre = soir_sem_tot - soir_sem_occ

    return {
        'annee': annee, 'nb_jours': len(dates), 'nb_jours_sem': len(dates_sem), 'par_temple_max': n * 3,
        'temples': lignes,
        'occ_global': round(100 * tot / grand), 'libres': libre,
        'libres_pct': round(100 * libre / grand),
        'soir_occ_pct': round(100 * soir_occ / soir_tot), 'soir_libres': soir_libre,
        'cap_tous_bleues': libre // 20, 'cap_tous_hg': libre // 10,
        'cap_soir_bleues': soir_libre // 20, 'cap_soir_hg': soir_libre // 10,
        # Soir en semaine (Lun-Ven) — capacité réaliste pour les loges
        'soir_sem_occ_pct': round(100 * soir_sem_occ / soir_sem_tot), 'soir_sem_libres': soir_sem_libre,
        'cap_sem_bleues': soir_sem_libre // 20, 'cap_sem_hg': soir_sem_libre // 10,
    }


def _creneaux_libres(annee, temple_id=None):
    """Catalogue des créneaux récurrents entièrement LIBRES sur la saison
    (temple × jour × semaine 1-4 × moment). Une loge bleue en combine 2 (mêmes
    ou différents temples/semaines), un haut grade en prend 1."""
    import calendar
    from datetime import time
    # (moment, début, fin, jours concernés) — tous les moments, tous les jours :
    # le soir en semaine reste le cas principal, mais la journée (matin/après-midi)
    # est proposable aussi, à la marge (personnes âgées, professions particulières).
    DP = [('soir', time(18, 0), time(23, 59), range(0, 7)),
          ('après-midi', time(13, 0), time(18, 0), range(0, 7)),
          ('matin', time(8, 0), time(13, 0), range(0, 7))]
    JF = ['lundi', 'mardi', 'mercredi', 'jeudi', 'vendredi', 'samedi', 'dimanche']
    POS = {1: '1re', 2: '2e', 3: '3e', 4: '4e'}
    MOIS = [(annee, m) for m in (9, 10, 11, 12)] + [(annee + 1, m) for m in (1, 2, 3, 4, 5, 6)]

    occ = set()
    for tid, dd, hd, hf in Reservation.objects.filter(
        date__gte=date(annee, 9, 1), date__lte=date(annee + 1, 6, 30),
        statut__in=['validee', 'attente'],
    ).values_list('temple_id', 'date', 'heure_debut', 'heure_fin'):
        for nm, ds, de, _ in DP:
            if hd < de and hf > ds:
                occ.add((tid, dd, nm))

    def dates_motif(wd, pos):
        out = []
        for an, mois in MOIS:
            js = [d for d in range(1, calendar.monthrange(an, mois)[1] + 1)
                  if date(an, mois, d).weekday() == wd]
            if pos <= len(js):
                out.append(date(an, mois, js[pos - 1]))
        return out

    temples = list(Temple.objects.filter(pk=temple_id) if temple_id else Temple.objects.all())
    charge = {t.id: sum(1 for (tid, d, nm) in occ if tid == t.id) for t in temples}
    temples.sort(key=lambda t: charge.get(t.id, 0))

    out = []
    for t in temples:
        nom = str(t).replace('Temple ', '')
        for nm, ds, de, wds in DP:
            for wd in wds:
                for pos in (1, 2, 3, 4):
                    dm = dates_motif(wd, pos)
                    if len(dm) >= 6 and all((t.id, d, nm) not in occ for d in dm):
                        out.append({'temple': nom, 'temple_id': t.id, 'jour': JF[wd],
                                    'jour_idx': wd, 'semaine': POS[pos], 'creneau': nm})
    return out


def _creneaux_fenetre(annee, deb, fin, jours):
    """Nombre de créneaux récurrents (temple × jour × position 1-4) où la fenêtre
    horaire [deb, fin[ est 100 % libre sur la saison — pour les jours donnés."""
    import calendar
    MOIS = [(annee, m) for m in (9, 10, 11, 12)] + [(annee + 1, m) for m in (1, 2, 3, 4, 5, 6)]
    occ = set()
    for tid, dd in Reservation.objects.filter(
        date__gte=date(annee, 9, 1), date__lte=date(annee + 1, 6, 30),
        statut__in=['validee', 'attente'], heure_debut__lt=fin, heure_fin__gt=deb,
    ).values_list('temple_id', 'date'):
        occ.add((tid, dd))

    def dates_motif(wd, pos):
        out = []
        for an, mois in MOIS:
            js = [d for d in range(1, calendar.monthrange(an, mois)[1] + 1)
                  if date(an, mois, d).weekday() == wd]
            if pos <= len(js):
                out.append(date(an, mois, js[pos - 1]))
        return out

    n = 0
    for t in Temple.objects.all():
        for wd in jours:
            for pos in (1, 2, 3, 4):
                dm = dates_motif(wd, pos)
                if len(dm) >= 6 and all((t.id, d) not in occ for d in dm):
                    n += 1
    return n


def _occupation_full(annee, temple_id=None, moment='', inclure_weekend=False):
    """Contexte complet occupation/capacité/potentiel d'une saison, partagé par la
    page et l'export PDF. Par défaut, les créneaux listés excluent le week-end
    (les loges se réunissent surtout en soirée du lundi au vendredi)."""
    tous_creneaux = _creneaux_libres(annee, temple_id)
    creneaux = tous_creneaux
    if moment in ('soir', 'après-midi', 'matin'):
        creneaux = [c for c in creneaux if c['creneau'] == moment]
    if not inclure_weekend:
        creneaux = [c for c in creneaux if c['jour_idx'] < 5]
    ctx = _occupation_temples(annee)

    # Capacité RÉELLE = créneaux récurrents 100 % libres le soir en semaine (cohérent
    # avec la liste des créneaux disponibles). Une loge bleue = 2 créneaux, un haut
    # grade = 1. Bien plus fiable que le décompte de soirées de calendrier isolées.
    all_slots = _creneaux_libres(annee)
    nb_slots = sum(1 for c in all_slots if c['creneau'] == 'soir' and c['jour_idx'] < 5)
    ctx['nb_slots_soir_sem'] = nb_slots
    ctx['cap_sem_bleues'] = nb_slots // 2
    ctx['cap_sem_hg'] = nb_slots
    # Maximum en créneaux récurrents, tous moments/jours confondus (même modèle)
    ctx['nb_slots_tous'] = len(all_slots)
    ctx['cap_tous_bleues'] = len(all_slots) // 2
    ctx['cap_tous_hg'] = len(all_slots)

    # Info : fenêtre spécifique 16h-19h en semaine (tenue courte de fin d'après-midi,
    # possible avant les tenues du soir de 20h) — capacité "cachée".
    from datetime import time as _t
    nb_16_19 = _creneaux_fenetre(annee, _t(16, 0), _t(19, 0), range(5))
    ctx['fin_aprem'] = {'slots': nb_16_19, 'cap_loges': nb_16_19 // 2, 'cap_hg': nb_16_19}
    params = Parametres.get_instance()
    MB_MIN, MB_MAX = 15, 20
    T_LOGE, T_HG = float(params.tarif_membre_loge), float(params.tarif_membre_hg)
    ctx['fin'] = {
        't_loge': T_LOGE, 't_hg': T_HG, 'mb_min': MB_MIN, 'mb_max': MB_MAX,
        'loge_min': MB_MIN * T_LOGE, 'loge_max': MB_MAX * T_LOGE,
        'hg_min': round(MB_MIN * T_HG, 2), 'hg_max': round(MB_MAX * T_HG, 2),
        'cap_loges': ctx['cap_sem_bleues'], 'cap_hg': ctx['cap_sem_hg'],
        'pot_loges_min': round(ctx['cap_sem_bleues'] * MB_MIN * T_LOGE),
        'pot_loges_max': round(ctx['cap_sem_bleues'] * MB_MAX * T_LOGE),
        'pot_hg_min': round(ctx['cap_sem_hg'] * MB_MIN * T_HG),
        'pot_hg_max': round(ctx['cap_sem_hg'] * MB_MAX * T_HG),
    }
    ctx['creneaux_libres'] = creneaux
    return ctx


@staff_required
def occupation(request):
    defaut = date.today().year if date.today().month >= 7 else date.today().year - 1
    try:
        annee = int(request.GET.get('annee', defaut))
    except (TypeError, ValueError):
        annee = defaut
    try:
        temple_id = int(request.GET.get('temple') or 0) or None
    except (TypeError, ValueError):
        temple_id = None
    moment = request.GET.get('moment', 'soir')  # défaut : le soir (cas principal)
    weekend = request.GET.get('weekend')  # '1' = inclure le week-end
    ctx = _occupation_full(annee, temple_id, moment, inclure_weekend=(weekend == '1'))
    ctx['annees'] = [defaut - 1, defaut, defaut + 1]
    ctx['tous_temples'] = Temple.objects.all().order_by('nom')
    ctx['temple_sel'] = temple_id
    ctx['moment_sel'] = moment
    ctx['weekend_sel'] = weekend
    return render(request, 'administration/occupation.html', ctx)


@staff_required
def occupation_export_pdf(request):
    """Export PDF « présentation » (bureau / AG) : constat d'occupation + capacité
    restante + potentiel financier + créneaux disponibles."""
    defaut = date.today().year if date.today().month >= 7 else date.today().year - 1
    try:
        annee = int(request.GET.get('annee', defaut))
    except (TypeError, ValueError):
        annee = defaut
    inclure_weekend = request.GET.get('weekend') == '1'
    ctx = _occupation_full(annee, moment='soir', inclure_weekend=inclure_weekend)
    log_evenement('export_occupation_pdf',
                  f"Export présentation occupation saison {annee}-{annee + 1}",
                  request=request, objet_type='systeme')
    return _occupation_pdf(ctx, annee, inclure_weekend)


def _occupation_pdf(ctx, annee, inclure_weekend=False):
    import io
    from django.utils import timezone
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.graphics.shapes import Drawing
    from reportlab.graphics.charts.barcharts import VerticalBarChart

    navy = colors.HexColor('#0F2137'); gold = colors.HexColor('#C8A84B')
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle('h1', parent=styles['Title'], textColor=navy, fontSize=17, alignment=0)
    h2 = ParagraphStyle('h2', parent=styles['Heading2'], textColor=navy, fontSize=12)
    body = ParagraphStyle('body', parent=styles['Normal'], fontSize=9.5, leading=13)
    small = ParagraphStyle('sm', parent=styles['Normal'], textColor=colors.HexColor('#64748B'), fontSize=9)

    def tstyle():
        return TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), navy), ('TEXTCOLOR', (0, 0), (-1, 0), gold),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'), ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 6), ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 4), ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ])

    temples = ctx['temples']; fin = ctx['fin']

    def chart_soir():
        dr = Drawing(460, 150)
        bc = VerticalBarChart()
        bc.x, bc.y, bc.width, bc.height = 30, 24, 420, 108
        bc.data = [[t['soir_pct'] for t in temples]]
        bc.categoryAxis.categoryNames = [t['nom'].replace('Temple ', '') for t in temples]
        bc.categoryAxis.labels.fontSize = 8
        bc.categoryAxis.labels.fillColor = colors.HexColor('#64748B')
        bc.valueAxis.valueMin = 0; bc.valueAxis.valueMax = 100
        bc.valueAxis.labels.fontSize = 8
        bc.valueAxis.labels.fillColor = colors.HexColor('#94A3B8')
        bc.bars[0].fillColor = gold; bc.bars[0].strokeColor = None
        bc.barWidth = 14
        dr.add(bc)
        return dr

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=1.4 * cm, bottomMargin=1.4 * cm,
                            leftMargin=1.5 * cm, rightMargin=1.5 * cm,
                            title=f"Occupation {annee}-{annee + 1}")
    E = [Paragraph(f"Occupation des temples et capacité — saison {annee}–{annee + 1}", h1),
         Paragraph(f"{ctx['nb_jours']} jours × 3 créneaux · édité le {timezone.localtime().strftime('%d/%m/%Y')}", small),
         Spacer(1, 0.35 * cm),
         Paragraph(f"Occupation globale : <b>{ctx['occ_global']} %</b> &nbsp;·&nbsp; occupation soir en semaine : "
                   f"<b>{ctx['soir_sem_occ_pct']} %</b> &nbsp;·&nbsp; créneaux récurrents soir en semaine libres (Lun-Ven) : "
                   f"<b>{ctx['nb_slots_soir_sem']}</b>", body),
         Spacer(1, 0.3 * cm),
         Paragraph("Taux d'occupation du soir par temple", h2), chart_soir(), Spacer(1, 0.25 * cm),
         Paragraph("Occupation par temple", h2), Spacer(1, 0.1 * cm)]

    rows = [['Temple', 'Global', 'Matin', 'Après-midi', 'Soir', 'Libres']]
    for t in temples:
        rows.append([t['nom'], f"{t['occ_pct']} %", f"{t['matin_pct']} %",
                     f"{t['aprem_pct']} %", f"{t['soir_pct']} %", t['libres']])
    rows.append(['GLOBAL', f"{ctx['occ_global']} %", '', '', f"{ctx['soir_occ_pct']} %", ctx['libres']])
    tt = Table(rows, hAlign='LEFT'); tt.setStyle(tstyle())
    E += [tt, Spacer(1, 0.4 * cm),
          Paragraph("Capacité d'accueil — le soir en semaine (Lun-Ven)", h2),
          Paragraph(f"Il reste <b>{ctx['nb_slots_soir_sem']}</b> créneaux récurrents 100 % libres le soir en semaine "
                    f"(hors vacances d'été). Une loge bleue en occupe 2, un haut grade 1 : on peut donc encore "
                    f"accueillir ~<b>{ctx['cap_sem_bleues']}</b> loges bleues <b>ou</b> ~<b>{ctx['cap_sem_hg']}</b> "
                    f"hauts grades. C'est la capacité réaliste, les loges se réunissant surtout du lundi au "
                    f"vendredi en soirée.", body),
          Spacer(1, 0.3 * cm),
          Paragraph("Potentiel financier annuel", h2),
          Paragraph(f"• En loges bleues : <b>{fin['pot_loges_min']} – {fin['pot_loges_max']} €/an</b> "
                    f"(~{fin['cap_loges']} loges × {fin['mb_min']}–{fin['mb_max']} membres × {fin['t_loge']} €).", body),
          Paragraph(f"• En hauts grades : <b>{fin['pot_hg_min']} – {fin['pot_hg_max']} €/an</b> "
                    f"(~{fin['cap_hg']} ateliers × {fin['mb_min']}–{fin['mb_max']} membres × {fin['t_hg']} €).", body),
          Paragraph("Une loge bleue rapporte ~2× plus par créneau qu'un haut grade. Tarifs paramétrables (Facturation).", small),
          Spacer(1, 0.35 * cm)]

    cl = ctx['creneaux_libres']
    _porte = "week-end inclus" if inclure_weekend else "soir en semaine, Lun-Ven"
    E += [Paragraph(f"Créneaux disponibles à proposer — {_porte} ({len(cl)})", h2), Spacer(1, 0.1 * cm)]
    crows = [['Temple', 'Jour', 'Semaine', 'Moment']]
    for c in cl[:40]:
        crows.append([c['temple'], c['jour'].capitalize(), f"{c['semaine']} sem.", c['creneau']])
    ct = Table(crows, repeatRows=1, hAlign='LEFT'); ct.setStyle(tstyle())
    E += [ct]
    if len(cl) > 40:
        E += [Spacer(1, 0.1 * cm), Paragraph(f"… et {len(cl) - 40} autres créneaux (voir la page en ligne).", small)]

    doc.build(E)
    buf.seek(0)
    resp = HttpResponse(buf.read(), content_type='application/pdf')
    resp['Content-Disposition'] = f'attachment; filename="occupation_kellermann_{annee}-{annee + 1}.pdf"'
    return resp


@staff_required
def statistiques(request):
    """Fusionnée dans la page unique Statistiques & reporting (exports:reporting)."""
    return redirect('exports:reporting')


def _statistiques_ancienne(request):
    """(Conservée pour référence, non routée.) Synthèse loges par type/statut/obédience."""
    from django.db.models import Count, Sum, Avg, Q
    actives = Loge.objects.exclude(statut='inactive')
    today = date.today()
    annee = today.year if today.month >= 7 else today.year - 1
    d1, d2 = date(annee, 9, 1), date(annee + 1, 6, 30)

    par_obd = list(actives.values('obedience__nom').annotate(
        n=Count('id'),
        loges=Count('id', filter=Q(type_loge='loge')),
        hg=Count('id', filter=Q(type_loge='haut_grade')),
        eff=Sum('effectif_total'),
    ).order_by('-n'))

    ctx = {
        'annee': annee,
        'nb_total': actives.count(),
        'nb_loges': actives.filter(type_loge='loge').count(),
        'nb_hg': actives.filter(type_loge='haut_grade').count(),
        'nb_active': actives.filter(statut='active').count(),
        'nb_reconf': actives.filter(statut='a_reconfirmer').count(),
        'nb_inactive': Loge.objects.filter(statut='inactive').count(),
        'effectif_total': actives.aggregate(s=Sum('effectif_total'))['s'] or 0,
        'agapes_moy': round(actives.filter(effectif_moyen_agapes__gt=0)
                            .aggregate(a=Avg('effectif_moyen_agapes'))['a'] or 0),
        'nb_avec_effectif': actives.filter(effectif_total__gt=0).count(),
        'par_obedience': par_obd,
        'nb_tenues': Reservation.objects.filter(date__gte=d1, date__lte=d2,
                                                statut='validee').count(),
        'nb_exc': Reservation.objects.filter(date__gte=d1, date__lte=d2, statut='validee',
                                             type_reservation='exceptionnelle').count(),
        'nb_regles': RegleRecurrence.objects.filter(actif=True).count(),
    }
    return render(request, 'administration/statistiques.html', ctx)


@staff_required
def relance_contacts(request):
    """Relance des loges qui n'ont pas confirmé la saison : envoi d'un rappel
    par email à celles qui en ont un, liste de celles à joindre autrement."""
    from django.db.models import Q

    if request.method == 'POST' and request.POST.get('action') == 'relancer':
        pks = [int(x) for x in request.POST.getlist('loges') if x.isdigit()]
        cibles = Loge.objects.filter(pk__in=pks).exclude(email='').exclude(email__isnull=True)
        n = 0
        for loge in cibles:
            send_mail_kellermann(
                subject="Rappel — Confirmation de votre calendrier / fiche de recensement",
                message=(
                    f"Bonjour,\n\n"
                    f"Sauf erreur, nous n'avons pas encore reçu la confirmation de votre "
                    f"calendrier de tenues (ou votre fiche de recensement) pour la nouvelle saison.\n\n"
                    f"Merci de nous revenir dès que possible afin de finaliser la réservation "
                    f"de vos temples et salles.\n\n"
                    f"Bien fraternellement,\nLes Temples Kellermann"
                ),
                recipient_list=[loge.email],
            )
            n += 1
        messages.success(request, f"{n} rappel(s) envoyé(s).")
        log_evenement('envoi_emails_saison', f"Relance contacts : {n} email(s)",
                      request=request, objet_type='systeme')
        return redirect('administration:relance_contacts')

    a_reconfirmer = Loge.objects.exclude(statut='inactive').filter(
        statut='a_reconfirmer').select_related('obedience').order_by('nom')
    avec_email = [l for l in a_reconfirmer if l.email]
    sans_email = Loge.objects.exclude(statut='inactive').filter(
        Q(email='') | Q(email__isnull=True)).select_related('obedience').order_by('nom')
    return render(request, 'administration/relance_contacts.html', {
        'avec_email': avec_email, 'sans_email': sans_email,
    })


@staff_required
def accueil(request):
    """Hub d'accueil : avancement de la saison, alertes, raccourcis."""
    from django.db.models import Q
    today = date.today()
    annee = today.year if today.month >= 7 else today.year - 1
    d1, d2 = date(annee, 9, 1), date(annee + 1, 6, 30)
    actives = Loge.objects.exclude(statut='inactive')
    nb_loges = actives.count()
    nb_tenues = Reservation.objects.filter(
        date__gte=d1, date__lte=d2, statut='validee', type_reservation='reguliere').count()
    vs_total = ValidationSaison.objects.filter(annee=annee).count()
    vs_soumises = ValidationSaison.objects.filter(
        annee=annee, statut__in=['soumise', 'traitee']).count()
    params = Parametres.get_instance()

    def etape(nom, ok, detail, encours=False):
        return {'nom': nom, 'ok': ok, 'encours': encours and not ok, 'detail': detail}

    etapes = [
        etape('Import', nb_loges > 0, f"{nb_loges} loges"),
        etape('Régénérer', nb_tenues > 0, f"{nb_tenues} tenues"),
        etape('Validation', vs_total > 0 and vs_soumises == vs_total,
              (f"{vs_soumises}/{vs_total}" if vs_total else "non ouverte"),
              encours=vs_total > 0),
        etape('Facturation', False,
              "active" if params.facturation_active else "désactivée"),
    ]
    occ = _occupation_temples(annee)
    alerts = {
        'conflits': len(_scan_conflits()),
        'a_reconfirmer': Loge.objects.filter(statut='a_reconfirmer').count(),
        'sans_email': actives.filter(Q(email='') | Q(email__isnull=True)).count(),
        'occ_soir_libre': 100 - occ['soir_occ_pct'],
    }
    return render(request, 'administration/accueil.html', {
        'annee': annee, 'nb_loges': nb_loges, 'etapes': etapes, 'alerts': alerts,
    })


@staff_required
def sante_donnees(request):
    """Tableau de bord de la qualité des données : conflits, tenues orphelines,
    doublons, loges sans contact/récurrence, à reconfirmer."""
    from django.db.models import Count, Q

    if request.method == 'POST' and request.POST.get('action') == 'purger_orphelines':
        n = Reservation.objects.filter(loge__isnull=True, type_reservation='reguliere').delete()[0]
        messages.success(request, f"{n} tenue(s) orpheline(s) supprimée(s).")
        return redirect('administration:sante_donnees')

    conflits = _scan_conflits()
    orph_temple = Reservation.objects.filter(loge__isnull=True, type_reservation='reguliere').count()
    orph_salle = ReservationSalle.objects.filter(loge__isnull=True).count()
    doublons_regles = (RegleRecurrence.objects
                       .values('loge', 'temple', 'jour_semaine', 'numero_semaine',
                               'heure_debut', 'heure_fin')
                       .annotate(n=Count('id')).filter(n__gt=1).count())
    actives = Loge.objects.exclude(statut='inactive')
    sans_email = actives.filter(Q(email='') | Q(email__isnull=True)).count()
    sans_regle = actives.annotate(
        nr=Count('regles', filter=Q(regles__actif=True))).filter(nr=0).count()
    a_reconfirmer = Loge.objects.filter(statut='a_reconfirmer').count()

    indicateurs = [
        {'label': 'Catapultages (conflits de créneau)', 'valeur': len(conflits),
         'url': 'administration:conflits', 'icone': '⚠️',
         'aide': 'Deux structures sur le même lieu/créneau.'},
        {'label': 'Tenues orphelines (sans loge)', 'valeur': orph_temple,
         'action': 'purger_orphelines', 'icone': '👻',
         'aide': 'Tenues récurrentes détachées de leur loge — à purger.'},
        {'label': 'Réservations de salle sans loge', 'valeur': orph_salle,
         'url': 'administration:rattachement_salles', 'icone': '🪑',
         'aide': 'Salles non rattachées — cliquez pour les relier à une loge.'},
        {'label': 'Doublons de règles de récurrence', 'valeur': doublons_regles,
         'url': 'administration:doublons_regles', 'icone': '🔁',
         'aide': 'Cliquez pour voir et supprimer les règles en double.'},
        {'label': 'Loges actives sans email', 'valeur': sans_email,
         'url': 'administration:annuaire', 'icone': '✉️',
         'aide': 'Impossible de les contacter / leur envoyer la validation.'},
        {'label': 'Loges actives sans récurrence', 'valeur': sans_regle,
         'url': 'administration:annuaire', 'icone': '📅',
         'aide': 'Aucune règle active — structures à clarifier.'},
        {'label': 'Loges à reconfirmer', 'valeur': a_reconfirmer,
         'url': 'administration:loges_saison', 'icone': '⏳',
         'aide': 'N’ont pas confirmé la saison.'},
    ]
    total_pb = sum(1 for i in indicateurs if i['valeur'])
    return render(request, 'administration/sante_donnees.html', {
        'indicateurs': indicateurs, 'total_pb': total_pb, 'conflits': conflits,
    })


@staff_required
def annuaire(request):
    """Annuaire admin : toutes les loges avec coordonnées (association, contact,
    email, téléphone) et créneaux, pour joindre facilement une structure."""
    from django.db.models import Q
    q       = request.GET.get('q', '').strip()
    f_type  = request.GET.get('type', '')
    f_statut = request.GET.get('statut', '')
    f_obd   = request.GET.get('obedience', '')

    loges = Loge.objects.select_related('obedience').order_by('nom')
    if q:
        loges = loges.filter(
            Q(nom__icontains=q) | Q(abreviation__icontains=q) | Q(association__icontains=q)
            | Q(nom_contact__icontains=q) | Q(email__icontains=q))
    if f_type in ('loge', 'haut_grade'):
        loges = loges.filter(type_loge=f_type)
    if f_statut in ('active', 'a_reconfirmer', 'inactive'):
        loges = loges.filter(statut=f_statut)
    if f_obd:
        loges = loges.filter(obedience__nom=f_obd)

    JF = ['Lun', 'Mar', 'Mer', 'Jeu', 'Ven', 'Sam', 'Dim']
    POS = {1: '1re', 2: '2e', 3: '3e', 4: '4e', -1: 'Der'}
    lignes = []
    nb = {'active': 0, 'a_reconfirmer': 0, 'inactive': 0, 'sans_email': 0}
    for l in loges:
        rec = [f"{POS.get(r.numero_semaine, r.numero_semaine)} {JF[r.jour_semaine]} "
               f"{str(r.temple).replace('Temple ', '')}"
               for r in l.regles.filter(actif=True).select_related('temple')]
        lignes.append({'loge': l, 'recurrences': ' · '.join(rec)})
        nb[l.statut] = nb.get(l.statut, 0) + 1
        if not l.email:
            nb['sans_email'] += 1

    return render(request, 'administration/annuaire.html', {
        'lignes': lignes, 'total': len(lignes), 'nb': nb,
        'q': q, 'f_type': f_type, 'f_statut': f_statut, 'f_obd': f_obd,
        'obediences': Obedience.objects.order_by('nom'),
    })


def _fusionner_loges(garder, suppr, appliquer):
    """Réaffecte toutes les données de `suppr` vers `garder`, puis supprime `suppr`.
    Renvoie un rapport {libellé: nombre}. Avec appliquer=False, ne fait qu'estimer
    (lecture seule). Logique alignée sur la commande `fusionner_loges`."""
    from django.db import transaction
    from temple_project.apps.reservations.models import ValidationSaison, RegleRecurrence
    report = {}

    def _collecte(apply_writes):
        # ValidationSaison est unique par (loge, année) : on retire les doublons d'abord
        annees = set(ValidationSaison.objects.filter(loge=garder).values_list('annee', flat=True))
        conflits = ValidationSaison.objects.filter(loge=suppr, annee__in=annees)
        if conflits.exists():
            report['Validations de saison en doublon (supprimées)'] = conflits.count()
            if apply_writes:
                conflits.delete()
        # Réaffectation générique de toutes les FK pointant vers Loge
        for rel in Loge._meta.related_objects:
            fname = rel.field.name
            Model = rel.related_model
            base = Model.objects.filter(**{fname: suppr})
            n = base.count()
            if n:
                report[f"{Model._meta.verbose_name} ({fname})"] = n
                if apply_writes:
                    base.update(**{fname: garder})

    if not appliquer:
        _collecte(False)
        return report

    with transaction.atomic():
        _collecte(True)
        # Complète les coordonnées manquantes de la loge conservée
        for champ in ('email', 'telephone', 'nom_contact', 'rite', 'rite_precision', 'association'):
            if not getattr(garder, champ, '') and getattr(suppr, champ, ''):
                setattr(garder, champ, getattr(suppr, champ))
        garder.save()
        suppr.delete()
        # Dédoublonnage des règles après réaffectation
        vues, doublons = set(), 0
        for r in RegleRecurrence.objects.filter(loge=garder).order_by('pk'):
            key = (r.temple_id, r.jour_semaine, r.numero_semaine, r.heure_debut, r.heure_fin)
            if key in vues:
                r.delete(); doublons += 1
            else:
                vues.add(key)
        if doublons:
            report['Règles en double supprimées'] = doublons
    return report


@staff_required
def fusion_loges(request):
    """Fusion de deux loges en double depuis l'admin (remplace la commande shell).
    Étape 1 : simulation (aperçu des objets réaffectés). Étape 2 : application."""
    loges  = Loge.objects.select_related('obedience').order_by('nom')
    garder = suppr = report = None

    if request.method == 'POST':
        try:
            gid = int(request.POST.get('garder') or 0)
            sid = int(request.POST.get('supprimer') or 0)
        except (ValueError, TypeError):
            gid = sid = 0
        garder = Loge.objects.filter(pk=gid).first()
        suppr  = Loge.objects.filter(pk=sid).first()
        action = request.POST.get('action')

        if not garder or not suppr:
            messages.error(request, "Sélectionnez deux loges valides.")
            garder = suppr = None
        elif garder.pk == suppr.pk:
            messages.error(request, "La loge à conserver et celle à supprimer doivent être différentes.")
            garder = suppr = None
        elif action == 'appliquer':
            nom_suppr = suppr.nom
            report = _fusionner_loges(garder, suppr, appliquer=True)
            log_evenement('fusion_loges',
                f"Fusion : « {nom_suppr} » absorbée dans « {garder.nom} » "
                f"({sum(report.values())} objet(s) réaffectés)",
                request=request, objet=garder)
            messages.success(request,
                f"Fusion effectuée : « {nom_suppr} » a été absorbée dans « {garder.nom} ».")
            return redirect('administration:fusion_loges')
        else:  # simuler
            report = _fusionner_loges(garder, suppr, appliquer=False)

    return render(request, 'administration/fusion_loges.html', {
        'loges': loges, 'garder': garder, 'suppr': suppr, 'report': report,
    })


@staff_required
def recherche_globale(request):
    """Recherche unique : loges, tenues et salles à venir, ou occupation d'un jour
    précis si la requête est une date (jj/mm ou jj/mm/aaaa)."""
    import re, datetime
    from django.db.models import Q

    q = request.GET.get('q', '').strip()
    loges = tenues = salles = []
    date_detectee = None

    if q:
        loges = (Loge.objects.filter(
                    Q(nom__icontains=q) | Q(abreviation__icontains=q) | Q(association__icontains=q)
                    | Q(nom_contact__icontains=q) | Q(email__icontains=q))
                 .select_related('obedience').order_by('nom')[:40])

        m = re.match(r'^(\d{1,2})[/-](\d{1,2})(?:[/-](\d{2,4}))?$', q)
        if m:
            d, mo, y = int(m.group(1)), int(m.group(2)), m.group(3)
            y = int(y) if y else datetime.date.today().year
            if y < 100:
                y += 2000
            try:
                date_detectee = datetime.date(y, mo, d)
            except ValueError:
                date_detectee = None

        today = datetime.date.today()
        if date_detectee:
            tenues = (Reservation.objects.filter(date=date_detectee, statut__in=['validee', 'attente'])
                      .select_related('loge', 'temple').order_by('heure_debut')[:60])
            salles = (ReservationSalle.objects.filter(date=date_detectee, statut__in=['validee', 'attente'])
                      .select_related('salle', 'loge').order_by('heure_debut')[:60])
        else:
            tenues = (Reservation.objects.filter(
                        Q(loge__nom__icontains=q) | Q(loge__abreviation__icontains=q)
                        | Q(temple__nom__icontains=q) | Q(nom_organisation__icontains=q),
                        date__gte=today, statut__in=['validee', 'attente'])
                      .select_related('loge', 'temple').order_by('date')[:40])
            salles = (ReservationSalle.objects.filter(
                        Q(organisation__icontains=q) | Q(loge__nom__icontains=q) | Q(salle__nom__icontains=q),
                        date__gte=today, statut__in=['validee', 'attente'])
                      .select_related('salle', 'loge').order_by('date')[:40])

    return render(request, 'administration/recherche.html', {
        'q': q, 'loges': loges, 'tenues': tenues, 'salles': salles,
        'date_detectee': date_detectee,
    })


@staff_required
def doublons_regles(request):
    """Liste les règles de récurrence en double (même loge/temple/jour/numéro de
    semaine) et permet de supprimer les surnuméraires en un clic (on garde la plus
    ancienne). Remplace la commande shell dedup_regles."""
    from django.db.models import Count

    def _groupes():
        # Vrai doublon = tout identique, HORAIRES COMPRIS. Deux règles le même
        # jour/semaine mais à des heures différentes sont deux tenues distinctes
        # (ex. une le matin, une l'après-midi) — pas des doublons.
        cles = (RegleRecurrence.objects
                .values('loge', 'temple', 'jour_semaine', 'numero_semaine',
                        'heure_debut', 'heure_fin')
                .annotate(n=Count('id')).filter(n__gt=1))
        out = []
        for c in cles:
            regles = list(RegleRecurrence.objects.filter(
                loge=c['loge'], temple=c['temple'],
                jour_semaine=c['jour_semaine'], numero_semaine=c['numero_semaine'],
                heure_debut=c['heure_debut'], heure_fin=c['heure_fin'],
            ).select_related('loge', 'temple').order_by('id'))
            out.append(regles)
        return out

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'tout':
            total = 0
            for regles in _groupes():
                garder = regles[0]
                for extra in regles[1:]:
                    extra.delete(); total += 1
            log_evenement('dedup_regles', f"{total} doublon(s) de règle supprimé(s) (dédup globale)",
                          request=request)
            messages.success(request, f"{total} doublon(s) de règle supprimé(s).")
        else:
            r = RegleRecurrence.objects.filter(pk=request.POST.get('supprimer') or 0).select_related('loge').first()
            if r:
                nom = str(r)
                r.delete()
                log_evenement('dedup_regles', f"Règle en double supprimée : {nom} (id={r.pk})",
                              request=request, objet=None)
                messages.success(request, "Règle en double supprimée.")
        return redirect('administration:doublons_regles')

    JOUR = dict(RegleRecurrence.JOUR_CHOICES)
    SEM  = dict(RegleRecurrence.SEMAINE_CHOICES)
    groupes = []
    for regles in _groupes():
        groupes.append({
            'loge': regles[0].loge, 'temple': regles[0].temple,
            'jour': JOUR.get(regles[0].jour_semaine), 'semaine': SEM.get(regles[0].numero_semaine),
            'garder': regles[0], 'extras': regles[1:],
        })
    return render(request, 'administration/doublons_regles.html', {'groupes': groupes})


@staff_required
def messagerie(request):
    """Boîte de réception des messages du formulaire de contact."""
    from temple_project.apps.reservations.models import MessageContact
    f = request.GET.get('statut', '')
    qs = MessageContact.objects.all()
    if f in ('nouveau', 'lu', 'traite'):
        qs = qs.filter(statut=f)
    counts = {
        'tous':    MessageContact.objects.count(),
        'nouveau': MessageContact.objects.filter(statut='nouveau').count(),
        'lu':      MessageContact.objects.filter(statut='lu').count(),
        'traite':  MessageContact.objects.filter(statut='traite').count(),
    }
    return render(request, 'administration/messagerie.html',
                  {'messages_list': list(qs[:300]), 'counts': counts, 'f': f})


@staff_required
def message_detail(request, pk):
    """Détail d'un message + réponse par email (enregistrée)."""
    from temple_project.apps.reservations.models import MessageContact
    m = get_object_or_404(MessageContact, pk=pk)

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'repondre':
            reponse = request.POST.get('reponse', '').strip()
            if not reponse:
                messages.error(request, "Le message de réponse est vide.")
            else:
                send_mail_kellermann(
                    subject=f"Re : {m.sujet or 'votre message'} — Temples Kellermann",
                    message=(
                        f"Bonjour {m.nom},\n\n{reponse}\n\n"
                        f"— — —\n"
                        f"En réponse à votre message du {m.created_at:%d/%m/%Y} :\n"
                        f"« {m.message} »"
                    ),
                    recipient_list=[m.email],
                )
                m.reponse = reponse
                m.date_reponse = timezone.now()
                m.repondu_par = request.user.get_username()
                m.statut = 'traite'
                m.save()
                log_evenement('reponse_message',
                              f"Réponse envoyée à {m.email} (message #{m.pk})",
                              request=request, objet=m)
                messages.success(request, f"Réponse envoyée à {m.email}.")
            return redirect('administration:message_detail', pk=m.pk)
        elif action == 'traite':
            m.statut = 'traite'; m.save(update_fields=['statut'])
            messages.success(request, "Message marqué comme traité.")
            return redirect('administration:messagerie')
        elif action == 'rouvrir':
            m.statut = 'lu'; m.save(update_fields=['statut'])
            return redirect('administration:message_detail', pk=m.pk)
        elif action == 'supprimer':
            m.delete()
            messages.success(request, "Message supprimé.")
            return redirect('administration:messagerie')

    # Marque comme lu à la première ouverture
    if m.statut == 'nouveau':
        m.statut = 'lu'
        m.save(update_fields=['statut'])
    return render(request, 'administration/message_detail.html', {'m': m})


@staff_required
def rattachement_salles(request):
    """Rattache manuellement à une loge les réservations de salle sans loge que
    l'auto-rattachement (par nom d'organisation) n'a pas su relier. En un clic,
    par organisation (toutes les réservations d'un même nom) ou à l'unité."""
    from collections import OrderedDict
    from temple_project.apps.reservations.signals import match_loge_par_organisation

    if request.method == 'POST':
        loge = Loge.objects.filter(pk=request.POST.get('loge') or 0).first()
        resa_id = request.POST.get('resa_id')
        org     = request.POST.get('organisation')
        if not loge:
            messages.error(request, "Sélectionnez une loge avant de rattacher.")
        elif resa_id:
            n = ReservationSalle.objects.filter(pk=resa_id, loge__isnull=True).update(loge=loge)
            if n:
                log_evenement('rattachement_salle',
                    f"Réservation salle #{resa_id} rattachée à « {loge.nom} »",
                    request=request, objet=loge)
                messages.success(request, f"Réservation rattachée à « {loge.nom} ».")
        elif org is not None:
            n = ReservationSalle.objects.filter(loge__isnull=True, organisation=org).update(loge=loge)
            log_evenement('rattachement_salle',
                f"{n} réservation(s) « {org} » rattachée(s) à « {loge.nom} »",
                request=request, objet=loge)
            messages.success(request, f"{n} réservation(s) « {org or '—'} » rattachée(s) à « {loge.nom} ».")
        return redirect('administration:rattachement_salles')

    orphelines = (ReservationSalle.objects.filter(loge__isnull=True)
                  .select_related('salle').order_by('organisation', 'date'))
    groupes, isolees = OrderedDict(), []
    for r in orphelines:
        key = (r.organisation or '').strip()
        if key:
            g = groupes.get(key)
            if not g:
                g = {'organisation': key, 'count': 0, 'exemples': [],
                     'suggestion': match_loge_par_organisation(key)}
                groupes[key] = g
            g['count'] += 1
            if len(g['exemples']) < 4:
                g['exemples'].append(r)
        else:
            isolees.append(r)

    return render(request, 'administration/rattachement_salles.html', {
        'groupes': list(groupes.values()), 'isolees': isolees,
        'loges': Loge.objects.order_by('nom'), 'total': orphelines.count(),
    })


@staff_required
def audit_export_excel(request):
    """Export Excel d'audit (à transmettre par mail) : tenues orphelines,
    doublons le même jour, loges sans contact, loges sans récurrence."""
    from collections import defaultdict
    from django.db.models import Q, Count
    from openpyxl.utils import get_column_letter

    JF = ['lundi', 'mardi', 'mercredi', 'jeudi', 'vendredi', 'samedi', 'dimanche']
    hf = Font(bold=True, color="FFFFFF")
    hfill = PatternFill("solid", fgColor="0F2137")
    ctr = Alignment(horizontal="center", vertical="center", wrap_text=True)

    wb = openpyxl.Workbook()

    def remplir(ws, headers, rows, widths):
        for c, h in enumerate(headers, 1):
            cell = ws.cell(1, c, h)
            cell.font = hf; cell.fill = hfill; cell.alignment = ctr
        for r, row in enumerate(rows, 2):
            for c, v in enumerate(row, 1):
                ws.cell(r, c, v)
        ws.freeze_panes = "A2"
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # 1) Tenues orphelines (sans loge)
    orph = Reservation.objects.filter(loge__isnull=True).select_related('temple').order_by('date', 'heure_debut')
    rows1 = [[
        t.date.strftime('%d/%m/%Y'), JF[t.date.weekday()],
        str(t.temple), t.heure_debut.strftime('%H:%M'), t.heure_fin.strftime('%H:%M'),
        t.get_type_reservation_display(), t.get_statut_display(),
        t.nom_demandeur or '', (t.commentaire or '')[:250],
    ] for t in orph]
    ws1 = wb.active; ws1.title = "Tenues orphelines"
    remplir(ws1, ["Date", "Jour", "Temple", "Début", "Fin", "Type", "Statut",
                  "Demandeur", "Commentaire"], rows1, [12, 11, 18, 8, 8, 16, 12, 22, 40])

    # 2) Doublons le même jour (tenues qui se chevauchent)
    parjour = defaultdict(list)
    for t in Reservation.objects.filter(
        loge__isnull=False, statut__in=['validee', 'attente']
    ).select_related('loge', 'temple'):
        parjour[(t.loge_id, t.date)].append(t)
    rows2 = []
    for (lid, d), items in parjour.items():
        if len(items) < 2:
            continue
        items.sort(key=lambda x: x.heure_debut)
        overlap = any(
            items[i].heure_debut < items[j].heure_fin and items[i].heure_fin > items[j].heure_debut
            for i in range(len(items)) for j in range(i + 1, len(items))
        )
        if overlap:
            lg = items[0].loge
            detail = ' ; '.join(f"{x.temple} {x.heure_debut:%H:%M}-{x.heure_fin:%H:%M}" for x in items)
            rows2.append([lg.abreviation, lg.nom, d.strftime('%d/%m/%Y'), JF[d.weekday()], len(items), detail])
    rows2.sort(key=lambda r: (r[0], r[2]))
    ws2 = wb.create_sheet("Doublons meme jour")
    remplir(ws2, ["Abrév.", "Loge", "Date", "Jour", "Nb tenues", "Détail (temple heure)"],
            rows2, [10, 34, 12, 11, 10, 55])

    actives = Loge.objects.exclude(statut='inactive')

    # 3) Loges sans contact
    sans_contact = actives.filter(
        Q(email='') | Q(email__isnull=True)
    ).filter(telephone='').filter(nom_contact='').select_related('obedience').order_by('nom')
    rows3 = [[l.abreviation, l.nom, l.get_statut_display(),
              l.obedience.nom if l.obedience else ''] for l in sans_contact]
    ws3 = wb.create_sheet("Loges sans contact")
    remplir(ws3, ["Abrév.", "Loge", "Statut", "Obédience"], rows3, [10, 34, 16, 16])

    # 4) Loges sans règle de récurrence
    sans_regle = actives.annotate(
        nr=Count('regles', filter=Q(regles__actif=True))).filter(nr=0).order_by('nom')
    rows4 = [[l.abreviation, l.nom, l.get_statut_display()] for l in sans_regle]
    ws4 = wb.create_sheet("Loges sans recurrence")
    remplir(ws4, ["Abrév.", "Loge", "Statut"], rows4, [10, 34, 16])

    # 5) Capacité / créneaux disponibles
    annee = date.today().year if date.today().month >= 9 else date.today().year - 1
    cap = _audit_capacite(annee)
    h = cap['homes']; cal = cap['calendrier']
    rows5 = [
        [f"Saison {annee}/{annee+1} — 1er sept. au 30 juin (hors juillet & août)", "", "", ""],
        ["HOMES RÉCURRENTS DU SOIR (temple x jour lun-sam x position 1-4)", "", "", ""],
        ["", "Total", "Occupés", "Libres"],
        ["Lun-Sam", h['total'], h['occ'], h['libres']],
        ["Lun-Ven (cœur)", h['total_lv'], h['occ_lv'], h['libres_lv']],
        ["", "", "", ""],
        ["Par temple (lun-sam)", "Total", "Occupés", "Libres"],
    ]
    for pt in h['par_temple']:
        rows5.append([pt['temple'], pt['total'], pt['occ'], pt['libres']])
    rows5.append(["", "", "", ""])
    rows5.append(["Créneaux-soir LIBRES par jour", "", "", ""])
    for j, v in h['libres_par_jour'].items():
        rows5.append([j, v, "", ""])
    rows5.append(["", "", "", ""])
    rows5.append([f"OCCUPATION CALENDAIRE DU SOIR ({cal['dates_par_temple']} soirées éligibles/temple)", "", "", ""])
    rows5.append(["Temple", "Soirées", "Occupées", "Libres"])
    for c in cal['par_temple']:
        rows5.append([c['temple'], c['dates'], c['occ'], f"{c['libres']} ({100-c['taux']:.0f}% libre)"])
    rows5.append(["GLOBAL", cal['total_dates'], cal['total_occ'], f"{cal['total_libres']} ({100-cal['taux']:.0f}% libre)"])
    we = cap['weekend']
    rows5.append(["", "", "", ""])
    rows5.append(["WEEK-END MATIN & APRÈS-MIDI (samedi + dimanche)", "", "", ""])
    rows5.append([f"Homes récurrents (temple x jour x partie x pos 1-4)", we['homes']['total'],
                  we['homes']['occ'], we['homes']['libres']])
    rows5.append(["Par case", "Total", "Occupés", "Libres"])
    for cs in we['homes']['par_case']:
        rows5.append([f"{cs['jour']} {cs['partie']}", cs['total'], cs['occ'], cs['libres']])
    rows5.append(["Occupation calendaire week-end", "Créneaux", "Occupés", "Libres"])
    for cs in we['calendrier']['par_case']:
        rows5.append([f"{cs['jour']} {cs['partie']}", cs['dates'], cs['occ'],
                      f"{cs['libres']} ({100-cs['taux']:.0f}% libre)"])
    ws5 = wb.create_sheet("Capacite creneaux")
    remplir(ws5, ["Indicateur", "Total", "Occupés", "Libres"], rows5, [52, 12, 12, 20])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Kellermann_Audit_{date.today():%Y%m%d}.xlsx"'
    wb.save(response)
    return response


# ── Import Excel ──────────────────────────────────────────────────────────────

@staff_required
def import_excel(request):
    errors = []
    stats  = None
    analyse = None
    if request.method == 'POST' and request.FILES.get('fichier'):
        try:
            wb = openpyxl.load_workbook(request.FILES['fichier'], data_only=True)
            if 'confirmer' in request.POST:
                stats, errors = _importer_donnees(wb)
                if not errors:
                    messages.success(request, (
                        f"Import réussi — Loges : {stats['loges']} nouvelle(s), {stats.get('loges_maj', 0)} mise(s) à jour · "
                        f"Règles : {stats['regles']} nouvelle(s), {stats.get('regles_maj', 0)} mise(s) à jour · "
                        f"Réservations ponctuelles : {stats.get('ponctuelles', 0)}."))
                    log_evenement('import_excel',
                        f"Import Excel : loges {stats['loges']}+{stats.get('loges_maj', 0)}maj, "
                        f"règles {stats['regles']}+{stats.get('regles_maj', 0)}maj, "
                        f"ponctuelles {stats.get('ponctuelles', 0)}",
                        request=request, objet_type='systeme')
                    return redirect('administration:tableau_de_bord')
            else:
                analyse = _analyser_import(wb)
        except Exception as e:
            errors.append(f"Erreur : {e}")
    return render(request, 'administration/import_excel.html', {'errors': errors, 'stats': stats, 'analyse': analyse})


# ── Template Excel ────────────────────────────────────────────────────────────

def _style_header(ws, row, cols, hf, hfill, ctr, thin):
    for col, h in enumerate(cols, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = hf; c.fill = hfill; c.alignment = ctr; c.border = thin

def _style_row(ws, row, vals, thin, ctr, fill=None):
    for ci, v in enumerate(vals, 1):
        c = ws.cell(row=row, column=ci, value=v)
        c.border = thin; c.alignment = ctr
        if fill: c.fill = fill

@staff_required
def telecharger_template_excel(request):
    """Template vierge avec exemples, listes déroulantes et onglet Référence."""
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = openpyxl.Workbook()
    hf    = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill("solid", fgColor="0F2137")
    ex    = PatternFill("solid", fgColor="EFF6FF")   # ligne exemple
    ctr   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin  = Border(left=Side(style='thin'), right=Side(style='thin'),
                   top=Side(style='thin'), bottom=Side(style='thin'))

    # ── Onglet RÉFÉRENCE ──────────────────────────────────────────────────────
    ws_ref = wb.active
    ws_ref.title = "RÉFÉRENCE"
    _style_header(ws_ref, 1, ["Temples","Obédiences","Types loge","Rites","Jours","N° semaine","Mois (n°)","Mois (nom)"], hf, hfill, ctr, thin)
    ref_data = [
        ("Lafayette",  "GODF",   "loge",       "reaa", "Lundi",    1, 1, "Janvier"),
        ("Égalité",    "GLdF",   "haut_grade", "rer",  "Mardi",    2, 2, "Février"),
        ("Fraternité", "GLNF",   "",           "rf",   "Mercredi", 3, 3, "Mars"),
        ("Liberté",    "GLAMF",  "",           "rem",  "Jeudi",    4, 4, "Avril"),
        ("",           "GODF-RF","",           "dh",   "Vendredi",-1, 5, "Mai"),
        ("",           "",       "",           "mem",  "Samedi",   "", 6, "Juin"),
        ("",           "",       "",           "autre","Dimanche", "", 7, "Juillet"),
        ("",           "",       "",           "",     "",         "", 8, "Août"),
        ("",           "",       "",           "",     "",         "", 9, "Septembre"),
        ("",           "",       "",           "",     "",         "",10, "Octobre"),
        ("",           "",       "",           "",     "",         "",11, "Novembre"),
        ("",           "",       "",           "",     "",         "",12, "Décembre"),
    ]
    for ri, row in enumerate(ref_data, 2):
        _style_row(ws_ref, ri, row, thin, ctr)
    ws_ref.cell(row=14, column=1, value="N° semaine : 1=1re, 2=2e, 3=3e, 4=4e, -1=Dernière")
    ws_ref.cell(row=15, column=4, value="Rites : reaa, rer, rf, rem, dh, mem, autre (laisser vide si inconnu)")
    for col, w in zip(['A','B','C','D','E','F','G','H'], [14,12,12,8,12,12,10,12]):
        ws_ref.column_dimensions[col].width = w
    ws_ref.freeze_panes = "A2"

    # ── Onglet LOGES ─────────────────────────────────────────────────────────
    ws_l = wb.create_sheet("LOGES")
    headers_l = ["Abréviation *","Nom complet *","Obédience *","Type *","Rite","Email","Effectif total","Moy. agapes","Nom du contact","Téléphone","Association"]
    _style_header(ws_l, 1, headers_l, hf, hfill, ctr, thin)
    # Lignes exemple
    _style_row(ws_l, 2, ["3P","Les 3 Piliers","GODF","loge","reaa","contact@loge.fr",45,30,"Jean Dupont","06 12 34 56 78","Les Amis des 3 Piliers"], thin, ctr, ex)
    _style_row(ws_l, 3, ["14GO","14/Consistoire GODF","GODF","haut_grade","rf","",20,0,"","","P12 - Consistoire"], thin, ctr, ex)
    # Validations
    dv_obe  = DataValidation(type="list", formula1="RÉFÉRENCE!$B$2:$B$6", allow_blank=True,  showDropDown=False)
    dv_type = DataValidation(type="list", formula1='"loge,haut_grade"',   allow_blank=False, showDropDown=False)
    dv_rite = DataValidation(type="list", formula1="RÉFÉRENCE!$D$2:$D$8", allow_blank=True,  showDropDown=False)
    ws_l.add_data_validation(dv_obe);  dv_obe.sqref  = "C2:C500"
    ws_l.add_data_validation(dv_type); dv_type.sqref = "D2:D500"
    ws_l.add_data_validation(dv_rite); dv_rite.sqref = "E2:E500"
    for col, w in zip(['A','B','C','D','E','F','G','H','I','J','K'], [12,38,12,12,8,28,14,12,22,16,26]):
        ws_l.column_dimensions[col].width = w
    ws_l.freeze_panes = "A2"
    ws_l.row_dimensions[1].height = 30

    # ── Onglet RÈGLES RÉCURRENCE ─────────────────────────────────────────────
    ws_r = wb.create_sheet("RÈGLES RÉCURRENCE")
    headers_r = ["Abréviation *","Nom complet *","Obédience *","Type *",
                 "Temple *","Jour *","N° semaine *","Heure début","Heure fin",
                 "Mois actifs (ex: 9,10,11,12,1,2,3,4,5,6)"]
    _style_header(ws_r, 1, headers_r, hf, hfill, ctr, thin)
    _style_row(ws_r, 2, ["3P","Les 3 Piliers","GODF","loge","Lafayette","Dimanche",2,"19:30","22:30","9,10,11,12,1,2,3,4,5,6"], thin, ctr, ex)
    _style_row(ws_r, 3, ["14GO","14/Consistoire","GODF","haut_grade","Égalité","Lundi",1,"14:00","17:00","5,6,9,10"], thin, ctr, ex)
    # Validations
    dv_t = DataValidation(type="list", formula1="RÉFÉRENCE!$A$2:$A$5", allow_blank=False, showDropDown=False)
    dv_j = DataValidation(type="list", formula1="RÉFÉRENCE!$D$2:$D$8", allow_blank=False, showDropDown=False)
    dv_s = DataValidation(type="list", formula1="RÉFÉRENCE!$E$2:$E$6", allow_blank=False, showDropDown=False)
    ws_r.add_data_validation(dv_t); dv_t.sqref = "E2:E500"
    ws_r.add_data_validation(dv_j); dv_j.sqref = "F2:F500"
    ws_r.add_data_validation(dv_s); dv_s.sqref = "G2:G500"
    for col, w in zip(range(1, 11), [12,38,12,12,14,12,12,12,12,38]):
        ws_r.column_dimensions[get_column_letter(col)].width = w
    ws_r.freeze_panes = "A2"
    ws_r.row_dimensions[1].height = 30

    # ── Onglet DATES PONCTUELLES ───────────────────────────────────────────────
    ws_p = wb.create_sheet("DATES PONCTUELLES")
    headers_p = ["Abréviation *","Nom complet *","Date * (AAAA-MM-JJ)","Heure début","Heure fin",
                 "Temple / Salle *","Type (temple/salle_reunion/cabinet_reflexion/banquet)","Objet / Remarque"]
    _style_header(ws_p, 1, headers_p, hf, hfill, ctr, thin)
    _style_row(ws_p, 2, ["3P","Les 3 Piliers","2026-11-14","09:00","17:00","Lafayette","temple","Tenue exceptionnelle"], thin, ctr, ex)
    _style_row(ws_p, 3, ["3P","Les 3 Piliers","2026-12-06","19:00","22:30","Oie et le Grill","banquet","Banquet d'ordre"], thin, ctr, ex)
    _style_row(ws_p, 4, ["14GO","14/Consistoire","2026-10-18","14:00","17:00","Salle James Anderson","salle_reunion","Réunion de bureau"], thin, ctr, ex)
    for col, w in zip(range(1, 9), [12,34,20,12,12,26,44,34]):
        ws_p.column_dimensions[get_column_letter(col)].width = w
    ws_p.freeze_panes = "A2"
    ws_p.row_dimensions[1].height = 30

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Kellermann_Import_Template.xlsx"'
    wb.save(response)
    return response


@staff_required
def telecharger_export_excel(request):
    """Export des données existantes (loges + règles) au même format que le template."""
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    hf    = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill("solid", fgColor="0F2137")
    alt   = PatternFill("solid", fgColor="F8FAFC")
    ctr   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin  = Border(left=Side(style='thin'), right=Side(style='thin'),
                   top=Side(style='thin'), bottom=Side(style='thin'))

    # ── Loges ────────────────────────────────────────────────────────────────
    ws_l = wb.active
    ws_l.title = "LOGES"
    headers_l = ["Abréviation","Nom complet","Obédience","Type","Rite","Email","Effectif total","Moy. agapes","Nom du contact","Téléphone","Association"]
    _style_header(ws_l, 1, headers_l, hf, hfill, ctr, thin)
    for ri, loge in enumerate(Loge.objects.select_related('obedience').order_by('nom'), 2):
        fill = None if ri % 2 == 0 else alt
        _style_row(ws_l, ri, [
            loge.abreviation, loge.nom,
            loge.obedience.nom if loge.obedience else "",
            loge.type_loge, loge.rite or "",
            loge.email or "",
            loge.effectif_total or "", loge.effectif_moyen_agapes or "",
            loge.nom_contact or "", loge.telephone or "", loge.association or "",
        ], thin, ctr, fill)
    for col, w in zip(['A','B','C','D','E','F','G','H','I','J','K'], [12,38,12,12,8,28,14,12,22,16,26]):
        ws_l.column_dimensions[col].width = w
    ws_l.freeze_panes = "A2"

    # ── Règles ───────────────────────────────────────────────────────────────
    ws_r = wb.create_sheet("RÈGLES RÉCURRENCE")
    headers_r = ["Abréviation","Nom complet","Obédience","Type",
                 "Temple","Jour","N° semaine","Heure début","Heure fin","Mois actifs"]
    _style_header(ws_r, 1, headers_r, hf, hfill, ctr, thin)
    JOURS = dict(RegleRecurrence.JOUR_CHOICES)
    for ri, reg in enumerate(RegleRecurrence.objects.select_related('loge','loge__obedience','temple').order_by('loge__nom'), 2):
        fill = None if ri % 2 == 0 else alt
        mois_str = ",".join(str(m) for m in reg.mois_actifs) if reg.mois_actifs else ""
        _style_row(ws_r, ri, [
            reg.loge.abreviation, reg.loge.nom,
            reg.loge.obedience.nom if reg.loge.obedience else "",
            reg.loge.type_loge,
            reg.temple.get_nom_display().replace("Temple ", ""),
            JOURS.get(reg.jour_semaine, ""),
            reg.numero_semaine,
            reg.heure_debut.strftime("%H:%M"),
            reg.heure_fin.strftime("%H:%M"),
            mois_str,
        ], thin, ctr, fill)
    for col, w in zip(range(1, 11), [12,38,12,12,14,12,12,12,12,30]):
        ws_r.column_dimensions[get_column_letter(col)].width = w
    ws_r.freeze_panes = "A2"

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Kellermann_Export_{date.today():%Y%m%d}.xlsx"'
    wb.save(response)
    return response


# ── Génération annuelle ───────────────────────────────────────────────────────

@staff_required
def generer_reservations_annuelles(request):
    if request.method == 'POST':
        annee = int(request.POST.get('annee', date.today().year + 1))
        regles = RegleRecurrence.objects.filter(actif=True).select_related('loge', 'temple')
        created = 0
        for regle in regles:
            for d in _calculer_dates_regle(regle, annee):
                if d.month in [7, 8]:
                    continue
                if not Reservation.objects.filter(loge=regle.loge, date=d, regle_source=regle).exists():
                    Reservation.objects.create(
                        loge=regle.loge, temple=regle.temple, date=d,
                        heure_debut=regle.heure_debut, heure_fin=regle.heure_fin,
                        type_reservation='reguliere', statut='validee',
                        nom_demandeur='Generation automatique',
                        email_demandeur=regle.loge.email or settings.DEFAULT_FROM_EMAIL,
                        regle_source=regle,
                    )
                    created += 1
        messages.success(request, f"{created} réservations générées pour {annee}.")
    return redirect('administration:tableau_de_bord')


# ── Reset / Nettoyage calendrier ─────────────────────────────────────────────

@staff_required
def reset_calendrier(request):
    today = date.today()
    annees = list(range(today.year - 2, today.year + 3))

    # Compteurs pour la prévisualisation
    def _compter(annee, loge_id, temple_id):
        qs_auto = Reservation.objects.filter(regle_source__isnull=False, date__year=annee)
        qs_tout = Reservation.objects.filter(date__year=annee)
        qs_regles = RegleRecurrence.objects.all()
        if loge_id:
            qs_auto  = qs_auto.filter(loge_id=loge_id)
            qs_tout  = qs_tout.filter(loge_id=loge_id)
            qs_regles = qs_regles.filter(loge_id=loge_id)
        if temple_id:
            qs_auto  = qs_auto.filter(temple_id=temple_id)
            qs_tout  = qs_tout.filter(temple_id=temple_id)
            qs_regles = qs_regles.filter(temple_id=temple_id)
        return {'auto': qs_auto.count(), 'tout': qs_tout.count(), 'regles': qs_regles.count()}

    if request.method == 'POST':
        if not request.POST.get('confirmer'):
            messages.error(request, "Cochez la case de confirmation pour valider.")
            return redirect('administration:reset_calendrier')

        action    = request.POST.get('action')
        annee     = request.POST.get('annee')
        loge_id   = request.POST.get('loge') or None
        temple_id = request.POST.get('temple') or None

        nb = 0
        if action == 'auto':
            # Supprimer réservations auto-générées pour l'année
            qs = Reservation.objects.filter(regle_source__isnull=False, date__year=int(annee))
            if loge_id:   qs = qs.filter(loge_id=loge_id)
            if temple_id: qs = qs.filter(temple_id=temple_id)
            nb, _ = qs.delete()
            messages.success(request, f"{nb} réservation(s) automatique(s) supprimée(s) pour {annee}.")

        elif action == 'tout':
            # Supprimer TOUTES les réservations temple pour l'année
            qs = Reservation.objects.filter(date__year=int(annee))
            if loge_id:   qs = qs.filter(loge_id=loge_id)
            if temple_id: qs = qs.filter(temple_id=temple_id)
            nb, _ = qs.delete()
            messages.warning(request, f"{nb} réservation(s) supprimée(s) pour {annee} (régulières + exceptionnelles).")

        elif action == 'regles':
            # Supprimer les règles de récurrence (+ réservations liées en cascade si souhaité)
            qs = RegleRecurrence.objects.all()
            if loge_id:   qs = qs.filter(loge_id=loge_id)
            if temple_id: qs = qs.filter(temple_id=temple_id)
            nb, _ = qs.delete()
            messages.warning(request, f"{nb} règle(s) de récurrence supprimée(s).")

        elif action == 'tout_absolu':
            # Tout supprimer : règles + réservations sans filtre année
            qs_r = Reservation.objects.all()
            qs_reg = RegleRecurrence.objects.all()
            if loge_id:
                qs_r   = qs_r.filter(loge_id=loge_id)
                qs_reg = qs_reg.filter(loge_id=loge_id)
            if temple_id:
                qs_r   = qs_r.filter(temple_id=temple_id)
                qs_reg = qs_reg.filter(temple_id=temple_id)
            nb_r, _ = qs_r.delete()
            nb_reg, _ = qs_reg.delete()
            messages.error(request, f"Nettoyage complet : {nb_reg} règle(s) et {nb_r} réservation(s) supprimée(s).")

        return redirect('administration:tableau_de_bord')

    # GET — afficher les compteurs
    annee_sel  = int(request.GET.get('annee', today.year))
    loge_id    = request.GET.get('loge') or None
    temple_id  = request.GET.get('temple') or None
    compteurs  = _compter(annee_sel, loge_id, temple_id)

    return render(request, 'administration/reset_calendrier.html', {
        'annees'   : annees,
        'annee_sel': annee_sel,
        'loges'    : Loge.objects.filter(actif=True).order_by('nom'),
        'temples'  : Temple.objects.all(),
        'loge_id'  : loge_id,
        'temple_id': temple_id,
        'compteurs': compteurs,
    })


# ── Gestion saison ────────────────────────────────────────────────────────────

_JOURS_FR = ['Lundi', 'Mardi', 'Mercredi', 'Jeudi', 'Vendredi', 'Samedi', 'Dimanche']


def _dates_saison(regle, annee):
    """
    Retourne toutes les dates de la règle dans la saison annee → annee+1
    (01/09/annee → 30/06/annee+1), en respectant date_debut/date_fin de la règle.
    """
    debut_saison = date(annee, 9, 1)
    fin_saison   = date(annee + 1, 6, 30)
    dates = []
    # Sep–Déc de l'année annee
    for d in _calculer_dates_regle(regle, annee):
        if d >= debut_saison:
            dates.append(d)
    # Jan–Jun de l'année annee+1
    for d in _calculer_dates_regle(regle, annee + 1):
        if d <= fin_saison:
            dates.append(d)
    # Filtre dates_debut/fin de la règle + dates exclues (déplacées/annulées)
    exclues = set(regle.dates_exclues or [])
    return [
        d for d in dates
        if d.isoformat() not in exclues
        and not (regle.date_fin and d > regle.date_fin)
        and not (regle.date_debut and d < regle.date_debut)
    ]


def _dry_run_saison(annee):
    """
    Simule la génération d'une saison sans aucune écriture en base.
    Couvre 01/09/annee → 30/06/annee+1.
    Retourne une liste de dicts triée par date :
      statut  'ok'          → sera créée (nouvelle)
              'existe_deja' → réservation auto déjà présente, sera remplacée
              'conflit'     → conflit avec une réservation manuelle, sera ignorée
    """
    regles = RegleRecurrence.objects.filter(actif=True).select_related('loge', 'temple')
    lignes = []
    for regle in regles:
        for d in _dates_saison(regle, annee):
            cle = f"{regle.pk}:{d.isoformat()}"

            # Conflit avec réservation NON-auto sur le même créneau ?
            conflit_qs = Reservation.objects.filter(
                temple=regle.temple,
                date=d,
                statut__in=['validee', 'attente'],
                heure_debut__lt=regle.heure_fin,
                heure_fin__gt=regle.heure_debut,
            ).exclude(regle_source=regle).select_related('loge')

            if conflit_qs.exists():
                c = conflit_qs.first()
                statut = 'conflit'
                conflict_detail = (
                    f"{c.loge.nom if c.loge else c.nom_demandeur} "
                    f"({c.get_statut_display()}, "
                    f"{c.heure_debut:%H:%M}–{c.heure_fin:%H:%M})"
                )
            elif Reservation.objects.filter(regle_source=regle, date=d).exists():
                statut = 'existe_deja'
                conflict_detail = ''
            else:
                statut = 'ok'
                conflict_detail = ''

            lignes.append({
                'regle_id':       regle.pk,
                'regle_label':    str(regle),
                'loge':           regle.loge,
                'temple':         regle.temple,
                'date':           d,
                'jour':           _JOURS_FR[d.weekday()],
                'heure_debut':    regle.heure_debut,
                'heure_fin':      regle.heure_fin,
                'statut':         statut,
                'conflict_detail': conflict_detail,
                'cle':            cle,
            })

    lignes.sort(key=lambda x: x['date'])
    return lignes


@staff_required
def gestion_saison(request):
    # Statistiques par saison
    current_year = date.today().year
    saisons = []
    for annee in range(current_year - 2, current_year + 3):
        saison_debut = date(annee, 9, 1)
        saison_fin = date(annee + 1, 6, 30)

        auto = Reservation.objects.filter(
            regle_source__isnull=False,
            date__gte=saison_debut,
            date__lte=saison_fin
        ).count()

        manuel = Reservation.objects.filter(
            regle_source__isnull=True,
            date__gte=saison_debut,
            date__lte=saison_fin
        ).count()

        saisons.append({
            'annee': annee,
            'auto': auto,
            'manuel': manuel,
            'total': auto + manuel
        })

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'supprimer_saison':
            annee = int(request.POST.get('annee'))
            saison_debut = date(annee, 9, 1)
            saison_fin = date(annee + 1, 6, 30)

            # Supprimer UNIQUEMENT les réservations automatiques
            deleted = Reservation.objects.filter(
                regle_source__isnull=False,
                date__gte=saison_debut,
                date__lte=saison_fin
            ).delete()[0]

            messages.success(request, f"Saison {annee}-{annee+1} : {deleted} réservations automatiques supprimées.")
            return redirect('administration:gestion_saison')

        elif action == 'generer_saison':
            annee = int(request.POST.get('annee_cible'))

            # Utiliser la fonction de régénération intelligente existante
            regles = RegleRecurrence.objects.filter(actif=True).select_related('loge', 'temple')
            cree = conflit = 0

            for regle in regles:
                for d in _calculer_dates_regle(regle, annee):
                    if d.month in [7, 8]:  # Éviter juillet-août
                        continue
                    if regle.date_fin and d > regle.date_fin:
                        continue
                    if regle.date_debut and d < regle.date_debut:
                        continue

                    # Supprimer les anciennes réservations automatiques pour cette règle/date
                    Reservation.objects.filter(regle_source=regle, date=d).delete()

                    # Vérifier conflits avec réservations manuelles
                    if Reservation.objects.filter(
                        temple=regle.temple, date=d, statut__in=['validee','attente'],
                        heure_debut__lt=regle.heure_fin, heure_fin__gt=regle.heure_debut
                    ).exclude(regle_source=regle).exists():
                        conflit += 1
                        continue

                    # Créer la nouvelle réservation
                    Reservation.objects.create(
                        loge=regle.loge, temple=regle.temple, date=d,
                        heure_debut=regle.heure_debut, heure_fin=regle.heure_fin,
                        type_reservation='reguliere', statut='validee',
                        nom_demandeur='Generation automatique',
                        email_demandeur=regle.loge.email or settings.DEFAULT_FROM_EMAIL,
                        regle_source=regle,
                    )
                    cree += 1

            if conflit:
                messages.warning(request, f"Saison {annee} : {cree} tenues créées, {conflit} conflits ignorés.")
            else:
                messages.success(request, f"Saison {annee} : {cree} tenues créées.")
            return redirect('administration:gestion_saison')

        elif action == 'reset_complet':
            confirmation = request.POST.get('confirmation', '').strip()
            if confirmation != 'CONFIRMER':
                messages.error(request, "Veuillez saisir 'CONFIRMER' pour valider la suppression.")
                return redirect('administration:gestion_saison')

            # Supprimer TOUTES les réservations automatiques
            deleted = Reservation.objects.filter(regle_source__isnull=False).delete()[0]
            messages.success(request, f"Reset complet : {deleted} réservations automatiques supprimées.")
            return redirect('administration:gestion_saison')

        elif action == 'backup':
            return telecharger_backup(request)

        elif action == 'previsualiser_saison':
            annee_cible = int(request.POST.get('annee_cible'))
            lignes_preview = _dry_run_saison(annee_cible)
            nb_ok      = sum(1 for l in lignes_preview if l['statut'] == 'ok')
            nb_existe  = sum(1 for l in lignes_preview if l['statut'] == 'existe_deja')
            nb_conflit = sum(1 for l in lignes_preview if l['statut'] == 'conflit')
            return render(request, 'administration/gestion_saison.html', {
                'saisons':         saisons,
                'current_year':    current_year,
                'annees':          list(range(current_year - 1, current_year + 4)),
                'db_last_modified': _get_db_last_modified(),
                'preview_lignes':  lignes_preview,
                'preview_annee':   annee_cible,
                'preview_nb_ok':   nb_ok,
                'preview_nb_existe': nb_existe,
                'preview_nb_conflit': nb_conflit,
            })

        elif action == 'generer_saison_confirme':
            annee_cible   = int(request.POST.get('annee_cible'))
            selectionnees = set(request.POST.getlist('lignes_selectionnees'))
            appliquer_retours = request.POST.get('appliquer_retours') == '1'
            regles = RegleRecurrence.objects.filter(actif=True).select_related('loge', 'temple')
            cree = conflit = ignore = ignore_retour = 0

            # Pré-charger les lignes de validation marquées 'annuler' pour cette saison
            # Clé : (regle_id, date_iso) → True si la loge a demandé l'annulation
            annulations_loge: set = set()
            if appliquer_retours:
                for ligne in ValidationSaisonLigne.objects.filter(
                    validation__annee=annee_cible,
                    validation__statut__in=['soumise', 'traitee'],
                    avis='annuler',
                ).select_related('regle'):
                    if ligne.regle_id:
                        annulations_loge.add((ligne.regle_id, ligne.date.isoformat()))

            for regle in regles:
                for d in _dates_saison(regle, annee_cible):
                    cle = f"{regle.pk}:{d.isoformat()}"
                    if cle not in selectionnees:
                        ignore += 1
                        continue

                    # Retour loge : annulation demandée → on skip
                    if appliquer_retours and (regle.pk, d.isoformat()) in annulations_loge:
                        ignore_retour += 1
                        continue

                    Reservation.objects.filter(regle_source=regle, date=d).delete()

                    if Reservation.objects.filter(
                        temple=regle.temple, date=d, statut__in=['validee', 'attente'],
                        heure_debut__lt=regle.heure_fin, heure_fin__gt=regle.heure_debut,
                    ).exclude(regle_source=regle).exists():
                        conflit += 1
                        continue

                    Reservation.objects.create(
                        loge=regle.loge, temple=regle.temple, date=d,
                        heure_debut=regle.heure_debut, heure_fin=regle.heure_fin,
                        type_reservation='reguliere', statut='validee',
                        nom_demandeur='Generation automatique',
                        email_demandeur=regle.loge.email or settings.DEFAULT_FROM_EMAIL,
                        regle_source=regle,
                    )
                    cree += 1

            parts = [f"{cree} tenue(s) créée(s)"]
            if ignore:
                parts.append(f"{ignore} ignorée(s) (décochées)")
            if ignore_retour:
                parts.append(f"{ignore_retour} annulée(s) sur demande loge")
            if conflit:
                parts.append(f"{conflit} conflit(s) détecté(s)")
            messages.success(request, f"Saison {annee_cible} : {', '.join(parts)}.")
            return redirect('administration:gestion_saison')

    return render(request, 'administration/gestion_saison.html', {
        'saisons':         saisons,
        'current_year':    current_year,
        'annees':          list(range(current_year - 1, current_year + 4)),
        'db_last_modified': _get_db_last_modified(),
    })


@staff_required
def preview_saison_excel(request):
    """Export Excel du dry-run groupé par loge."""
    from collections import defaultdict
    annee  = int(request.GET.get('annee', date.today().year))
    lignes = _dry_run_saison(annee)

    saison_label  = f"{annee}-{annee + 1}"
    periode_label = f"01/09/{annee} → 30/06/{annee + 1}"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Saison {saison_label}"

    # ── Styles ──────────────────────────────────────────────────────────────────
    navy_fill  = PatternFill("solid", fgColor="0F2137")
    loge_fill  = PatternFill("solid", fgColor="1E3A5F")
    col_fill   = PatternFill("solid", fgColor="E2E8F0")
    total_fill = PatternFill("solid", fgColor="F1F5F9")
    grand_fill = PatternFill("solid", fgColor="0F2137")
    fill_ok      = PatternFill("solid", fgColor="D1FAE5")
    fill_existe  = PatternFill("solid", fgColor="FEF9C3")
    fill_conflit = PatternFill("solid", fgColor="FEE2E2")

    thin  = Border(left=Side(style='thin'), right=Side(style='thin'),
                   top=Side(style='thin'),  bottom=Side(style='thin'))
    ctr   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left  = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    NCOLS = 7  # Date | Jour | Temple | Horaires | Règle | Statut | Détail
    col_widths = [13, 10, 22, 13, 34, 22, 36]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    def _cell(r, c, val='', font=None, fill=None, align=None, border=thin):
        cell = ws.cell(row=r, column=c, value=val)
        if font:   cell.font      = font
        if fill:   cell.fill      = fill
        if align:  cell.alignment = align
        if border: cell.border    = border
        return cell

    def _merge_row(r, val, font, fill, height=18):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOLS)
        _cell(r, 1, val, font=font, fill=fill, align=ctr)
        for c in range(2, NCOLS + 1):
            ws.cell(row=r, column=c).fill  = fill
            ws.cell(row=r, column=c).border = thin
        ws.row_dimensions[r].height = height

    # ── En-tête document ────────────────────────────────────────────────────────
    row = 1
    _merge_row(row, "TEMPLES KELLERMANN",
               Font(bold=True, size=14, color="C8A84B"), navy_fill, height=26)
    row += 1
    _merge_row(row, f"Prévisualisation saison {saison_label}  ·  {periode_label}",
               Font(bold=True, size=11, color="FFFFFF"), navy_fill, height=20)
    row += 1

    nb_ok      = sum(1 for l in lignes if l['statut'] == 'ok')
    nb_existe  = sum(1 for l in lignes if l['statut'] == 'existe_deja')
    nb_conflit = sum(1 for l in lignes if l['statut'] == 'conflit')
    _merge_row(row, f"✓ {nb_ok} à créer   ·   ↻ {nb_existe} remplace existantes   ·   ⚠ {nb_conflit} conflits ignorés",
               Font(size=9, color="0F2137"), PatternFill("solid", fgColor="F0F9FF"), height=16)
    row += 2  # ligne vide

    # ── Regroupement par loge ───────────────────────────────────────────────────
    groupes: dict = defaultdict(list)
    for l in lignes:
        groupes[l['loge'].nom if l['loge'] else '— Sans loge —'].append(l)
    groupes_tries = sorted(groupes.items(), key=lambda x: x[0])

    COL_HEADERS = ["Date", "Jour", "Temple", "Horaires", "Règle de récurrence", "Statut", "Détail conflit"]
    STATUT_LABELS = {'ok': '✓ À créer', 'existe_deja': '↻ Remplace', 'conflit': '⚠ Conflit'}

    for loge_nom, loge_lignes in groupes_tries:
        nb_loge = len(loge_lignes)

        # En-tête loge
        _merge_row(row, f"  {loge_nom.upper()}  —  {nb_loge} tenue{'s' if nb_loge > 1 else ''}",
                   Font(bold=True, size=10, color="FFFFFF"), loge_fill, height=18)
        row += 1

        # En-têtes colonnes
        for c, h in enumerate(COL_HEADERS, 1):
            _cell(row, c, h,
                  font=Font(bold=True, size=9, color="0F2137"),
                  fill=col_fill, align=ctr)
        ws.row_dimensions[row].height = 15
        row += 1

        # Lignes de données
        loge_lignes_sorted = sorted(loge_lignes, key=lambda x: x['date'])
        for l in loge_lignes_sorted:
            fill = {'ok': fill_ok, 'existe_deja': fill_existe, 'conflit': fill_conflit}[l['statut']]
            font_data = Font(size=9, color="991B1B" if l['statut'] == 'conflit' else "000000")
            vals = [
                l['date'].strftime('%d/%m/%Y'),
                l['jour'],
                str(l['temple']),
                f"{l['heure_debut']:%H:%M}–{l['heure_fin']:%H:%M}",
                l['regle_label'],
                STATUT_LABELS[l['statut']],
                l['conflict_detail'] or '',
            ]
            for c, v in enumerate(vals, 1):
                _cell(row, c, v, font=font_data, fill=fill,
                      align=ctr if c in (1, 2, 4, 6) else left)
            ws.row_dimensions[row].height = 14
            row += 1

        # Total loge
        nb_ok_l      = sum(1 for l in loge_lignes if l['statut'] == 'ok')
        nb_existe_l  = sum(1 for l in loge_lignes if l['statut'] == 'existe_deja')
        nb_conflit_l = sum(1 for l in loge_lignes if l['statut'] == 'conflit')
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOLS)
        _cell(row, 1,
              f"Total {loge_nom} : {nb_loge} tenue(s)   "
              f"[✓ {nb_ok_l}  ↻ {nb_existe_l}  ⚠ {nb_conflit_l}]",
              font=Font(bold=True, size=9, color="0F2137"),
              fill=total_fill, align=left)
        for c in range(2, NCOLS + 1):
            ws.cell(row=row, column=c).fill   = total_fill
            ws.cell(row=row, column=c).border = thin
        ws.row_dimensions[row].height = 14
        row += 2  # saut entre loges

    # ── Grand total ─────────────────────────────────────────────────────────────
    _merge_row(row,
               f"TOTAL SAISON {saison_label}  :  {len(lignes)} tenues   "
               f"[✓ {nb_ok} à créer   ↻ {nb_existe} remplace   ⚠ {nb_conflit} conflits]",
               Font(bold=True, size=11, color="C8A84B"), grand_fill, height=22)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = (
        f'attachment; filename="kellermann_saison_{saison_label}.xlsx"')
    wb.save(response)
    return response


@staff_required
def preview_saison_pdf(request):
    """Export PDF du dry-run groupé par loge."""
    from io import BytesIO
    from collections import defaultdict
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Spacer

    annee = int(request.GET.get('annee', date.today().year))
    lignes = _dry_run_saison(annee)

    saison_label  = f"{annee}-{annee + 1}"
    periode_label = f"01/09/{annee} \u2192 30/06/{annee + 1}"

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)

    navy        = colors.HexColor('#0F2137')
    gold        = colors.HexColor('#C8A84B')
    loge_color  = colors.HexColor('#1E3A5F')
    col_bg      = colors.HexColor('#E2E8F0')
    total_bg    = colors.HexColor('#DBEAFE')
    fill_ok     = colors.HexColor('#D1FAE5')
    fill_existe = colors.HexColor('#FEF9C3')
    fill_conflit= colors.HexColor('#FEE2E2')
    green_dark  = colors.HexColor('#166534')
    yellow_dark = colors.HexColor('#92400E')
    red_dark    = colors.HexColor('#991B1B')
    grid_color  = colors.HexColor('#CBD5E1')

    nb_ok      = sum(1 for l in lignes if l['statut'] == 'ok')
    nb_existe  = sum(1 for l in lignes if l['statut'] == 'existe_deja')
    nb_conflit = sum(1 for l in lignes if l['statut'] == 'conflit')

    STATUT_LABELS = {'ok': '\u2713 \u00c0 cr\u00e9er', 'existe_deja': '\u21bb Remplace', 'conflit': '\u26a0 Conflit'}
    COL_HEADERS   = ["Date", "Jour", "Temple", "Horaires", "R\u00e8gle de r\u00e9currence", "Statut", "D\u00e9tail conflit"]

    # Columns: Date | Jour | Temple | Horaires | Règle | Statut | Conflit
    col_widths = [2.2*cm, 1.5*cm, 6.2*cm, 2.4*cm, 7.2*cm, 2.5*cm, 4.2*cm]
    NCOLS      = len(col_widths)
    total_w    = sum(col_widths)

    story = []

    # ── En-tête document ────────────────────────────────────────────────────────
    header_data = [
        ["TEMPLES KELLERMANN"],
        [f"Pr\u00e9visualisation saison {saison_label}  \u00b7  {periode_label}"],
        [f"\u2713 {nb_ok} \u00e0 cr\u00e9er   \u00b7   \u21bb {nb_existe} remplace existantes   \u00b7   \u26a0 {nb_conflit} conflits ignor\u00e9s"],
    ]
    header_table = Table(header_data, colWidths=[total_w])
    header_table.setStyle(TableStyle([
        ('BACKGROUND',   (0,0), (-1,0), navy),
        ('TEXTCOLOR',    (0,0), (-1,0), gold),
        ('FONTNAME',     (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',     (0,0), (-1,0), 16),
        ('BACKGROUND',   (0,1), (-1,1), navy),
        ('TEXTCOLOR',    (0,1), (-1,1), colors.white),
        ('FONTNAME',     (0,1), (-1,1), 'Helvetica-Bold'),
        ('FONTSIZE',     (0,1), (-1,1), 11),
        ('BACKGROUND',   (0,2), (-1,2), colors.HexColor('#EFF6FF')),
        ('TEXTCOLOR',    (0,2), (-1,2), navy),
        ('FONTSIZE',     (0,2), (-1,2), 9),
        ('ALIGN',        (0,0), (-1,-1), 'CENTER'),
        ('VALIGN',       (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING',   (0,0), (-1,0), 6),
        ('BOTTOMPADDING',(0,0), (-1,0), 6),
        ('TOPPADDING',   (0,1), (-1,1), 4),
        ('BOTTOMPADDING',(0,1), (-1,1), 4),
        ('TOPPADDING',   (0,2), (-1,2), 3),
        ('BOTTOMPADDING',(0,2), (-1,2), 3),
    ]))
    story.append(header_table)
    story.append(Spacer(1, 0.3*cm))

    # ── Regroupement par loge ───────────────────────────────────────────────────
    groupes: dict = defaultdict(list)
    for l in lignes:
        groupes[l['loge'].nom if l['loge'] else '\u2014 Sans loge \u2014'].append(l)
    groupes_tries = sorted(groupes.items(), key=lambda x: x[0])

    for loge_nom, loge_lignes in groupes_tries:
        nb_loge      = len(loge_lignes)
        nb_ok_l      = sum(1 for l in loge_lignes if l['statut'] == 'ok')
        nb_existe_l  = sum(1 for l in loge_lignes if l['statut'] == 'existe_deja')
        nb_conflit_l = sum(1 for l in loge_lignes if l['statut'] == 'conflit')

        loge_lignes_sorted = sorted(loge_lignes, key=lambda x: x['date'])
        data       = []
        style_cmds = []

        # Ligne 0 — en-tête loge
        data.append([f"  {loge_nom.upper()}  \u2014  {nb_loge} tenue{'s' if nb_loge > 1 else ''}"] + [''] * (NCOLS - 1))
        style_cmds += [
            ('SPAN',         (0,0), (NCOLS-1, 0)),
            ('BACKGROUND',   (0,0), (-1,0), loge_color),
            ('TEXTCOLOR',    (0,0), (-1,0), colors.white),
            ('FONTNAME',     (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE',     (0,0), (-1,0), 10),
            ('ALIGN',        (0,0), (-1,0), 'LEFT'),
            ('VALIGN',       (0,0), (-1,0), 'MIDDLE'),
            ('TOPPADDING',   (0,0), (-1,0), 5),
            ('BOTTOMPADDING',(0,0), (-1,0), 5),
        ]

        # Ligne 1 — en-têtes colonnes
        data.append(COL_HEADERS)
        style_cmds += [
            ('BACKGROUND',   (0,1), (-1,1), col_bg),
            ('TEXTCOLOR',    (0,1), (-1,1), navy),
            ('FONTNAME',     (0,1), (-1,1), 'Helvetica-Bold'),
            ('FONTSIZE',     (0,1), (-1,1), 8),
            ('ALIGN',        (0,1), (-1,1), 'CENTER'),
            ('VALIGN',       (0,1), (-1,1), 'MIDDLE'),
            ('TOPPADDING',   (0,1), (-1,1), 3),
            ('BOTTOMPADDING',(0,1), (-1,1), 3),
        ]

        # Lignes de données
        for i, l in enumerate(loge_lignes_sorted):
            ri = 2 + i
            data.append([
                l['date'].strftime('%d/%m/%Y'),
                l['jour'][:3],
                str(l['temple']),
                f"{l['heure_debut']:%H:%M}\u2013{l['heure_fin']:%H:%M}",
                l['regle_label'],
                STATUT_LABELS[l['statut']],
                l['conflict_detail'] or '',
            ])
            if l['statut'] == 'ok':
                style_cmds.append(('BACKGROUND', (0,ri), (-1,ri), fill_ok))
                style_cmds.append(('TEXTCOLOR',  (5,ri), (5,ri),  green_dark))
            elif l['statut'] == 'existe_deja':
                style_cmds.append(('BACKGROUND', (0,ri), (-1,ri), fill_existe))
                style_cmds.append(('TEXTCOLOR',  (5,ri), (5,ri),  yellow_dark))
            else:
                style_cmds.append(('BACKGROUND', (0,ri), (-1,ri), fill_conflit))
                style_cmds.append(('TEXTCOLOR',  (0,ri), (-1,ri), red_dark))
                style_cmds.append(('FONTNAME',   (0,ri), (-1,ri), 'Helvetica-Bold'))

        # Ligne total loge
        ti = len(data)
        data.append(
            [f"Total {loge_nom} : {nb_loge} tenue(s)   "
             f"[\u2713 {nb_ok_l}  \u21bb {nb_existe_l}  \u26a0 {nb_conflit_l}]"]
            + [''] * (NCOLS - 1)
        )
        style_cmds += [
            ('SPAN',         (0,ti), (NCOLS-1, ti)),
            ('BACKGROUND',   (0,ti), (-1,ti), total_bg),
            ('TEXTCOLOR',    (0,ti), (-1,ti), navy),
            ('FONTNAME',     (0,ti), (-1,ti), 'Helvetica-Bold'),
            ('FONTSIZE',     (0,ti), (-1,ti), 8),
            ('ALIGN',        (0,ti), (-1,ti), 'LEFT'),
            ('TOPPADDING',   (0,ti), (-1,ti), 4),
            ('BOTTOMPADDING',(0,ti), (-1,ti), 4),
        ]

        # Style global données
        style_cmds += [
            ('FONTSIZE',     (0,2), (-1,ti-1), 7.5),
            ('VALIGN',       (0,2), (-1,ti-1), 'MIDDLE'),
            ('ALIGN',        (0,2), (-1,ti-1), 'LEFT'),
            ('ALIGN',        (0,2), (0,ti-1),  'CENTER'),
            ('ALIGN',        (1,2), (1,ti-1),  'CENTER'),
            ('ALIGN',        (3,2), (3,ti-1),  'CENTER'),
            ('ALIGN',        (5,2), (5,ti-1),  'CENTER'),
            ('TOPPADDING',   (0,2), (-1,ti-1), 2),
            ('BOTTOMPADDING',(0,2), (-1,ti-1), 2),
            ('GRID',         (0,1), (-1,-1), 0.3, grid_color),
        ]

        loge_table = Table(data, colWidths=col_widths, repeatRows=2)
        loge_table.setStyle(TableStyle(style_cmds))
        story.append(loge_table)
        story.append(Spacer(1, 0.2*cm))

    # ── Grand total ─────────────────────────────────────────────────────────────
    grand_data = [
        [f"TOTAL SAISON {saison_label}  :  {len(lignes)} tenues   "
         f"[\u2713 {nb_ok} \u00e0 cr\u00e9er   \u21bb {nb_existe} remplace   \u26a0 {nb_conflit} conflits]"]
        + [''] * (NCOLS - 1)
    ]
    grand_table = Table(grand_data, colWidths=col_widths)
    grand_table.setStyle(TableStyle([
        ('SPAN',         (0,0), (NCOLS-1, 0)),
        ('BACKGROUND',   (0,0), (-1,0), navy),
        ('TEXTCOLOR',    (0,0), (-1,0), gold),
        ('FONTNAME',     (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',     (0,0), (-1,0), 11),
        ('ALIGN',        (0,0), (-1,0), 'CENTER'),
        ('VALIGN',       (0,0), (-1,0), 'MIDDLE'),
        ('TOPPADDING',   (0,0), (-1,0), 8),
        ('BOTTOMPADDING',(0,0), (-1,0), 8),
    ]))
    story.append(grand_table)

    doc.build(story)
    buf.seek(0)
    response = HttpResponse(buf, content_type='application/pdf')
    response['Content-Disposition'] = (
        f'attachment; filename="kellermann_saison_{saison_label}.pdf"')
    return response


@staff_required
def validation_saison_admin(request):
    """Dashboard de validation de saison par les loges."""
    from django.utils import timezone

    current_year = date.today().year
    annees = list(range(current_year - 1, current_year + 4))
    annee = int(request.GET.get('annee', current_year))
    saison_label  = f"{annee}-{annee + 1}"
    periode_label = f"01/09/{annee} → 30/06/{annee + 1}"

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'ouvrir_validation':
            # ── ÉTAPE 1 : calcul + création des fiches, AUCUN email ──────────
            annee_cible = int(request.POST.get('annee_cible', annee))
            lignes_dry  = _dry_run_saison(annee_cible)

            from collections import defaultdict
            lignes_par_loge = defaultdict(list)
            for l in lignes_dry:
                if l['statut'] in ('ok', 'existe_deja') and l['loge']:
                    lignes_par_loge[l['loge']].append(l)

            nb_cree = nb_maj = nb_skip_soumise = 0
            for loge, loge_lignes in lignes_par_loge.items():
                val, created = ValidationSaison.objects.get_or_create(
                    loge=loge, annee=annee_cible,
                    defaults={'statut': 'attente'},
                )
                if not created and val.statut == 'soumise':
                    nb_skip_soumise += 1
                    continue

                # (Re)créer les lignes sans toucher au statut ni aux emails
                val.lignes.all().delete()
                ValidationSaisonLigne.objects.bulk_create([
                    ValidationSaisonLigne(
                        validation=val,
                        regle_id=l['regle_id'],
                        date=l['date'],
                        heure_debut=l['heure_debut'],
                        heure_fin=l['heure_fin'],
                        temple_nom=str(l['temple']),
                    )
                    for l in sorted(loge_lignes, key=lambda x: x['date'])
                ])

                # Statut attente uniquement si nouveau ou remis à zéro
                if created or val.statut not in ('ouverte', 'traitee'):
                    val.statut = 'attente'
                val.save()
                if created:
                    nb_cree += 1
                else:
                    nb_maj += 1

            parts = [f"{len(lignes_par_loge)} fiche(s) calculée(s)"]
            if nb_cree:
                parts.append(f"{nb_cree} nouvelle(s)")
            if nb_maj:
                parts.append(f"{nb_maj} mise(s) à jour")
            if nb_skip_soumise:
                parts.append(f"{nb_skip_soumise} déjà soumise(s) — non modifiée(s)")
            messages.info(request,
                "Récapitulatif calculé — aucun email envoyé. "
                "Vérifiez le tableau ci-dessous puis cliquez sur «\u00a0Envoyer les emails\u00a0». "
                f"({', '.join(parts)})")
            log_evenement('ouverture_validation_saison',
                f"Ouverture validation saison {annee_cible}-{annee_cible + 1} : {', '.join(parts)}",
                request=request, objet_type='systeme')
            return redirect(f"{request.path}?annee={annee_cible}")

        elif action == 'envoyer_emails':
            # ── ÉTAPE 2 : envoi des emails aux loges sélectionnées ───────────
            annee_cible  = int(request.POST.get('annee_cible', annee))
            periode_cible = f"01/09/{annee_cible} → 30/06/{annee_cible + 1}"
            pks_selectionnes = set(
                int(x) for x in request.POST.getlist('validation_pks') if x.isdigit()
            )
            if not pks_selectionnes:
                messages.warning(request, "Aucune loge sélectionnée.")
                return redirect(f"{request.path}?annee={annee_cible}")
            validations_attente = ValidationSaison.objects.filter(
                pk__in=pks_selectionnes, annee=annee_cible, statut='attente'
            ).select_related('loge')

            nb_email = nb_sans_email = nb_token_cree = 0
            for val in validations_attente:
                loge = val.loge
                nb_tenues = val.lignes.count()

                if loge.email:
                    demande = DemandeAccesPortail.objects.filter(
                        loge=loge, statut='validee'
                    ).order_by('-created_at').first()
                    if not demande:
                        # Pas d'accès portail → on le crée automatiquement pour
                        # que la loge puisse valider en ligne.
                        demande = DemandeAccesPortail.objects.create(
                            loge=loge, nom_venerable=loge.nom_contact or loge.nom,
                            email=loge.email, statut='validee',
                        )
                        nb_token_cree += 1
                    portail_url = request.build_absolute_uri(
                        f"/reservations/portail/{demande.token}/")
                    send_mail_kellermann(
                        subject=f"Validation de votre calendrier — Saison {annee_cible}-{annee_cible + 1}",
                        message=(
                            f"Bonjour,\n\n"
                            f"Nous vous invitons à valider le calendrier prévisionnel de vos tenues "
                            f"pour la saison {annee_cible}-{annee_cible + 1} ({periode_cible}).\n\n"
                            f"{nb_tenues} tenue(s) sont planifiées pour votre loge.\n\n"
                            f"Accédez à votre espace loge pour confirmer, signaler un déplacement "
                            f"ou une annulation :\n{portail_url}\n\n"
                            f"Bien fraternellement,\nLes Temples Kellermann"
                        ),
                        recipient_list=[loge.email],
                    )
                    val.statut     = 'ouverte'
                    val.date_envoi = timezone.now()
                    nb_email += 1
                else:
                    val.statut = 'ouverte'
                    nb_sans_email += 1

                val.save()

            parts = [f"{nb_email} email(s) envoyé(s)"]
            if nb_token_cree:
                parts.append(f"{nb_token_cree} accès portail créé(s)")
            if nb_sans_email:
                parts.append(f"{nb_sans_email} sans adresse email (non envoyé)")
            messages.success(request, "Emails envoyés — " + ", ".join(parts) + ".")
            log_evenement('envoi_emails_saison',
                f"Envoi emails validation saison {annee_cible}-{annee_cible + 1} : {', '.join(parts)}",
                request=request, objet_type='systeme')
            return redirect(f"{request.path}?annee={annee_cible}")

        elif action == 'marquer_traitee':
            pk = int(request.POST.get('validation_pk'))
            val = ValidationSaison.objects.get(pk=pk)
            val.statut = 'traitee'
            val.save()
            messages.success(request, f"{val.loge} — saison {val.annee}-{val.annee + 1} marquée comme traitée.")
            return redirect(f"{request.path}?annee={annee}")

        elif action == 'reinitialiser':
            pk = int(request.POST.get('validation_pk'))
            val = ValidationSaison.objects.get(pk=pk)
            val.statut = 'ouverte'
            val.commentaire_loge = ''
            val.date_reponse = None
            val.lignes.update(avis='ok', commentaire='')
            val.save()
            messages.success(request, f"Validation de {val.loge} réinitialisée.")
            return redirect(f"{request.path}?annee={annee}")

        elif action == 'relancer':
            annee_cible = int(request.POST.get('annee_cible', annee))
            periode_cible = f"01/09/{annee_cible} → 30/06/{annee_cible + 1}"
            pks_selectionnes = set(
                int(x) for x in request.POST.getlist('validation_pks') if x.isdigit()
            )
            if not pks_selectionnes:
                messages.warning(request, "Aucune loge sélectionnée.")
                return redirect(f"{request.path}?annee={annee_cible}")
            validations_ouverte = ValidationSaison.objects.filter(
                pk__in=pks_selectionnes, annee=annee_cible, statut='ouverte'
            ).select_related('loge')

            nb_email = nb_sans_email = 0
            for val in validations_ouverte:
                loge = val.loge
                nb_tenues = val.lignes.count()
                if loge.email:
                    demande = DemandeAccesPortail.objects.filter(
                        loge=loge, statut='validee'
                    ).order_by('-created_at').first()
                    if not demande:
                        demande = DemandeAccesPortail.objects.create(
                            loge=loge, nom_venerable=loge.nom_contact or loge.nom,
                            email=loge.email, statut='validee',
                        )
                    portail_url = request.build_absolute_uri(
                        f"/reservations/portail/{demande.token}/")
                    send_mail_kellermann(
                        subject=f"[RAPPEL] Validation de votre calendrier — Saison {annee_cible}-{annee_cible + 1}",
                        message=(
                            f"Bonjour,\n\n"
                            f"Nous n'avons pas encore reçu votre validation concernant le calendrier "
                            f"prévisionnel de vos tenues pour la saison {annee_cible}-{annee_cible + 1} "
                            f"({periode_cible}).\n\n"
                            f"{nb_tenues} tenue(s) sont planifiées pour votre loge.\n\n"
                            f"Accédez à votre espace loge pour confirmer, signaler un déplacement "
                            f"ou une annulation :\n{portail_url}\n\n"
                            f"Bien fraternellement,\nLes Temples Kellermann"
                        ),
                        recipient_list=[loge.email],
                    )
                    nb_email += 1
                else:
                    nb_sans_email += 1

            parts = [f"{nb_email} rappel(s) envoyé(s)"]
            if nb_sans_email:
                parts.append(f"{nb_sans_email} sans adresse email (non envoyé)")
            messages.success(request, "Relance effectuée — " + ", ".join(parts) + ".")
            log_evenement('relance_validation_saison',
                f"Relance validation saison {annee_cible}-{annee_cible + 1} : {', '.join(parts)}",
                request=request, objet_type='systeme')
            return redirect(f"{request.path}?annee={annee_cible}")

    # ── GET ──────────────────────────────────────────────────────────────────────
    validations = (
        ValidationSaison.objects
        .filter(annee=annee)
        .select_related('loge')
        .prefetch_related('lignes')
        .order_by('loge__nom')
    )

    # Statistiques globales
    nb_total   = validations.count()
    nb_attente = validations.filter(statut='attente').count()
    nb_ouverte = validations.filter(statut='ouverte').count()
    nb_soumise = validations.filter(statut='soumise').count()
    nb_traitee = validations.filter(statut='traitee').count()
    nb_anomalies_total = sum(v.nb_anomalies() for v in validations)

    # Loges avec au moins une tenue projetée pour cette année
    # mais sans fiche ValidationSaison — on réutilise le dry-run
    # uniquement si des fiches existent déjà (évite le calcul à froid)
    loges_validees = set(validations.values_list('loge_id', flat=True))
    if validations.exists():
        lignes_dry = _dry_run_saison(annee)
        loges_avec_tenues = {
            l['loge'].pk
            for l in lignes_dry
            if l['statut'] in ('ok', 'existe_deja') and l['loge']
        }
        loges_manquantes = Loge.objects.filter(
            pk__in=loges_avec_tenues - loges_validees
        ).order_by('nom')
    else:
        loges_manquantes = Loge.objects.none()

    validations_attente_list = [v for v in validations if v.statut == 'attente']
    validations_ouverte_list = [v for v in validations if v.statut == 'ouverte']

    return render(request, 'administration/validation_saison.html', {
        'annee':                   annee,
        'annees':                  annees,
        'saison_label':            saison_label,
        'periode_label':           periode_label,
        'validations':             validations,
        'validations_attente_list': validations_attente_list,
        'validations_ouverte_list': validations_ouverte_list,
        'nb_total':                nb_total,
        'nb_attente':              nb_attente,
        'nb_ouverte':              nb_ouverte,
        'nb_soumise':              nb_soumise,
        'nb_traitee':              nb_traitee,
        'nb_anomalies_total':      nb_anomalies_total,
        'loges_manquantes':        loges_manquantes,
    })


@staff_required
def modifier_reservation(request, pk):
    """Modifier le type et/ou le tarif d'une réservation déjà validée."""
    from decimal import Decimal, InvalidOperation
    resa = get_object_or_404(Reservation, pk=pk)
    params = Parametres.get_instance()

    if request.method == 'POST':
        nouveau_type = request.POST.get('type_reservation', '').strip()
        if nouveau_type not in ('reguliere', 'exceptionnelle', 'congres'):
            messages.error(request, "Type invalide.")
            return redirect('administration:modifier_reservation', pk=pk)

        ancien_type  = resa.get_type_reservation_display()
        ancien_tarif = resa.tarif
        resa.type_reservation = nouveau_type

        # Tarif : manuel si saisi, sinon calcul automatique
        tarif_manuel_str = request.POST.get('tarif_manuel', '').strip()
        if tarif_manuel_str:
            try:
                resa.tarif = Decimal(tarif_manuel_str.replace(',', '.'))
            except InvalidOperation:
                messages.error(request, "Montant invalide.")
                return redirect('administration:modifier_reservation', pk=pk)
        else:
            resa.tarif = tarif_reservation(resa, params)

        resa.save()

        log_evenement('modification_reservation',
            f"Type : {ancien_type} → {resa.get_type_reservation_display()} | "
            f"Tarif : {ancien_tarif} € → {resa.tarif} € | "
            f"{resa.loge} — {resa.date:%d/%m/%Y} ({resa.temple})",
            request=request, objet=resa)

        messages.success(request,
            f"Réservation mise à jour — type : {resa.get_type_reservation_display()}, "
            f"tarif : {resa.tarif} €.")
        next_url = request.POST.get('next', '')
        return redirect(next_url or 'administration:tableau_de_bord')

    # Prévisualisation tarif automatique selon le type
    from copy import copy
    def _tarif_auto(t):
        sim = copy(resa)
        sim.type_reservation = t
        sim.regle_source_id = None
        return tarif_reservation(sim, params)

    return render(request, 'administration/modifier_reservation.html', {
        'reservation':             resa,
        'tarif_auto_reguliere':    _tarif_auto('reguliere'),
        'tarif_auto_exceptionnelle': _tarif_auto('exceptionnelle'),
        'params':                  params,
    })


@staff_required
def telecharger_backup(request):
    """Télécharge la base de données SQLite en tant que sauvegarde."""
    import os
    from django.conf import settings

    # Chemin vers la base de données
    db_path = settings.DATABASES['default']['NAME']

    # Vérifier que le fichier existe
    if not os.path.exists(db_path):
        messages.error(request, "Fichier de base de données introuvable.")
        return redirect('administration:gestion_saison')

    # Nom du fichier de téléchargement
    today = date.today().strftime('%Y%m%d')
    filename = f'backup_kellermann_{today}.sqlite3'

    # Lire le fichier et le retourner en réponse
    log_evenement('backup_base',
        f"Téléchargement backup base de données : {filename}",
        request=request, objet_type='systeme')
    with open(db_path, 'rb') as f:
        response = HttpResponse(f.read(), content_type='application/octet-stream')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


@staff_required
def restaurer_backup(request):
    """Permet de restaurer une sauvegarde de la base de données."""
    import os
    import sqlite3
    from django.conf import settings

    if request.method == 'POST' and request.FILES.get('backup_file'):
        backup_file = request.FILES['backup_file']

        # Vérifier l'extension du fichier
        if not backup_file.name.endswith('.sqlite3'):
            messages.error(request, "Le fichier doit avoir l'extension .sqlite3")
            return redirect('administration:restaurer_backup')

        # Vérifier que c'est bien une base SQLite valide
        try:
            # Tester la connexion à la base uploadée
            import tempfile
            with tempfile.NamedTemporaryFile(delete=False) as temp_file:
                for chunk in backup_file.chunks():
                    temp_file.write(chunk)
                temp_path = temp_file.name

            # Tester la validité du fichier SQLite
            conn = sqlite3.connect(temp_path)
            conn.execute("SELECT name FROM sqlite_master WHERE type='table';")
            conn.close()

        except Exception as e:
            os.unlink(temp_path)
            messages.error(request, f"Fichier SQLite invalide : {e}")
            return redirect('administration:restaurer_backup')

        # Chemin de la base actuelle
        db_path = settings.DATABASES['default']['NAME']

        # Créer une sauvegarde automatique de la base actuelle
        backup_path = f"{db_path}.avant_restauration"
        try:
            import shutil
            shutil.copy2(db_path, backup_path)
        except Exception as e:
            messages.warning(request, f"Impossible de créer la sauvegarde automatique : {e}")

        # Remplacer la base actuelle par le fichier uploadé
        try:
            shutil.move(temp_path, db_path)
            messages.success(request, "Sauvegarde restaurée avec succès. Une copie de l'ancienne base a été sauvegardée.")
            return redirect('administration:gestion_saison')
        except Exception as e:
            messages.error(request, f"Erreur lors de la restauration : {e}")
            return redirect('administration:restaurer_backup')

    return render(request, 'administration/restaurer_backup.html')


# ── Helpers ───────────────────────────────────────────────────────────────────

def _get_db_last_modified():
    """Retourne la date de dernière modification de la base de données."""
    import os
    from django.conf import settings

    db_path = settings.DATABASES['default']['NAME']
    if os.path.exists(db_path):
        timestamp = os.path.getmtime(db_path)
        return date.fromtimestamp(timestamp).strftime('%d/%m/%Y %H:%M')
    return "Inconnue"


def _preview_excel(wb):
    preview = {}
    for sheet_name in wb.sheetnames[:4]:
        ws = wb[sheet_name]
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i > 6: break
            if any(v is not None for v in row):
                rows.append([str(v)[:40] if v is not None else '' for v in row[:10]])
        preview[sheet_name] = rows
    return preview


def _data_rows(ws):
    """Itère (numéro_ligne, ligne) en sautant une éventuelle ligne de titre :
    l'en-tête est la 1re ligne (parmi les 6 premières) dont la 1re cellule
    contient « abr » (Abréviation). Les données commencent juste après."""
    header = 1
    for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=6, values_only=True), 1):
        first = str(row[0]).strip().lower() if row and row[0] else ''
        if 'abr' in first:
            header = idx
            break
    for idx, row in enumerate(ws.iter_rows(min_row=header + 1, values_only=True), header + 1):
        yield idx, row


def _match_loge(abrev, nom):
    """Retrouve une loge existante par abréviation, sinon par nom (insensible à la
    casse). Évite les doublons quand l'abréviation OU le nom a changé."""
    loge = None
    if abrev:
        loge = Loge.objects.filter(abreviation__iexact=abrev).first()
    if not loge and nom:
        loge = Loge.objects.filter(nom__iexact=nom).first()
    return loge


_TEMPLES_ALIAS = {
    'lafayette': 'lafayette', 'liberte': 'liberte', 'liberté': 'liberte',
    'egalite': 'egalite', 'égalité': 'egalite', 'fraternite': 'fraternite',
    'fraternité': 'fraternite',
}


def _parse_date_cell(v):
    from datetime import datetime, date as _date
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, _date):
        return v
    if v is None:
        return None
    s = str(v).strip()
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def _parse_heure_cell(v, defaut=None):
    from datetime import datetime, time as _time
    if isinstance(v, _time):
        return v.strftime('%H:%M')
    if isinstance(v, datetime):
        return v.strftime('%H:%M')
    if v is None or str(v).strip() == '':
        return defaut
    s = str(v).strip().replace('h', ':').replace('H', ':')
    for fmt in ('%H:%M', '%H:%M:%S'):
        try:
            return datetime.strptime(s, fmt).strftime('%H:%M')
        except ValueError:
            pass
    return defaut


def _match_temple(lieu):
    key = lieu.lower().replace('temple', '').strip()
    tk = _TEMPLES_ALIAS.get(key)
    return Temple.objects.filter(nom=tk).first() if tk else None


def _match_salle(lieu):
    if not lieu:
        return None
    l = lieu.strip()
    s = SalleReunion.objects.filter(nom__iexact=l).first()
    if s:
        return s
    if 'humide' in l.lower():
        s = SalleReunion.objects.filter(type_salle='agapes').exclude(nom__icontains='oie').first()
        if s:
            return s
    core = l.lower().replace('salle', '').strip()
    if core:
        s = SalleReunion.objects.filter(nom__icontains=core).first()
        if s:
            return s
    return SalleReunion.objects.filter(nom__icontains=l).first()


def _resoudre_lieu(lieu, type_r):
    """Renvoie ('temple', Temple) ou ('salle', SalleReunion) ou (None, None)."""
    type_r = (type_r or '').strip().lower()
    if type_r == 'temple':
        return ('temple', _match_temple(lieu))
    if type_r in ('salle_reunion', 'cabinet_reflexion', 'banquet', 'agapes', 'salle'):
        return ('salle', _match_salle(lieu))
    # Type absent : on infère depuis le lieu
    t = _match_temple(lieu)
    if t:
        return ('temple', t)
    return ('salle', _match_salle(lieu))


def _analyser_import(wb):
    """Analyse (lecture seule) pour la prévisualisation : indique pour chaque loge
    si elle est nouvelle ou mise à jour (rapprochement), et le statut des règles."""
    analyse = {'loges': [], 'regles': [], 'ponctuelles': [],
               'has_loges': False, 'has_regles': False, 'has_ponctuelles': False}
    abrevs_sheet = set()

    if 'LOGES' in wb.sheetnames:
        analyse['has_loges'] = True
        for _i, row in _data_rows(wb['LOGES']):
            if not row or not row[0]:
                continue
            abrev = str(row[0]).strip()
            nom   = str(row[1]).strip() if len(row) > 1 and row[1] else abrev
            abrevs_sheet.add(abrev.lower())
            existing = _match_loge(abrev, nom)
            if existing:
                if existing.abreviation.lower() != abrev.lower():
                    match = f"↔ {existing.abreviation} → {abrev} (abréviation modifiée)"
                elif existing.nom.lower() != nom.lower():
                    match = f"↔ nom modifié (« {existing.nom} »)"
                else:
                    match = "déjà en base"
            else:
                match = ""
            analyse['loges'].append({
                'abrev': abrev, 'nom': nom,
                'obedience': str(row[2]).strip() if len(row) > 2 and row[2] else '',
                'contact': str(row[8]).strip() if len(row) > 8 and row[8] else '',
                'statut': 'maj' if existing else 'nouvelle',
                'match': match,
            })

    regles_sheet = next((n for n in wb.sheetnames if 'GLES' in n and 'CURRENCE' in n), None)
    if regles_sheet:
        analyse['has_regles'] = True
        existing_abrevs = {a.lower() for a in Loge.objects.values_list('abreviation', flat=True) if a}
        existing_noms = {n.lower() for n in Loge.objects.values_list('nom', flat=True) if n}
        for _i, row in _data_rows(wb[regles_sheet]):
            if not row or not row[0]:
                continue
            abrev = str(row[0]).strip()
            nom   = str(row[1]).strip() if len(row) > 1 and row[1] else ''
            loge_ok = (abrev.lower() in existing_abrevs or abrev.lower() in abrevs_sheet
                       or (nom and nom.lower() in existing_noms))
            analyse['regles'].append({
                'abrev': abrev,
                'temple': str(row[4]).strip() if len(row) > 4 and row[4] else '',
                'jour': str(row[5]).strip() if len(row) > 5 and row[5] else '',
                'sem': row[6] if len(row) > 6 else '',
                'heures': (f"{row[7]}–{row[8]}" if len(row) > 8 and row[7] else ''),
                'mois': str(row[9]) if len(row) > 9 and row[9] else '',
                'statut': 'ok' if loge_ok else 'erreur',
                'detail': '' if loge_ok else f"loge « {abrev} » introuvable",
            })

    ponct_sheet = next((n for n in wb.sheetnames if 'PONCTUEL' in n.upper()), None)
    if ponct_sheet:
        analyse['has_ponctuelles'] = True
        for _i, row in _data_rows(wb[ponct_sheet]):
            if not row or not row[0] or (len(row) > 2 and not row[2]):
                continue
            abrev = str(row[0]).strip()
            nom = str(row[1]).strip() if len(row) > 1 and row[1] else ''
            loge = _match_loge(abrev, nom)
            d = _parse_date_cell(row[2] if len(row) > 2 else None)
            hd = _parse_heure_cell(row[3] if len(row) > 3 else None, '19:00')
            lieu = str(row[5]).strip() if len(row) > 5 and row[5] else ''
            type_r = str(row[6]).strip().lower() if len(row) > 6 and row[6] else ''
            kind, ressource = _resoudre_lieu(lieu, type_r)
            in_sheet = abrev.lower() in abrevs_sheet
            statut, detail = 'nouvelle', ''
            if not loge and not in_sheet:
                statut, detail = 'erreur', f"loge « {abrev} » introuvable"
            elif not d:
                statut, detail = 'erreur', "date invalide"
            elif not ressource:
                statut, detail = 'erreur', f"lieu « {lieu} » introuvable"
            elif loge:
                # Loge déjà en base : on peut vérifier le doublon
                if kind == 'temple':
                    exists = Reservation.objects.filter(loge=loge, temple=ressource, date=d, heure_debut=hd).exists()
                else:
                    exists = ReservationSalle.objects.filter(loge=loge, salle=ressource, date=d, heure_debut=hd).exists()
                if exists:
                    statut, detail = 'existante', 'déjà en base'
            analyse['ponctuelles'].append({
                'abrev': abrev, 'date': d.strftime('%d/%m/%Y') if d else str(row[2] if len(row) > 2 else ''),
                'lieu': lieu, 'type': type_r or (kind or ''),
                'objet': str(row[7]).strip() if len(row) > 7 and row[7] else '',
                'statut': statut, 'detail': detail,
            })
        analyse['nb_ponct_new'] = sum(1 for p in analyse['ponctuelles'] if p['statut'] == 'nouvelle')
        analyse['nb_ponct_exist'] = sum(1 for p in analyse['ponctuelles'] if p['statut'] == 'existante')
        analyse['nb_ponct_err'] = sum(1 for p in analyse['ponctuelles'] if p['statut'] == 'erreur')

    analyse['nb_loges_new'] = sum(1 for l in analyse['loges'] if l['statut'] == 'nouvelle')
    analyse['nb_loges_maj'] = sum(1 for l in analyse['loges'] if l['statut'] == 'maj')
    analyse['nb_regles'] = len(analyse['regles'])
    analyse['nb_regles_err'] = sum(1 for r in analyse['regles'] if r['statut'] == 'erreur')
    return analyse


def _importer_donnees(wb):
    errors = []
    stats  = {'loges': 0, 'loges_maj': 0, 'obediences': 0,
              'regles': 0, 'regles_maj': 0, 'ponctuelles': 0}

    if 'LOGES' in wb.sheetnames:
        for i, row in _data_rows(wb['LOGES']):
            try:
                if not row[0]: continue
                ob, co = Obedience.objects.get_or_create(nom=str(row[2]).strip() if row[2] else 'Non définie')
                if co: stats['obediences'] += 1
                RITES_VALIDES = ['reaa','rer','rf','rf_reaa','rem','dh','mem','rapmm','rmfr','emulation','marque','autre']
                RITE_ALIASES  = {'rf/reaa': 'rf_reaa', 'reaa/rf': 'rf_reaa'}
                def _normalise_rite(raw):
                    r = raw.strip().lower()
                    return RITE_ALIASES.get(r, r if r in RITES_VALIDES else '')
                # Nouveau format : col4=rite, col5=email, col6=effectif, col7=agapes
                # Ancien format  : col4=email, col5=effectif, col6=agapes, col7=rite
                # Détection : si col4 est dans les rites valides → nouveau format
                col4_val = str(row[4]).strip().lower() if len(row) > 4 and row[4] else ''
                col4_norm = RITE_ALIASES.get(col4_val, col4_val)
                nouveau_format = col4_norm in RITES_VALIDES or col4_val == ''
                nom_contact = ''
                telephone   = ''
                association = ''
                if nouveau_format:
                    rite     = _normalise_rite(str(row[4]) if len(row) > 4 and row[4] else '')
                    email    = str(row[5]).strip() if len(row) > 5 and row[5] else ''
                    effectif = int(row[6]) if len(row) > 6 and row[6] and str(row[6]).isdigit() else 0
                    agapes   = int(row[7]) if len(row) > 7 and row[7] and str(row[7]).isdigit() else 0
                    nom_contact = str(row[8]).strip() if len(row) > 8 and row[8] else ''
                    telephone   = str(row[9]).strip() if len(row) > 9 and row[9] else ''
                    association = str(row[10]).strip() if len(row) > 10 and row[10] else ''
                else:
                    email    = col4_val
                    effectif = int(row[5]) if len(row) > 5 and row[5] and str(row[5]).isdigit() else 0
                    agapes   = int(row[6]) if len(row) > 6 and row[6] and str(row[6]).isdigit() else 0
                    rite     = _normalise_rite(str(row[7]) if len(row) > 7 and row[7] else '')
                abrev = str(row[0]).strip()
                nom   = str(row[1]).strip() if row[1] else abrev
                vals = {
                    'nom': nom, 'abreviation': abrev, 'obedience': ob,
                    'type_loge': str(row[3]).strip() if row[3] in ('loge', 'haut_grade') else 'loge',
                    'rite': rite, 'email': email,
                    'effectif_total': effectif, 'effectif_moyen_agapes': agapes,
                    # Présente dans le fichier = a renvoyé sa fiche → réactivée
                    'statut': 'active', 'actif': True,
                }
                # Ne pas écraser le contact si les colonnes ne sont pas renseignées
                if nom_contact:
                    vals['nom_contact'] = nom_contact
                if telephone:
                    vals['telephone'] = telephone
                if association:
                    vals['association'] = association
                # Rapprochement : abréviation OU nom (évite les doublons)
                loge = _match_loge(abrev, nom)
                if loge:
                    for k, v in vals.items():
                        setattr(loge, k, v)
                    loge.save()
                    stats['loges_maj'] += 1
                else:
                    Loge.objects.create(**vals)
                    stats['loges'] += 1
            except Exception as e:
                errors.append(f"LOGES ligne {i} : {e}")

    # Accepter l'ancien nom sans accents et le nouveau avec accents
    regles_sheet = next((n for n in wb.sheetnames if 'GLES' in n and 'CURRENCE' in n), None)
    if regles_sheet:
        JOURS  = {'Lundi':0,'Mardi':1,'Mercredi':2,'Jeudi':3,'Vendredi':4,'Samedi':5,'Dimanche':6}
        TEMPLES = {'Lafayette':'lafayette','Liberte':'liberte','Egalite':'egalite','Fraternite':'fraternite',
                   'Égalité':'egalite','Fraternité':'fraternite','Liberté':'liberte'}
        for i, row in _data_rows(wb[regles_sheet]):
            try:
                if not row[0] or not row[4] or not row[5] or row[6] is None: continue
                nom_r = str(row[1]).strip() if len(row) > 1 and row[1] else ''
                loge = _match_loge(str(row[0]).strip(), nom_r)
                if not loge:
                    errors.append(f"REGLES ligne {i} : loge '{row[0]}' introuvable"); continue
                tk = TEMPLES.get(str(row[4]).strip())
                if not tk: errors.append(f"REGLES ligne {i} : temple '{row[4]}' inconnu"); continue
                try: temple = Temple.objects.get(nom=tk)
                except Temple.DoesNotExist: errors.append(f"REGLES ligne {i} : temple absent"); continue
                jn = JOURS.get(str(row[5]).strip())
                if jn is None: errors.append(f"REGLES ligne {i} : jour '{row[5]}' inconnu"); continue
                mois_raw = str(row[9]) if len(row) > 9 and row[9] is not None else ''
                mois_actifs = [int(m) for m in mois_raw.replace(' ', '').split(',')
                               if m.strip().isdigit() and 1 <= int(m) <= 12]
                defaults = {
                    'heure_debut': str(row[7]) if len(row) > 7 and row[7] else '19:30',
                    'heure_fin': str(row[8]) if len(row) > 8 and row[8] else '22:30',
                    'mois_actifs': mois_actifs, 'actif': True,
                }
                # Robuste aux doublons existants (get() planterait sur >1)
                existantes = RegleRecurrence.objects.filter(
                    loge=loge, temple=temple, jour_semaine=jn, numero_semaine=int(row[6]),
                ).order_by('pk')
                if existantes.exists():
                    regle = existantes.first()
                    existantes.exclude(pk=regle.pk).delete()   # supprime les doublons
                    for k, v in defaults.items():
                        setattr(regle, k, v)
                    regle.save()
                    stats['regles_maj'] += 1
                else:
                    RegleRecurrence.objects.create(
                        loge=loge, temple=temple, jour_semaine=jn,
                        numero_semaine=int(row[6]), **defaults)
                    stats['regles'] += 1
            except Exception as e:
                errors.append(f"REGLES ligne {i} : {e}")

    # ── Dates ponctuelles (banquets, salles, cabinets, temples exceptionnels) ──
    ponct_sheet = next((n for n in wb.sheetnames if 'PONCTUEL' in n.upper()), None)
    if ponct_sheet:
        OBJET_DEFAUT = {'temple': "Tenue exceptionnelle", 'banquet': "Banquet d'ordre",
                        'cabinet_reflexion': "Cabinet de réflexion", 'salle_reunion': "Réunion"}
        for i, row in _data_rows(wb[ponct_sheet]):
            try:
                if not row or not row[0] or (len(row) > 2 and not row[2]):
                    continue
                abrev = str(row[0]).strip()
                nom_r = str(row[1]).strip() if len(row) > 1 and row[1] else ''
                loge = _match_loge(abrev, nom_r)
                if not loge:
                    errors.append(f"PONCTUELLES ligne {i} : loge '{abrev}' introuvable"); continue
                d = _parse_date_cell(row[2] if len(row) > 2 else None)
                if not d:
                    errors.append(f"PONCTUELLES ligne {i} : date invalide"); continue
                hd = _parse_heure_cell(row[3] if len(row) > 3 else None, '19:00')
                hf = _parse_heure_cell(row[4] if len(row) > 4 else None, '22:00')
                lieu = str(row[5]).strip() if len(row) > 5 and row[5] else ''
                type_r = str(row[6]).strip().lower() if len(row) > 6 and row[6] else ''
                objet = str(row[7]).strip() if len(row) > 7 and row[7] else ''
                kind, ressource = _resoudre_lieu(lieu, type_r)
                if not ressource:
                    errors.append(f"PONCTUELLES ligne {i} : lieu '{lieu}' introuvable"); continue

                if kind == 'temple':
                    if Reservation.objects.filter(loge=loge, temple=ressource, date=d, heure_debut=hd).exists():
                        continue  # doublon
                    Reservation.objects.create(
                        loge=loge, temple=ressource, type_reservation='exceptionnelle',
                        sous_type='standard', statut='validee', date=d,
                        heure_debut=hd, heure_fin=hf,
                        nom_demandeur=loge.nom_contact or loge.nom,
                        email_demandeur=loge.email or '', commentaire=objet,
                    )
                else:
                    if ReservationSalle.objects.filter(loge=loge, salle=ressource, date=d, heure_debut=hd).exists():
                        continue  # doublon
                    ReservationSalle.objects.create(
                        loge=loge, salle=ressource, date=d, heure_debut=hd, heure_fin=hf,
                        statut='validee', nom_demandeur=loge.nom_contact or loge.nom,
                        email_demandeur=loge.email or '', organisation=loge.nom,
                        objet=objet or OBJET_DEFAUT.get(type_r, "Réservation"),
                        nombre_participants=loge.effectif_moyen_agapes or 1, commentaire=objet,
                    )
                stats['ponctuelles'] += 1
            except Exception as e:
                errors.append(f"PONCTUELLES ligne {i} : {e}")

    return stats, errors


def _calculer_dates_regle(regle, annee):
    mois_list = regle.mois_actifs if regle.mois_actifs else list(range(1, 13))
    return [d for mois in mois_list for d in [_nieme_jour_du_mois(annee, mois, regle.numero_semaine, regle.jour_semaine)] if d]


# ── Paramètres ────────────────────────────────────────────────────────────────

@staff_required
def parametres(request):
    params = Parametres.get_instance()
    if request.method == 'POST':
        params.mot_de_passe_annuel = request.POST.get('mot_de_passe_annuel', params.mot_de_passe_annuel)
        params.email_admin    = request.POST.get('email_admin',    params.email_admin)
        params.email_traiteur = request.POST.get('email_traiteur', params.email_traiteur)
        params.email_from     = request.POST.get('email_from',     params.email_from)
        params.smtp_host = request.POST.get('smtp_host', params.smtp_host)
        params.smtp_port = int(request.POST.get('smtp_port', params.smtp_port))
        params.smtp_user = request.POST.get('smtp_user', params.smtp_user)
        params.smtp_password = request.POST.get('smtp_password', params.smtp_password)
        params.smtp_tls = request.POST.get('smtp_tls') == 'on'
        params.save()
        messages.success(request, "Paramètres sauvegardés.")
        return redirect('administration:parametres')
    return render(request, 'administration/parametres.html', {'params': params})


@staff_required
def tester_smtp(request):
    if request.method != 'POST':
        return redirect('administration:parametres')
    dest = get_email_admin()
    try:
        send_mail_kellermann(
            subject="[Kellermann] Test SMTP",
            message=(
                "Cet email confirme que la configuration SMTP est fonctionnelle.\n\n"
                "Si vous recevez ce message, les paramètres SMTP sont correctement configurés."
            ),
            recipient_list=[dest],
            fail_silently=False,
        )
        messages.success(request, f"Email de test envoyé avec succès à {dest}.")
    except Exception as e:
        messages.error(request, f"Échec de l'envoi : {e}")
    return redirect('administration:parametres')


# ── Gestion des salles ────────────────────────────────────────────────────────

@staff_required
def salles_liste(request):
    salles = SalleReunion.objects.all().order_by('nom')
    return render(request, 'administration/salles_liste.html', {
        'salles': salles,
        'nb_salles': salles.count(),
    })


@staff_required
def salle_form(request, pk=None):
    salle = get_object_or_404(SalleReunion, pk=pk) if pk else None
    if request.method == 'POST':
        try:
            data = {
                'nom': request.POST['nom'],
                'type_salle': request.POST['type_salle'],
                'capacite': int(request.POST['capacite']),
                'description': request.POST.get('description', '').strip(),
                'actif': request.POST.get('actif') == 'on',
            }
            if salle:
                for k, v in data.items():
                    setattr(salle, k, v)
                salle.save()
                messages.success(request, "Salle modifiée.")
            else:
                SalleReunion.objects.create(**data)
                messages.success(request, "Salle ajoutée.")
            return redirect('administration:salles_liste')
        except Exception as e:
            messages.error(request, f"Erreur : {e}")

    return render(request, 'administration/salle_form.html', {
        'salle': salle,
        'type_choices': SalleReunion.TYPE_CHOICES,
    })


@staff_required
def salle_supprimer(request, pk):
    salle = get_object_or_404(SalleReunion, pk=pk)
    if request.method == 'POST':
        salle.delete()
        messages.success(request, "Salle supprimée.")
        return redirect('administration:salles_liste')
    return render(request, 'administration/salle_supprimer.html', {'salle': salle})


def _couverts_admin(t):
    """(valeur_int, est_estimation)
    est_estimation=True + valeur=0  →  non renseigné (afficher '~?')
    est_estimation=True + valeur>0  →  estimation via effectif_moyen_agapes
    est_estimation=False            →  valeur confirmée
    """
    nombre = getattr(t, 'nombre_repas', 0) or 0
    if nombre > 0:
        return (nombre, False)
    loge = getattr(t, 'loge', None)
    if loge is not None:
        effectif = getattr(loge, 'effectif_moyen_agapes', 0) or 0
        if effectif > 0:
            return (effectif, True)
    # Aucune donnée disponible — signalé comme estimation inconnue
    return (0, True)


@staff_required
def agapes_traiteur(request):
    """Vue synthétique agapes + banquets pour le traiteur."""
    today = date.today()
    annee_courante = today.year if today.month >= 9 else today.year - 1
    annee_param = int(request.GET.get('annee', annee_courante))
    debut_saison = date(annee_param, 9, 1)
    fin_saison   = date(annee_param + 1, 6, 30)

    # Tenues avec agapes
    tenues = (
        Reservation.objects
        .select_related('loge', 'temple')
        .filter(
            besoin_agapes=True,
            statut='validee',
            date__gte=debut_saison,
            date__lte=fin_saison,
        )
        .order_by('date')
    )

    # Banquets (ReservationSalle type agapes)
    banquets = (
        ReservationSalle.objects
        .select_related('salle')
        .filter(
            salle__type_salle='agapes',
            statut='validee',
            date__gte=debut_saison,
            date__lte=fin_saison,
        )
        .order_by('date')
    )

    # Fusion en liste normalisée
    lignes = []
    for t in tenues:
        couverts, est_estim = _couverts_admin(t)
        lignes.append({
            'date':           t.date,
            'organisation':   t.loge.nom if t.loge else (t.nom_organisation or t.nom_demandeur),
            'type':           'Tenue + agapes',
            'couverts':       couverts,
            'est_estimation': est_estim,
            'lieu':           str(t.temple),
            'horaires':       f"{t.heure_debut:%H:%M} – {t.heure_fin:%H:%M}",
            'commentaire':    t.commentaire,
        })
    for b in banquets:
        lignes.append({
            'date':           b.date,
            'organisation':   b.organisation or b.nom_demandeur,
            'type':           "Banquet d'ordre",
            'couverts':       b.nombre_participants,
            'est_estimation': False,
            'lieu':           str(b.salle),
            'horaires':       f"{b.heure_debut:%H:%M} – {b.heure_fin:%H:%M}",
            'commentaire':    b.commentaire,
        })
    lignes.sort(key=lambda x: x['date'])

    # Totaux par mois — tous les mois de la saison, même vides
    MOIS_ORDRE = [9, 10, 11, 12, 1, 2, 3, 4, 5, 6]
    MOIS_NOMS  = {1:'Janvier',2:'Février',3:'Mars',4:'Avril',5:'Mai',6:'Juin',
                  9:'Septembre',10:'Octobre',11:'Novembre',12:'Décembre'}
    totaux_mois = {}
    for m in MOIS_ORDRE:
        sous_liste = [l for l in lignes if l['date'].month == m]
        totaux_mois[m] = {
            'nom':            MOIS_NOMS[m],
            'lignes':         sous_liste,
            'total':          sum(l['couverts'] for l in sous_liste),
            'has_estimations': any(l['est_estimation'] for l in sous_liste),
        }

    context = {
        'lignes':        lignes,
        'totaux_mois':   totaux_mois,
        'total_saison':  sum(l['couverts'] for l in lignes),
        'has_estimations': any(l['est_estimation'] for l in lignes),
        'annee':         annee_param,
        'annees':        list(range(annee_courante - 2, annee_courante + 4)),
        'saison_label':  f"{annee_param}/{annee_param + 1}",
        'mois_liste':    [(m, MOIS_NOMS[m]) for m in MOIS_ORDRE],
    }
    return render(request, 'administration/agapes_traiteur.html', context)


@staff_required
def agapes_export_excel(request):
    """Export Excel agapes/banquets — période et type filtrables."""
    from datetime import datetime as dt
    today = date.today()
    annee_courante = today.year if today.month >= 9 else today.year - 1

    # Paramètres de filtrage
    annee_param  = int(request.GET.get('annee', annee_courante))
    date_debut_s = request.GET.get('date_debut', '')
    date_fin_s   = request.GET.get('date_fin', '')
    type_export  = request.GET.get('type_export', 'tout')  # tout | agapes | banquet

    try:
        debut_saison = dt.strptime(date_debut_s, '%Y-%m-%d').date() if date_debut_s else date(annee_param, 9, 1)
        fin_saison   = dt.strptime(date_fin_s,   '%Y-%m-%d').date() if date_fin_s   else date(annee_param + 1, 6, 30)
    except ValueError:
        debut_saison = date(annee_param, 9, 1)
        fin_saison   = date(annee_param + 1, 6, 30)

    tenues = (
        Reservation.objects
        .select_related('loge', 'temple')
        .filter(besoin_agapes=True, statut='validee',
                date__gte=debut_saison, date__lte=fin_saison)
        .order_by('date')
    )
    banquets = (
        ReservationSalle.objects
        .select_related('salle')
        .filter(salle__type_salle='agapes', statut='validee',
                date__gte=debut_saison, date__lte=fin_saison)
        .order_by('date')
    )
    if type_export == 'agapes':
        banquets = banquets.none()
    elif type_export == 'banquet':
        tenues = tenues.none()

    lignes = []
    for t in tenues:
        couverts, est_estim = _couverts_admin(t)
        couverts_affiche = f"~{couverts} (estim.)" if est_estim else couverts
        lignes.append((
            t.date,
            t.loge.nom if t.loge else (t.nom_organisation or t.nom_demandeur),
            'Tenue + agapes',
            couverts_affiche,
            couverts,
            str(t.temple),
            f"{t.heure_debut:%H:%M} – {t.heure_fin:%H:%M}",
            t.commentaire,
        ))
    for b in banquets:
        lignes.append((
            b.date,
            b.organisation or b.nom_demandeur,
            "Banquet d'ordre",
            b.nombre_participants,
            b.nombre_participants,
            str(b.salle),
            f"{b.heure_debut:%H:%M} – {b.heure_fin:%H:%M}",
            b.commentaire,
        ))
    lignes.sort(key=lambda x: x[0])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Agapes {debut_saison:%d/%m/%Y}-{fin_saison:%d/%m/%Y}"[:31]

    # Styles
    hf    = Font(bold=True, color="C8A84B")
    hfill = PatternFill("solid", fgColor="0F2137")
    ctr   = Alignment(horizontal="center", vertical="center")
    thin  = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    total_fill = PatternFill("solid", fgColor="F1F5F9")
    total_font = Font(bold=True)

    headers = ["Date", "Loge / Organisation", "Type", "Couverts", "Lieu", "Horaires", "Commentaire"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hf; cell.fill = hfill; cell.alignment = ctr; cell.border = thin
    ws.row_dimensions[1].height = 20

    col_widths = [14, 36, 20, 12, 22, 18, 40]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    MOIS_NOMS = {1:'Janvier',2:'Février',3:'Mars',4:'Avril',5:'Mai',6:'Juin',
                 7:'Juillet',8:'Août',9:'Septembre',10:'Octobre',11:'Novembre',12:'Décembre'}

    # Construire l'ordre des mois couverts par la période réelle
    mois_presents = sorted({l[0].month for l in lignes}, key=lambda m: (m < debut_saison.month, m))
    # Fallback saison classique si aucune donnée
    if not mois_presents:
        mois_presents = []

    # Tuple layout : (date_obj, organisation, type, couverts_affiche, couverts_num, lieu, horaires, commentaire)
    row_idx = 2
    for mois in mois_presents:
        mois_lignes = sorted([l for l in lignes if l[0].month == mois], key=lambda l: l[0])
        if not mois_lignes:
            continue
        # Séparateur de mois
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=7)
        sep = ws.cell(row=row_idx, column=1, value=MOIS_NOMS[mois].upper())
        sep.font = Font(bold=True, color="0F2137")
        sep.fill = PatternFill("solid", fgColor="E2E8F0")
        sep.alignment = ctr; sep.border = thin
        row_idx += 1
        # Lignes — 7 colonnes Excel : date, org, type, couverts, lieu, horaires, commentaire
        for l in mois_lignes:
            excel_row = [l[0].strftime('%d/%m/%Y'), l[1], l[2], l[3], l[5], l[6], l[7]]
            for col, val in enumerate(excel_row, 1):
                c = ws.cell(row=row_idx, column=col, value=val)
                c.border = thin
                if col == 4:  # Couverts
                    c.alignment = ctr
            row_idx += 1
        # Total mois (valeur numérique = index 4)
        total = sum(l[4] for l in mois_lignes)
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=3)
        ws.cell(row=row_idx, column=1, value=f"Total {MOIS_NOMS[mois]}").font = total_font
        ws.cell(row=row_idx, column=1).fill = total_fill
        ws.cell(row=row_idx, column=1).border = thin
        tc = ws.cell(row=row_idx, column=4, value=total)
        tc.font = total_font; tc.fill = total_fill
        tc.alignment = ctr; tc.border = thin
        for col in range(5, 8):
            ws.cell(row=row_idx, column=col).fill = total_fill
            ws.cell(row=row_idx, column=col).border = thin
        row_idx += 1

    # Total saison
    total_saison = sum(l[4] for l in lignes)
    row_idx += 1
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=3)
    ws.cell(row=row_idx, column=1, value=f"TOTAL  {debut_saison:%d/%m/%Y} → {fin_saison:%d/%m/%Y}").font = Font(bold=True, color="C8A84B")
    ws.cell(row=row_idx, column=1).fill = PatternFill("solid", fgColor="0F2137")
    ws.cell(row=row_idx, column=1).border = thin
    ts = ws.cell(row=row_idx, column=4, value=total_saison)
    ts.font = Font(bold=True, color="C8A84B")
    ts.fill = PatternFill("solid", fgColor="0F2137")
    ts.alignment = ctr; ts.border = thin

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    periode_label = f"{debut_saison:%d%m%Y}-{fin_saison:%d%m%Y}"
    response['Content-Disposition'] = f'attachment; filename="agapes_{periode_label}.xlsx"'
    wb.save(response)
    return response


@staff_required
def agapes_export_pdf(request):
    """Export PDF de la synthèse agapes/banquets (mensuel ou 7 jours)."""
    from io import BytesIO
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT

    today = date.today()
    annee_courante = today.year if today.month >= 9 else today.year - 1
    annee_param = int(request.GET.get('annee', annee_courante))
    periode = request.GET.get('periode', 'mensuel')
    mois_param = request.GET.get('mois')

    MOIS_NOMS = {1:'Janvier',2:'Février',3:'Mars',4:'Avril',5:'Mai',6:'Juin',
                 7:'Juillet',8:'Août',9:'Septembre',10:'Octobre',11:'Novembre',12:'Décembre'}
    JOURS_FR  = {0:'Lun',1:'Mar',2:'Mer',3:'Jeu',4:'Ven',5:'Sam',6:'Dim'}

    if periode == 'hebdo':
        date_debut = today
        date_fin   = today + timedelta(days=6)
        titre_periode = f"7 prochains jours ({date_debut.strftime('%d/%m')} – {date_fin.strftime('%d/%m/%Y')})"
        nom_fichier = f"agapes_7jours_{today.strftime('%Y%m%d')}.pdf"
    else:
        if mois_param:
            mois_int = int(mois_param)
        else:
            mois_int = today.month
        annee_mois = annee_param if mois_int >= 9 else annee_param + 1
        import calendar as _cal
        dernier_jour = _cal.monthrange(annee_mois, mois_int)[1]
        date_debut = date(annee_mois, mois_int, 1)
        date_fin   = date(annee_mois, mois_int, dernier_jour)
        titre_periode = f"{MOIS_NOMS[mois_int]} {annee_mois}"
        nom_fichier = f"agapes_{annee_mois}_{mois_int:02d}.pdf"

    # Requêtes
    tenues = (
        Reservation.objects
        .select_related('loge', 'temple')
        .filter(besoin_agapes=True, statut='validee',
                date__gte=date_debut, date__lte=date_fin)
        .order_by('date')
    )
    banquets = (
        ReservationSalle.objects
        .select_related('salle')
        .filter(salle__type_salle='agapes', statut='validee',
                date__gte=date_debut, date__lte=date_fin)
        .order_by('date')
    )

    lignes = []
    for t in tenues:
        lignes.append({
            'date':         t.date,
            'organisation': t.loge.nom if t.loge else (t.nom_organisation or t.nom_demandeur),
            'type':         'Tenue + agapes',
            'couverts':     t.nombre_repas,
            'lieu':         str(t.temple),
            'horaires':     f"{t.heure_debut:%H:%M}–{t.heure_fin:%H:%M}",
        })
    for b in banquets:
        lignes.append({
            'date':         b.date,
            'organisation': b.organisation or b.nom_demandeur,
            'type':         "Banquet d'ordre",
            'couverts':     b.nombre_participants,
            'lieu':         str(b.salle),
            'horaires':     f"{b.heure_debut:%H:%M}–{b.heure_fin:%H:%M}",
        })
    lignes.sort(key=lambda x: x['date'])

    # Couleurs
    C_NAVY  = colors.HexColor('#0F2137')
    C_GOLD  = colors.HexColor('#C8A84B')
    C_LIGHT = colors.HexColor('#F8FAFC')
    C_TOTAL = colors.HexColor('#E2E8F0')

    # Construction du PDF
    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=1.5*cm, bottomMargin=2*cm,
    )

    sty_titre = ParagraphStyle('titre', fontSize=16, textColor=C_NAVY,
                               spaceAfter=4, fontName='Helvetica-Bold')
    sty_sous  = ParagraphStyle('sous',  fontSize=11, textColor=C_GOLD,
                               spaceAfter=12, fontName='Helvetica-Bold')
    sty_pied  = ParagraphStyle('pied',  fontSize=8,  textColor=colors.grey,
                               alignment=TA_CENTER)

    story = []

    # En-tête
    story.append(Paragraph("Kellermann — Recapitulatif Agapes", sty_titre))
    story.append(Paragraph(titre_periode, sty_sous))

    if not lignes:
        story.append(Paragraph("Aucun evenement sur cette periode.", ParagraphStyle('x', fontSize=10)))
    else:
        # Tableau
        headers = ["Date", "Loge / Organisation", "Type", "Couverts", "Lieu", "Horaires"]
        col_widths = [2.5*cm, 5.5*cm, 3.2*cm, 2*cm, 3.2*cm, 2.6*cm]

        table_data = [headers]
        for l in lignes:
            jour_fr = JOURS_FR.get(l['date'].weekday(), '')
            table_data.append([
                f"{jour_fr} {l['date'].strftime('%d/%m/%Y')}",
                l['organisation'],
                l['type'],
                str(l['couverts']),
                l['lieu'],
                l['horaires'],
            ])

        # Ligne total
        total = sum(l['couverts'] for l in lignes)
        table_data.append(["", "TOTAL", "", str(total), "", ""])

        tbl = Table(table_data, colWidths=col_widths, repeatRows=1)

        n = len(table_data)
        style = TableStyle([
            # En-tête
            ('BACKGROUND',   (0,0), (-1,0),  C_NAVY),
            ('TEXTCOLOR',    (0,0), (-1,0),  C_GOLD),
            ('FONTNAME',     (0,0), (-1,0),  'Helvetica-Bold'),
            ('FONTSIZE',     (0,0), (-1,0),  8),
            ('ALIGN',        (0,0), (-1,0),  'CENTER'),
            ('BOTTOMPADDING',(0,0), (-1,0),  6),
            ('TOPPADDING',   (0,0), (-1,0),  6),
            # Corps
            ('FONTNAME',     (0,1), (-1,n-2), 'Helvetica'),
            ('FONTSIZE',     (0,1), (-1,n-2), 8),
            ('ROWBACKGROUNDS',(0,1),(-1,n-2), [colors.white, C_LIGHT]),
            ('ALIGN',        (3,1), (3,n-2),  'CENTER'),
            ('FONTNAME',     (3,1), (3,n-2),  'Helvetica-Bold'),
            # Ligne total
            ('BACKGROUND',   (0,n-1), (-1,n-1), C_TOTAL),
            ('FONTNAME',     (0,n-1), (-1,n-1), 'Helvetica-Bold'),
            ('FONTSIZE',     (0,n-1), (-1,n-1), 9),
            ('ALIGN',        (1,n-1), (1,n-1),  'RIGHT'),
            ('ALIGN',        (3,n-1), (3,n-1),  'CENTER'),
            ('TEXTCOLOR',    (0,n-1), (-1,n-1), C_NAVY),
            ('TOPPADDING',   (0,n-1), (-1,n-1), 6),
            ('BOTTOMPADDING',(0,n-1), (-1,n-1), 6),
            # Bordures
            ('GRID',         (0,0),  (-1,-1),  0.4, colors.HexColor('#CBD5E1')),
            ('BOX',          (0,0),  (-1,-1),  1,   C_NAVY),
        ])
        tbl.setStyle(style)
        story.append(tbl)
        story.append(Spacer(1, 0.5*cm))

    # Pied de page
    story.append(Spacer(1, 0.3*cm))
    story.append(Paragraph(
        f"Document genere le {today.strftime('%d/%m/%Y')} — Temples Kellermann",
        sty_pied,
    ))

    doc.build(story)
    buf.seek(0)
    response = HttpResponse(buf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{nom_fichier}"'
    return response


def _nieme_jour_du_mois(annee, mois, n, jour):
    premier = date(annee, mois, 1)
    dernier = date(annee, mois, calendar.monthrange(annee, mois)[1])
    if n > 0:
        delta = (jour - premier.weekday()) % 7
        cible = premier + timedelta(days=delta + (n - 1) * 7)
        return cible if cible.month == mois else None
    else:
        delta = (dernier.weekday() - jour) % 7
        cible = dernier - timedelta(days=delta)
        return cible if cible.month == mois else None


# ── Réservation directe (admin) ───────────────────────────────────────────────

def _to_time(val):
    """Convertit 'HH:MM' en datetime.time (ou renvoie tel quel si déjà un time)."""
    from datetime import datetime, time as _time
    if isinstance(val, _time):
        return val
    return datetime.strptime(val, "%H:%M").time()


def _temple_occupe(temple, date_r, hd_t, hf_t):
    if Reservation.objects.filter(
        temple=temple, date=date_r, heure_debut__lt=hf_t, heure_fin__gt=hd_t,
        statut__in=['attente', 'validee'],
    ).exists():
        return True
    if Indisponibilite.objects.filter(
        temples=temple, date_debut__lte=date_r, date_fin__gte=date_r,
    ).exists():
        return True
    return False


def _salle_occupee(salle, date_r, hd_t, hf_t):
    if ReservationSalle.objects.filter(
        salle=salle, date=date_r, heure_debut__lt=hf_t, heure_fin__gt=hd_t,
        statut__in=['attente', 'validee'],
    ).exists():
        return True
    if BlocageCreneaux.objects.filter(
        salles=salle, date=date_r, heure_debut__lt=hf_t, heure_fin__gt=hd_t,
    ).exists():
        return True
    if Indisponibilite.objects.filter(
        salles=salle, date_debut__lte=date_r, date_fin__gte=date_r,
    ).exists():
        return True
    return False


_SALLE_FIELD_IDS = {
    "reunion": "id_salle_reunion",
    "cabinet": "id_salle_cabinet",
    "agapes":  "id_salle_agapes",
}


def _cabinets_libres(date_r, hd, hf):
    """Liste des cabinets de réflexion libres sur le créneau."""
    hd_t = _to_time(hd)
    hf_t = _to_time(hf)
    libres = []
    for c in SalleReunion.objects.filter(
        type_salle='cabinet_reflexion', actif=True,
    ).order_by('nom'):
        if not _salle_occupee(c, date_r, hd_t, hf_t):
            libres.append(c)
    return libres


def _cabinets_etat(date_r, hd, hf):
    """État de chaque cabinet (libre / occupé par qui) sur le créneau, par nom."""
    hd_t = _to_time(hd)
    hf_t = _to_time(hf)
    etat = []
    for c in SalleReunion.objects.filter(
        type_salle='cabinet_reflexion', actif=True,
    ).order_by('nom'):
        occ = []
        for r in ReservationSalle.objects.filter(
            salle=c, date=date_r, heure_debut__lt=hf_t, heure_fin__gt=hd_t,
            statut__in=['attente', 'validee'],
        ).select_related('loge'):
            occ.append(r.organisation or (r.loge.nom if r.loge else r.nom_demandeur))
        etat.append({'nom': c.nom, 'libre': not occ, 'detail': ', '.join(occ)})
    return etat


def _grille_congres(temples, date_debut, date_fin, hd, hf, salles=None):
    """Grille de disponibilité d'un congrès : pour chaque jour et chaque temple,
    libre ou occupé (avec détail), plus l'état des salles de réunion."""
    from datetime import timedelta
    hd_t = _to_time(hd)
    hf_t = _to_time(hf)
    fin = date_fin or date_debut
    jours_list = []
    j = date_debut
    while j <= fin:
        jours_list.append(j)
        j += timedelta(days=1)

    grille = []
    nb_conflits = 0
    for jour in jours_list:
        cells = []
        for tp in temples:
            occup = []
            for r in Reservation.objects.filter(
                temple=tp, date=jour, heure_debut__lt=hf_t, heure_fin__gt=hd_t,
                statut__in=['attente', 'validee'],
            ).select_related('loge'):
                qui = r.loge or r.nom_organisation or r.nom_demandeur
                occup.append(f"{r.heure_debut:%H:%M}–{r.heure_fin:%H:%M} {qui}")
            for i in Indisponibilite.objects.filter(
                temples=tp, date_debut__lte=jour, date_fin__gte=jour,
            ):
                occup.append(f"indispo : {i.motif}")
            libre = not occup
            if not libre:
                nb_conflits += 1
            cells.append({'temple': str(tp), 'libre': libre, 'detail': ' ; '.join(occup)})
        grille.append({'date': jour, 'cells': cells})

    salles_etat = []
    for s in (salles or []):
        occ_days = [jour.strftime('%d/%m') for jour in jours_list
                    if _salle_occupee(s, jour, hd_t, hf_t)]
        if occ_days:
            nb_conflits += 1
        salles_etat.append({'nom': str(s), 'libre': not occ_days, 'detail': ', '.join(occ_days)})

    return {'temples': [str(t) for t in temples], 'jours': grille,
            'salles': salles_etat, 'nb_conflits': nb_conflits}


def _analyser_disponibilite(type_resa, ressource, date_r, hd, hf):
    """Renvoie (conflits, alternatives) pour la ressource et le créneau demandés."""
    conflits = []
    alternatives = []
    if not ressource:
        return conflits, alternatives
    hd_t = _to_time(hd)
    hf_t = _to_time(hf)

    if type_resa == "temple":
        temple = ressource
        for r in Reservation.objects.filter(
            temple=temple, date=date_r, heure_debut__lt=hf_t, heure_fin__gt=hd_t,
            statut__in=['attente', 'validee'],
        ).select_related('loge'):
            qui = r.loge or r.nom_organisation or r.nom_demandeur
            conflits.append(f"{r.heure_debut:%H:%M}–{r.heure_fin:%H:%M} : {qui} ({r.get_statut_display()})")
        for i in Indisponibilite.objects.filter(
            temples=temple, date_debut__lte=date_r, date_fin__gte=date_r,
        ):
            conflits.append(f"Indisponibilité : {i.motif}")
        if conflits:
            for t in Temple.objects.exclude(pk=temple.pk).order_by('nom'):
                if not _temple_occupe(t, date_r, hd_t, hf_t):
                    alternatives.append({'field': 'id_temple', 'id': t.pk, 'label': str(t)})
    else:
        salle = ressource
        for r in ReservationSalle.objects.filter(
            salle=salle, date=date_r, heure_debut__lt=hf_t, heure_fin__gt=hd_t,
            statut__in=['attente', 'validee'],
        ).select_related('loge'):
            qui = r.organisation or (r.loge.nom if r.loge else r.nom_demandeur)
            conflits.append(f"{r.heure_debut:%H:%M}–{r.heure_fin:%H:%M} : {qui} — {r.objet} ({r.get_statut_display()})")
        for b in BlocageCreneaux.objects.filter(
            salles=salle, date=date_r, heure_debut__lt=hf_t, heure_fin__gt=hd_t,
        ):
            conflits.append(f"Blocage {b.heure_debut:%H:%M}–{b.heure_fin:%H:%M} : {b.motif}")
        for i in Indisponibilite.objects.filter(
            salles=salle, date_debut__lte=date_r, date_fin__gte=date_r,
        ):
            conflits.append(f"Indisponibilité : {i.motif}")
        if conflits:
            champ = _SALLE_FIELD_IDS.get(type_resa, "id_salle_agapes")
            for s in SalleReunion.objects.filter(
                actif=True, type_salle=salle.type_salle,
            ).exclude(pk=salle.pk).order_by('nom'):
                if not _salle_occupee(s, date_r, hd_t, hf_t):
                    alternatives.append({'field': champ, 'id': s.pk, 'label': str(s)})
    return conflits, alternatives


@staff_required
def reservation_directe(request):
    """Créer une réservation directement validée, avec contrôle de disponibilité."""
    from temple_project.apps.traiteur.forms import ReservationDirecteForm

    form = ReservationDirecteForm(request.POST or None)
    conflits = []
    alternatives = []
    dispo_verifiee = False
    creneau = None

    if request.method == "POST" and form.is_valid():
        cd        = form.cleaned_data
        type_resa = cd["type_resa"]
        loge      = cd.get("loge")
        org       = cd.get("organisation") or ""
        nom_dem   = cd.get("nom_demandeur") or ""
        email_dem = cd.get("email_demandeur") or ""
        date_r    = cd["date"]
        hd        = cd["heure_debut"]
        hf        = cd["heure_fin"]
        couverts  = cd.get("nombre_repas") or 0
        note      = cd.get("note") or ""
        action    = request.POST.get("action", "creer")
        forcer    = request.POST.get("forcer") == "on"

        # ── Cas particulier : cabinets de réflexion (par quantité 1/2/3) ─────
        if type_resa == "cabinet":
            nb = int(cd.get("nombre_cabinets") or 1)
            libres = _cabinets_libres(date_r, hd, hf)
            etat = _cabinets_etat(date_r, hd, hf)
            nb_libres = len(libres)
            conflits = []
            if nb_libres < nb:
                conflits = [f"Seulement {nb_libres} cabinet(s) libre(s) sur ce créneau "
                            f"(vous en demandez {nb})."]
            creneau = {'date': date_r, 'hd': hd, 'hf': hf,
                       'ressource': f"{nb} cabinet(s) de réflexion"}
            ctx_cab = {
                "form": form, "conflits": conflits, "alternatives": [],
                "dispo_verifiee": True, "creneau": creneau,
                "cabinets_libres_count": nb_libres, "cabinets_etat": etat,
            }
            if action == "verifier":
                return render(request, "administration/reservation_directe.html", ctx_cab)
            if not nom_dem or not email_dem:
                messages.error(request, "Le nom et l'email du demandeur sont requis pour créer la réservation.")
                return render(request, "administration/reservation_directe.html", ctx_cab)
            if conflits and not forcer:
                messages.warning(request, "Pas assez de cabinets libres : réservation non créée. "
                                          "Réduisez le nombre ou cochez « Forcer ».")
                return render(request, "administration/reservation_directe.html",
                              {**ctx_cab, "bloque": True})
            # Choix des cabinets : libres en priorité, puis occupés si on force
            cibles = list(libres)
            if len(cibles) < nb:
                occupes = [c for c in SalleReunion.objects.filter(
                    type_salle='cabinet_reflexion', actif=True).order_by('nom')
                    if c not in libres]
                cibles = (cibles + occupes)[:nb]
            else:
                cibles = cibles[:nb]
            crees = []
            for cab in cibles:
                rs = ReservationSalle.objects.create(
                    loge=loge, salle=cab, date=date_r, heure_debut=hd, heure_fin=hf,
                    statut="validee", nom_demandeur=nom_dem, email_demandeur=email_dem,
                    organisation=loge.nom if loge else org,
                    objet=note or "Cabinet de réflexion", nombre_cabinets=1, commentaire=note,
                )
                crees.append(rs)
            noms = ', '.join(c.salle.nom for c in crees)
            messages.success(request, f"{len(crees)} cabinet(s) de réflexion réservé(s) et validé(s) : {noms}.")
            log_evenement('creation_reservation_directe',
                f"Réservation directe cabinets : {loge.nom if loge else org} — "
                f"{date_r:%d/%m/%Y} {hd}–{hf} ({len(crees)} cabinet(s))",
                request=request, objet=crees[0] if crees else None)
            return redirect("administration:tableau_de_bord")

        # ── Congrès multi-temples ────────────────────────────────────────────
        if type_resa == "congres":
            temples_c = list(cd.get("temples_congres") or [])
            date_fin_c = cd.get("date_fin")
            grille = _grille_congres(temples_c, date_r, date_fin_c, hd, hf,
                                     salles=list(cd.get("salles_reunion") or []))
            conflits_exist = grille['nb_conflits'] > 0
            creneau = {'date': date_r, 'hd': hd, 'hf': hf, 'date_fin': date_fin_c,
                       'ressource': "Congrès — " + ", ".join(str(t) for t in temples_c)}
            ctx_c = {"form": form, "alternatives": [], "dispo_verifiee": True,
                     "creneau": creneau, "grille_congres": grille}
            if action == "verifier":
                return render(request, "administration/reservation_directe.html", ctx_c)
            if not nom_dem or not email_dem:
                messages.error(request, "Le nom et l'email du demandeur sont requis pour créer la réservation.")
                return render(request, "administration/reservation_directe.html", ctx_c)
            if conflits_exist and not forcer:
                messages.warning(request, "Conflit détecté sur le congrès : réservation non créée. "
                                          "Consultez la grille de disponibilité ou cochez « Forcer ».")
                return render(request, "administration/reservation_directe.html", {**ctx_c, "bloque": True})
            premier = None
            for tp in temples_c:
                r = Reservation.objects.create(
                    loge=loge, nom_organisation=org, temple=tp,
                    type_reservation='congres', sous_type='standard', statut='validee',
                    date=date_r, date_fin=date_fin_c, heure_debut=hd, heure_fin=hf,
                    besoin_agapes=couverts > 0, nombre_repas=couverts,
                    nom_demandeur=nom_dem, email_demandeur=email_dem, commentaire=note,
                )
                r.tarif = tarif_reservation(r)
                r.save(update_fields=['tarif'])
                if premier is None:
                    premier = r
            org_nom = loge.nom if loge else org
            for salle in cd.get('salles_reunion') or []:
                ReservationSalle.objects.create(
                    loge=loge, salle=salle, date=date_r, heure_debut=hd, heure_fin=hf,
                    statut='validee', nom_demandeur=nom_dem, email_demandeur=email_dem,
                    organisation=org_nom, objet='Congrès', nombre_participants=couverts,
                    commentaire=note,
                )
            messages.success(request, f"Congrès créé et validé sur {len(temples_c)} temple(s).")
            log_evenement('creation_reservation_directe',
                f"Congrès direct : {org_nom} — {date_r:%d/%m/%Y}"
                + (f" → {date_fin_c:%d/%m/%Y}" if date_fin_c else "")
                + f" ({len(temples_c)} temple(s))",
                request=request, objet=premier)
            return redirect("administration:tableau_de_bord")

        ressource = cd.get("temple") if type_resa == "exceptionnelle" else form.salle_choisie()
        type_dispo = 'temple' if type_resa == "exceptionnelle" else type_resa

        # ── Contrôle de disponibilité ───────────────────────────────────────
        conflits, alternatives = _analyser_disponibilite(type_dispo, ressource, date_r, hd, hf)
        dispo_verifiee = True
        creneau = {'date': date_r, 'hd': hd, 'hf': hf, 'ressource': ressource}

        occupants_rec = (_occupants_recurrents(ressource, date_r, hd, hf)
                         if type_dispo == 'temple' and conflits else [])

        ctx_dispo = {
            "form": form, "conflits": conflits, "alternatives": alternatives,
            "dispo_verifiee": dispo_verifiee, "creneau": creneau,
            "occupants_recurrents": occupants_rec,
            "echange_next": "administration:reservation_directe",
        }

        # Bouton « Vérifier la disponibilité » : on affiche le résultat sans créer.
        # Aucune saisie du demandeur n'est requise à ce stade.
        if action == "verifier":
            return render(request, "administration/reservation_directe.html", ctx_dispo)

        # À partir d'ici on crée : le demandeur devient obligatoire
        if not nom_dem or not email_dem:
            messages.error(request, "Le nom et l'email du demandeur sont requis pour créer la réservation.")
            return render(request, "administration/reservation_directe.html", ctx_dispo)

        # Blocage si conflit non forcé
        if conflits and not forcer:
            messages.warning(
                request,
                "Conflit détecté : la réservation n'a pas été créée. "
                "Choisissez une alternative ou cochez « Forcer malgré le conflit »."
            )
            return render(request, "administration/reservation_directe.html",
                          {**ctx_dispo, "bloque": True})

        if type_resa == "exceptionnelle":
            temple = cd["temple"]
            resa = Reservation.objects.create(
                loge=loge,
                nom_organisation=org,
                temple=temple,
                type_reservation="exceptionnelle",
                sous_type="standard",
                statut="validee",
                date=date_r,
                heure_debut=hd,
                heure_fin=hf,
                besoin_agapes=couverts > 0,
                nombre_repas=couverts,
                nom_demandeur=nom_dem,
                email_demandeur=email_dem,
                commentaire=note,
            )
            resa.tarif = tarif_reservation(resa)
            resa.save(update_fields=['tarif'])
            messages.success(request, "Réservation temple créée et validée.")
            log_evenement('creation_reservation_directe',
                f"Réservation directe temple : {loge or org} — {date_r:%d/%m/%Y} {hd}–{hf} ({temple})",
                request=request, objet=resa)
        else:
            salle = ressource
            objet_defaut = {
                "reunion": "Réunion",
                "cabinet": "Cabinet de réflexion",
                "agapes":  "Agapes",
            }.get(type_resa, "Réservation")
            resa_salle = ReservationSalle.objects.create(
                loge=loge,
                salle=salle,
                date=date_r,
                heure_debut=hd,
                heure_fin=hf,
                statut="validee",
                nom_demandeur=nom_dem,
                email_demandeur=email_dem,
                organisation=loge.nom if loge else org,
                objet=note or objet_defaut,
                nombre_participants=couverts,
                commentaire=note,
            )
            messages.success(request, "Réservation créée et validée.")
            log_evenement('creation_reservation_directe',
                f"Réservation directe {type_resa} : {loge.nom if loge else org} — {date_r:%d/%m/%Y} {hd}–{hf} ({salle})",
                request=request, objet=resa_salle)

        return redirect("administration:tableau_de_bord")

    return render(request, "administration/reservation_directe.html", {
        "form": form, "conflits": conflits, "alternatives": alternatives,
        "dispo_verifiee": dispo_verifiee, "creneau": creneau,
    })


# ── Journal de traçabilité ────────────────────────────────────────────────────

@staff_required
def journal(request):
    """Journal de traçabilité — accès staff uniquement."""
    from django.contrib.auth import get_user_model
    from django.core.paginator import Paginator

    if not request.user.is_staff:
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden("Accès réservé aux administrateurs.")

    User = get_user_model()

    qs = JournalEvenement.objects.select_related('utilisateur').order_by('-date_heure')

    # ── Filtres ───────────────────────────────────────────────────────────────
    f_type       = request.GET.get('type', '').strip()
    f_date_debut = request.GET.get('date_debut', '').strip()
    f_date_fin   = request.GET.get('date_fin', '').strip()
    f_loge       = request.GET.get('loge', '').strip()
    f_user       = request.GET.get('utilisateur', '').strip()

    if f_type:
        qs = qs.filter(type_evenement=f_type)
    if f_date_debut:
        try:
            from datetime import date as _date
            qs = qs.filter(date_heure__date__gte=_date.fromisoformat(f_date_debut))
        except ValueError:
            pass
    if f_date_fin:
        try:
            from datetime import date as _date
            qs = qs.filter(date_heure__date__lte=_date.fromisoformat(f_date_fin))
        except ValueError:
            pass
    if f_loge:
        qs = qs.filter(objet_type='loge', objet_id=f_loge)
    if f_user:
        qs = qs.filter(utilisateur_id=f_user)

    # ── Pagination ────────────────────────────────────────────────────────────
    paginator   = Paginator(qs, 50)
    page_number = request.GET.get('page', 1)
    page_obj    = paginator.get_page(page_number)

    # GET params sans 'page' (pour les liens de pagination)
    get_copy = request.GET.copy()
    get_copy.pop('page', None)
    get_params = get_copy.urlencode()

    # Données pour les selects de filtres
    types_choices    = JournalEvenement.TYPE_CHOICES
    loges_list       = Loge.objects.order_by('nom')
    utilisateurs_list = User.objects.filter(
        evenements_journal__isnull=False
    ).distinct().order_by('username')

    return render(request, 'administration/journal.html', {
        'page_obj':          page_obj,
        'paginator':         paginator,
        'get_params':        get_params,
        'types_choices':     types_choices,
        'loges_list':        loges_list,
        'utilisateurs_list': utilisateurs_list,
        'f_type':            f_type,
        'f_date_debut':      f_date_debut,
        'f_date_fin':        f_date_fin,
        'f_loge':            f_loge,
        'f_user':            f_user,
        'total':             qs.count(),
    })


# ── Gestion des accès portail ─────────────────────────────────────────────────

@staff_required
def portail_acces_admin(request):
    if request.method == 'POST':
        action  = request.POST.get('action')
        loge_id = request.POST.get('loge_id')
        loge    = get_object_or_404(Loge, pk=loge_id)

        if action == 'creer_acces':
            existant = DemandeAccesPortail.objects.filter(loge=loge, statut='validee').first()
            if existant:
                messages.info(request, f"Un accès actif existe déjà pour {loge.nom}.")
            else:
                DemandeAccesPortail.objects.create(
                    loge=loge,
                    nom_venerable='Administration',
                    email=loge.email or '',
                    statut='validee',
                )
                messages.success(request, f"Accès portail créé pour {loge.nom}.")
                log_evenement('creation_acces_portail',
                    f"Accès portail créé administrativement pour : {loge.nom}",
                    request=request, objet=loge)

        elif action == 'envoyer_lien':
            demande = DemandeAccesPortail.objects.filter(loge=loge, statut='validee').first()
            if not demande:
                messages.error(request, f"Aucun accès actif pour {loge.nom} — créez-le d'abord.")
            elif not loge.email:
                messages.error(request, f"La loge {loge.nom} n'a pas d'email renseigné.")
            else:
                lien = request.build_absolute_uri(f'/reservations/portail/{demande.token}/')
                send_mail_kellermann(
                    subject=f"[Kellermann] Votre lien d'accès au portail loge",
                    message=(
                        f"Bonjour,\n\n"
                        f"Vous trouverez ci-dessous le lien personnel d'accès au portail loge "
                        f"des Temples Kellermann pour la loge {loge.nom}.\n\n"
                        f"Lien d'accès :\n{lien}\n\n"
                        f"Ce lien est personnel et unique à votre loge. "
                        f"Il vous permet de consulter vos réservations, votre calendrier de saison "
                        f"et de valider vos tenues.\n\n"
                        f"En cas de problème, contactez l'administration.\n\n"
                        f"Fraternellement,\nL'administration des Temples Kellermann"
                    ),
                    recipient_list=[loge.email],
                )
                messages.success(request, f"Lien envoyé à {loge.email}.")
                log_evenement('envoi_lien_portail',
                    f"Lien portail envoyé à {loge.email} pour : {loge.nom}",
                    request=request, objet=loge)

        return redirect('administration:portail_acces_admin')

    # ── GET — liste ───────────────────────────────────────────────────────────
    from django.db.models import Prefetch

    loges = Loge.objects.filter(actif=True).order_by('nom').prefetch_related(
        Prefetch(
            'demandes_portail',
            queryset=DemandeAccesPortail.objects.filter(statut='validee').order_by('-created_at'),
            to_attr='acces_valides',
        )
    )

    all_data = []
    for loge in loges:
        demande = loge.acces_valides[0] if loge.acces_valides else None
        all_data.append({
            'loge':    loge,
            'demande': demande,
            'lien':    request.build_absolute_uri(f'/reservations/portail/{demande.token}/') if demande else None,
        })

    nb_avec      = sum(1 for d in all_data if d['demande'])
    nb_sans      = len(all_data) - nb_avec
    nb_sans_email = sum(1 for d in all_data if not d['loge'].email)

    filtre = request.GET.get('filtre', 'tous')
    if filtre == 'avec_acces':
        loges_data = [d for d in all_data if d['demande']]
    elif filtre == 'sans_acces':
        loges_data = [d for d in all_data if not d['demande']]
    elif filtre == 'sans_email':
        loges_data = [d for d in all_data if not d['loge'].email]
    else:
        loges_data = all_data

    return render(request, 'administration/portail_acces.html', {
        'loges_data':    loges_data,
        'filtre':        filtre,
        'nb_total':      len(all_data),
        'nb_avec':       nb_avec,
        'nb_sans':       nb_sans,
        'nb_sans_email': nb_sans_email,
    })


# ── Annonces / Pop-up ─────────────────────────────────────────────────────────

def _parse_dt_local(valeur):
    """Parse une valeur d'input datetime-local en datetime aware (ou None)."""
    if not valeur:
        return None
    dt = parse_datetime(valeur)
    if dt and timezone.is_naive(dt):
        dt = timezone.make_aware(dt, timezone.get_current_timezone())
    return dt


@staff_required
def annonces_liste(request):
    annonces = Annonce.objects.all()
    return render(request, 'administration/annonces_liste.html', {
        'annonces': annonces,
        'nb_annonces': annonces.count(),
    })


@staff_required
def annonce_form(request, pk=None):
    annonce = get_object_or_404(Annonce, pk=pk) if pk else None
    if request.method == 'POST':
        try:
            titre = request.POST.get('titre', '').strip()
            message = request.POST.get('message', '').strip()
            if not titre or not message:
                raise ValueError("Le titre et le message sont obligatoires.")
            data = {
                'titre': titre,
                'message': message,
                'niveau': request.POST.get('niveau', 'info'),
                'actif': request.POST.get('actif') == 'on',
                'date_debut': _parse_dt_local(request.POST.get('date_debut')),
                'date_fin': _parse_dt_local(request.POST.get('date_fin')),
                'duree_affichage': int(request.POST.get('duree_affichage') or 0),
            }
            if (data['date_debut'] and data['date_fin']
                    and data['date_fin'] < data['date_debut']):
                raise ValueError("La date de fin doit être postérieure à la date de début.")
            if annonce:
                for k, v in data.items():
                    setattr(annonce, k, v)
                annonce.save()
                messages.success(request, "Annonce modifiée.")
            else:
                Annonce.objects.create(**data)
                messages.success(request, "Annonce créée.")
            return redirect('administration:annonces_liste')
        except Exception as e:
            messages.error(request, f"Erreur : {e}")

    return render(request, 'administration/annonce_form.html', {
        'annonce': annonce,
        'niveaux': Annonce.NIVEAU_CHOICES,
    })


@staff_required
def annonce_toggle(request, pk):
    annonce = get_object_or_404(Annonce, pk=pk)
    annonce.actif = not annonce.actif
    annonce.save(update_fields=['actif', 'updated_at'])
    messages.success(
        request,
        "Annonce activée." if annonce.actif else "Annonce désactivée."
    )
    return redirect('administration:annonces_liste')


@staff_required
def annonce_supprimer(request, pk):
    annonce = get_object_or_404(Annonce, pk=pk)
    if request.method == 'POST':
        annonce.delete()
        messages.success(request, "Annonce supprimée.")
        return redirect('administration:annonces_liste')
    return render(request, 'administration/annonce_supprimer.html', {'annonce': annonce})


# ── Facturation (trésorier) ───────────────────────────────────────────────────

def _periode_facturation(request):
    """Période demandée (GET date_debut/date_fin), défaut = saison courante."""
    today = date.today()
    annee = today.year if today.month >= 9 else today.year - 1
    defaut_debut, defaut_fin = date(annee, 9, 1), date(annee + 1, 6, 30)

    def _p(val, defaut):
        try:
            return date.fromisoformat(val)
        except (ValueError, TypeError):
            return defaut
    return _p(request.GET.get('date_debut'), defaut_debut), _p(request.GET.get('date_fin'), defaut_fin)


def _facturation_data(date_debut, date_fin, params):
    """Facturation par JOURNÉE (un tarif par organisation et par jour, quel que
    soit le nombre de temples occupés), pour les occupations exceptionnelles et
    congrès validés (hors récurrentes), groupée par loge."""
    from decimal import Decimal

    def _nom(r):
        return r.loge.nom if r.loge else (r.nom_organisation or r.nom_demandeur or '—')

    base = Reservation.objects.filter(
        statut='validee',
        regle_source__isnull=True,   # jamais les réservations récurrentes
        date__gte=date_debut, date__lte=date_fin,
    ).select_related('loge', 'temple')

    items = {}   # clé -> entrée consolidée

    # Occupations exceptionnelles : consolidées par (organisation, jour)
    for r in base.filter(type_reservation='exceptionnelle').order_by('date'):
        tarif = tarif_reservation(r, params)
        if tarif <= 0:
            continue   # tenue de semaine, hors barème, ou antérieure au vote
        nom = _nom(r)
        key = ('exc', nom, r.date)
        e = items.get(key)
        if e is None:
            items[key] = {
                'nom': nom, 'date': r.date, 'date_fin': None,
                'temples': [str(r.temple)] if r.temple else [],
                'tarif': tarif, 'type': r.get_type_reservation_display(),
                'type_code': r.type_reservation, 'agapes': r.besoin_agapes,
            }
        else:
            if tarif > e['tarif']:
                e['tarif'] = tarif
            e['agapes'] = e['agapes'] or r.besoin_agapes
            if r.temple and str(r.temple) not in e['temples']:
                e['temples'].append(str(r.temple))

    # Congrès : consolidés par (organisation, date début, date fin) — un tarif
    # pour la durée, quel que soit le nombre de temples
    for r in base.filter(type_reservation='congres').order_by('date'):
        tarif = tarif_reservation(r, params)
        if tarif <= 0:
            continue   # congrès antérieur au vote des tarifs
        nom = _nom(r)
        key = ('cong', nom, r.date, r.date_fin)
        e = items.get(key)
        if e is None:
            items[key] = {
                'nom': nom, 'date': r.date, 'date_fin': r.date_fin,
                'temples': [str(r.temple)] if r.temple else [],
                'tarif': tarif, 'type': r.get_type_reservation_display(),
                'type_code': r.type_reservation, 'agapes': r.besoin_agapes,
            }
        else:
            if tarif > e['tarif']:
                e['tarif'] = tarif
            if r.temple and str(r.temple) not in e['temples']:
                e['temples'].append(str(r.temple))

    groupes = {}
    total = Decimal('0')
    nb_lignes = 0
    for e in items.values():
        g = groupes.setdefault(e['nom'], {'nom': e['nom'], 'lignes': [], 'total': Decimal('0')})
        g['lignes'].append({
            'date': e['date'], 'date_fin': e['date_fin'],
            'temple': ', '.join(e['temples']),
            'type': e['type'], 'type_code': e['type_code'],
            'agapes': e['agapes'], 'tarif': e['tarif'],
        })
        g['total'] += e['tarif']
        total += e['tarif']
        nb_lignes += 1
    for g in groupes.values():
        g['lignes'].sort(key=lambda x: x['date'])
    groupes_list = sorted(groupes.values(), key=lambda d: d['nom'].lower())
    return groupes_list, total, nb_lignes


@staff_required
def facturation(request):
    from decimal import Decimal, InvalidOperation
    params = Parametres.get_instance()

    if request.method == 'POST' and request.POST.get('action') == 'maj_tarifs':
        try:
            params.tarif_exc_sans_agapes = Decimal(request.POST.get('tarif_exc_sans_agapes') or '0')
            params.tarif_exc_avec_agapes = Decimal(request.POST.get('tarif_exc_avec_agapes') or '0')
            params.tarif_congres_jour    = Decimal(request.POST.get('tarif_congres_jour') or '0')
            params.tarif_funebre         = Decimal(request.POST.get('tarif_funebre') or '0')
            de = (request.POST.get('tarif_date_effet') or '').strip()
            params.tarif_date_effet = date.fromisoformat(de) if de else None
            params.facturation_active = 'facturation_active' in request.POST
            params.tarif_membre_loge = Decimal(request.POST.get('tarif_membre_loge') or '0')
            params.tarif_membre_hg = Decimal(request.POST.get('tarif_membre_hg') or '0')
            params.save(update_fields=['tarif_exc_sans_agapes', 'tarif_exc_avec_agapes',
                                       'tarif_congres_jour', 'tarif_funebre', 'tarif_date_effet',
                                       'facturation_active', 'tarif_membre_loge', 'tarif_membre_hg'])
            messages.success(request, "Tarifs mis à jour. Ils ne s'appliquent pas aux dates antérieures à leur entrée en vigueur.")
        except (InvalidOperation, ValueError):
            messages.error(request, "Valeurs invalides : vérifiez les montants et la date.")
        qs = request.META.get('QUERY_STRING', '')
        return redirect(f"{request.path}?{qs}" if qs else request.path)

    date_debut, date_fin = _periode_facturation(request)
    groupes, total, nb = _facturation_data(date_debut, date_fin, params)

    return render(request, 'administration/facturation.html', {
        'params': params,
        'date_debut': date_debut, 'date_fin': date_fin,
        'groupes': groupes, 'total_global': total, 'nb_jours': nb,
    })


@staff_required
def facturation_pdf(request):
    """Facture PDF : d'une loge (?nom=...) ou de toutes (une page par loge)."""
    from io import BytesIO
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer, PageBreak)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    params = Parametres.get_instance()
    date_debut, date_fin = _periode_facturation(request)
    groupes, total, nb = _facturation_data(date_debut, date_fin, params)
    nom_filtre = request.GET.get('nom')
    if nom_filtre:
        groupes = [g for g in groupes if g['nom'] == nom_filtre]

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=1.5 * cm,
                            bottomMargin=1.5 * cm, leftMargin=1.5 * cm, rightMargin=1.5 * cm)
    styles = getSampleStyleSheet()
    h = ParagraphStyle('h', parent=styles['Title'], fontSize=15, textColor=colors.HexColor('#0F2137'))
    lg = ParagraphStyle('lg', parent=styles['Heading2'], textColor=colors.HexColor('#0F2137'))
    sub = ParagraphStyle('sub', parent=styles['Normal'], fontSize=9, textColor=colors.grey)
    elems = []
    for gi, g in enumerate(groupes):
        if gi > 0:
            elems.append(PageBreak())
        elems.append(Paragraph("Facture — Temples Kellermann", h))
        elems.append(Paragraph(g['nom'], lg))
        elems.append(Paragraph(f"Période : {date_debut:%d/%m/%Y} → {date_fin:%d/%m/%Y}", sub))
        elems.append(Spacer(1, 0.4 * cm))
        data = [['Date', 'Type', 'Temple(s)', 'Agapes', 'Montant']]
        for l in g['lignes']:
            d = f"{l['date']:%d/%m/%Y}" + (f" → {l['date_fin']:%d/%m/%Y}" if l['date_fin'] else '')
            data.append([d, l['type'], l['temple'], 'Oui' if l['agapes'] else '—', f"{l['tarif']:.0f} €"])
        data.append(['', '', '', 'Total', f"{g['total']:.0f} €"])
        t = Table(data, colWidths=[3.4 * cm, 3.4 * cm, 5.4 * cm, 2 * cm, 2.5 * cm])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0F2137')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.grey),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#DBEAFE')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('ALIGN', (-1, 0), (-1, -1), 'RIGHT'),
        ]))
        elems.append(t)
        elems.append(Spacer(1, 0.3 * cm))
        elems.append(Paragraph("Montants hors prestations traiteur. Réf. tarifs votés en AG.", sub))
    if not groupes:
        elems.append(Paragraph("Aucune occupation facturable sur la période.", styles['Normal']))
    doc.build(elems)
    buf.seek(0)
    resp = HttpResponse(buf, content_type='application/pdf')
    fn = (nom_filtre or 'toutes').replace(' ', '_').replace('/', '-')[:40]
    resp['Content-Disposition'] = f'attachment; filename="Facture_{fn}_{date.today():%Y%m%d}.pdf"'
    return resp


@staff_required
def facturation_export_excel(request):
    params = Parametres.get_instance()
    date_debut, date_fin = _periode_facturation(request)
    groupes, total, nb = _facturation_data(date_debut, date_fin, params)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Facturation"
    navy = Font(bold=True, color="FFFFFF")
    navy_fill = PatternFill("solid", fgColor="0F2137")
    loge_fill = PatternFill("solid", fgColor="1E3A5F")
    total_fill = PatternFill("solid", fgColor="DBEAFE")
    bold = Font(bold=True)
    euro = '#,##0.00\\ €'

    ws.append([f"Récapitulatif de facturation — {date_debut:%d/%m/%Y} au {date_fin:%d/%m/%Y}"])
    ws.merge_cells('A1:E1')
    ws['A1'].font = Font(bold=True, size=13, color="0F2137")
    ws.append([])

    headers = ["Date", "Type", "Agapes", "Temple", "Tarif (€)"]
    for grp in groupes:
        ws.append([grp['nom']])
        r = ws.max_row
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        ws.cell(r, 1).font = navy
        ws.cell(r, 1).fill = loge_fill
        ws.append(headers)
        hr = ws.max_row
        for c in range(1, 6):
            ws.cell(hr, c).font = navy
            ws.cell(hr, c).fill = navy_fill
        for l in grp['lignes']:
            date_str = l['date'].strftime('%d/%m/%Y')
            if l.get('date_fin'):
                date_str += ' → ' + l['date_fin'].strftime('%d/%m/%Y')
            ws.append([
                date_str, l['type'],
                "Oui" if l['agapes'] else "Non", l['temple'], float(l['tarif']),
            ])
            ws.cell(ws.max_row, 5).number_format = euro
        ws.append(["", "", "", "Sous-total", float(grp['total'])])
        sr = ws.max_row
        ws.cell(sr, 4).font = bold
        ws.cell(sr, 5).font = bold
        ws.cell(sr, 5).number_format = euro
        ws.cell(sr, 5).fill = total_fill
        ws.append([])

    ws.append(["", "", "", "TOTAL GÉNÉRAL", float(total)])
    tr = ws.max_row
    ws.cell(tr, 4).font = Font(bold=True, size=12)
    ws.cell(tr, 5).font = Font(bold=True, size=12)
    ws.cell(tr, 5).number_format = euro
    ws.cell(tr, 5).fill = total_fill

    widths = [14, 28, 9, 20, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = (
        f'attachment; filename="facturation_{date_debut}_{date_fin}.xlsx"')
    wb.save(response)
    return response


# ── Suivi des loges par saison (statut Active / À reconfirmer / Inactive) ──────

@staff_required
def loges_saison(request):
    """Écran de suivi : loges sans récurrence / à reconfirmer, avec gestion du statut."""
    from django.db.models import Count, Q as _Q

    if request.method == 'POST':
        if request.POST.get('action') == 'reset_all':
            n = Loge.objects.filter(statut='active').update(statut='a_reconfirmer')
            messages.warning(request, f"{n} loge(s) actives repassées en « À reconfirmer ». "
                                      "Importez les fiches reçues pour les réactiver.")
            return redirect(f"{request.path}?{request.GET.urlencode()}")
        loge = get_object_or_404(Loge, pk=request.POST.get('loge_id'))
        nouveau = request.POST.get('statut')
        if nouveau in ('active', 'a_reconfirmer', 'inactive'):
            loge.statut = nouveau
            loge.actif = (nouveau != 'inactive')
            loge.save(update_fields=['statut', 'actif'])
            messages.success(request, f"{loge.nom} → « {loge.get_statut_display()} ».")
        return redirect(f"{request.path}?{request.GET.urlencode()}")

    filtre = request.GET.get('filtre', 'tous')
    loges = (Loge.objects.exclude(statut='inactive')
             .select_related('obedience')
             .annotate(nb_regles=Count('regles', filter=_Q(regles__actif=True)))
             .order_by('nom'))
    if filtre == 'sans_recurrence':
        loges = loges.filter(nb_regles=0)
    elif filtre == 'a_reconfirmer':
        loges = loges.filter(statut='a_reconfirmer')
    elif filtre == 'active':
        loges = loges.filter(statut='active')

    inactives = Loge.objects.filter(statut='inactive').order_by('nom')

    # Compteurs (sur l'ensemble, hors filtre courant)
    base = Loge.objects.exclude(statut='inactive').annotate(
        nb_regles=Count('regles', filter=_Q(regles__actif=True)))
    nb_sans = base.filter(nb_regles=0).count()
    nb_a_reconfirmer = Loge.objects.filter(statut='a_reconfirmer').count()

    return render(request, 'administration/loges_saison.html', {
        'loges': loges, 'inactives': inactives, 'filtre': filtre,
        'nb_sans': nb_sans, 'nb_a_reconfirmer': nb_a_reconfirmer,
        'nb_inactives': inactives.count(),
    })


# ── Règles de récurrence — Salles ─────────────────────────────────────────────

def _regle_salle_form_class():
    """Retourne la classe Form pour une RegleRecurrenceSalle (définie ici pour éviter un forms.py séparé)."""
    from django import forms as _forms

    class RegleSalleForm(_forms.ModelForm):
        class Meta:
            model  = RegleRecurrenceSalle
            fields = [
                'loge', 'salles', 'jour_semaine', 'numero_semaine',
                'heure_debut', 'heure_fin', 'mois_actifs',
                'actif', 'date_debut', 'date_fin',
                'objet', 'nombre_participants',
            ]
            widgets = {
                'loge':           _forms.Select(attrs={'class': 'form-select'}),
                'salles':         _forms.CheckboxSelectMultiple(),
                'jour_semaine':   _forms.Select(attrs={'class': 'form-select no-select2'}),
                'numero_semaine': _forms.Select(attrs={'class': 'form-select no-select2'}),
                'heure_debut':    _forms.TimeInput(attrs={'class': 'form-control', 'type': 'time'}),
                'heure_fin':      _forms.TimeInput(attrs={'class': 'form-control', 'type': 'time'}),
                'mois_actifs':    _forms.CheckboxSelectMultiple(
                    choices=[
                        (1,'Janvier'),(2,'Février'),(3,'Mars'),(4,'Avril'),
                        (5,'Mai'),(6,'Juin'),(7,'Juillet'),(8,'Août'),
                        (9,'Septembre'),(10,'Octobre'),(11,'Novembre'),(12,'Décembre'),
                    ]
                ),
                'date_debut':     _forms.DateInput(attrs={'class': 'form-control', 'type': 'date'}),
                'date_fin':       _forms.DateInput(attrs={'class': 'form-control', 'type': 'date'}),
                'objet':          _forms.TextInput(attrs={'class': 'form-control'}),
                'nombre_participants': _forms.NumberInput(attrs={'class': 'form-control'}),
            }

        def clean(self):
            cleaned = super().clean()
            hd = cleaned.get('heure_debut')
            hf = cleaned.get('heure_fin')
            if hd and hf and hf <= hd:
                self.add_error('heure_fin', "L'heure de fin doit être après l'heure de début.")
            return cleaned

    return RegleSalleForm


@staff_required
def regles_salle_liste(request):
    """Liste de toutes les règles de récurrence pour les salles."""
    regles = (
        RegleRecurrenceSalle.objects
        .select_related('loge')
        .prefetch_related('salles')
        .order_by('loge__nom', 'jour_semaine', 'numero_semaine')
    )
    return render(request, 'administration/regles_salle_liste.html', {'regles': regles})


@staff_required
def regle_salle_form(request, pk=None):
    """Créer ou modifier une règle de récurrence salle."""
    RegleSalleForm = _regle_salle_form_class()
    instance = get_object_or_404(RegleRecurrenceSalle, pk=pk) if pk else None
    form = RegleSalleForm(request.POST or None, instance=instance)

    if request.method == 'POST' and form.is_valid():
        form.save()
        verb = "modifiée" if pk else "créée"
        messages.success(request, f"Règle {verb} avec succès.")
        return redirect('administration:regles_salle_liste')

    return render(request, 'administration/regle_salle_form.html', {
        'form': form, 'instance': instance,
    })


@staff_required
def regle_salle_supprimer(request, pk):
    """Supprimer une règle de récurrence salle."""
    regle = get_object_or_404(RegleRecurrenceSalle, pk=pk)
    if request.method == 'POST':
        nom = str(regle)
        regle.delete()
        messages.success(request, f"Règle supprimée : {nom}.")
        return redirect('administration:regles_salle_liste')
    return render(request, 'administration/regle_salle_supprimer.html', {'regle': regle})


@staff_required
def regenerer_salles(request):
    """Génère les ReservationSalle depuis les RegleRecurrenceSalle actives."""
    if request.method != 'POST':
        annee_def = date.today().year if date.today().month >= 9 else date.today().year - 1
        return render(request, 'administration/regenerer_salles.html', {
            'annees': list(range(annee_def - 1, annee_def + 3)),
            'annee_def': annee_def,
            'regles': RegleRecurrenceSalle.objects.filter(actif=True).select_related('loge').prefetch_related('salles'),
        })

    import uuid as uuid_module
    annee  = int(request.POST.get('annee', date.today().year))
    mode   = request.POST.get('mode', 'ajouter')  # ajouter | remplacer
    d1, d2 = date(annee, 9, 1), date(annee + 1, 6, 30)

    regles = RegleRecurrenceSalle.objects.filter(actif=True).prefetch_related('salles')

    cree = 0
    for regle in regles:
        salles_list = list(regle.salles.all())
        if not salles_list:
            continue

        dates_saison = [
            d for d in (
                _calculer_dates_regle(regle, annee) +
                _calculer_dates_regle(regle, annee + 1)
            )
            if d1 <= d <= d2 and d.month not in [7, 8]
            and not (regle.date_fin and d > regle.date_fin)
            and not (regle.date_debut and d < regle.date_debut)
        ]

        if mode == 'remplacer':
            ReservationSalle.objects.filter(
                regle_source=regle,
                date__gte=d1, date__lte=d2,
            ).delete()

        exclues = set(regle.dates_exclues or [])
        for d in dates_saison:
            if d.isoformat() in exclues:
                continue
            if ReservationSalle.objects.filter(regle_source=regle, date=d).exists():
                continue

            group_id = uuid_module.uuid4()
            nom = regle.loge.nom if regle.loge else "—"
            for salle in salles_list:
                ReservationSalle.objects.create(
                    loge=regle.loge,
                    salle=salle,
                    date=d,
                    heure_debut=regle.heure_debut,
                    heure_fin=regle.heure_fin,
                    statut='validee',
                    nom_demandeur=nom,
                    email_demandeur='admin@kellermann.local',
                    organisation=nom,
                    objet=regle.objet or 'Réunion',
                    nombre_participants=regle.nombre_participants or 0,
                    facturable=False,
                    group_uuid=group_id,
                    regle_source=regle,
                )
                cree += 1

    messages.success(request, f"{cree} réservation(s) salle créée(s) pour la saison {annee}/{annee+1}.")
    return redirect('administration:regles_salle_liste')


# ── Réservation multi-salles (depuis l'admin) ─────────────────────────────────

@staff_required
def reserver_multi_salles(request):
    """Réservation directe de plusieurs salles simultanées pour une loge (ex : temple + salles complémentaires)."""
    import uuid as uuid_module

    loges = Loge.objects.filter(actif=True).order_by('nom')
    salles_par_type = {}
    for salle in SalleReunion.objects.filter(actif=True).order_by('type_salle', 'nom'):
        label = salle.get_type_salle_display()
        salles_par_type.setdefault(label, []).append(salle)

    if request.method == 'POST':
        loge_pk  = request.POST.get('loge_pk', '').strip()
        org      = request.POST.get('organisation', '').strip()
        salle_pks = request.POST.getlist('salles')
        date_str  = request.POST.get('date', '').strip()
        hd_str    = request.POST.get('heure_debut', '').strip()
        hf_str    = request.POST.get('heure_fin', '').strip()
        objet     = request.POST.get('objet', '').strip() or 'Réservation groupée'
        participants = request.POST.get('nombre_participants', '0').strip()
        facturable_val = request.POST.get('facturable') == 'on'

        errors = []
        loge = None
        if loge_pk:
            try:
                loge = Loge.objects.get(pk=int(loge_pk))
            except (Loge.DoesNotExist, ValueError):
                errors.append("Loge introuvable.")
        if not loge and not org:
            errors.append("Indiquez une loge ou un nom d'organisation.")
        if not salle_pks:
            errors.append("Sélectionnez au moins une salle.")
        try:
            from datetime import date as date_cls, time as time_cls
            date_r = date_cls.fromisoformat(date_str)
        except (ValueError, TypeError):
            date_r = None
            errors.append("Date invalide.")
        try:
            hd = time_cls.fromisoformat(hd_str)
            hf = time_cls.fromisoformat(hf_str)
            if hf <= hd:
                errors.append("L'heure de fin doit être après l'heure de début.")
        except (ValueError, TypeError):
            hd = hf = None
            errors.append("Horaires invalides.")

        salles_sel = list(SalleReunion.objects.filter(pk__in=salle_pks, actif=True))

        if not errors and date_r and hd and hf and salles_sel:
            nom      = loge.nom if loge else org
            try:
                nb_part = int(participants)
            except ValueError:
                nb_part = 0
            group_id = uuid_module.uuid4()
            for salle in salles_sel:
                ReservationSalle.objects.create(
                    loge=loge, salle=salle, date=date_r,
                    heure_debut=hd, heure_fin=hf, statut='validee',
                    nom_demandeur=nom,
                    email_demandeur='admin@kellermann.local',
                    organisation=nom,
                    objet=objet,
                    nombre_participants=nb_part,
                    facturable=facturable_val,
                    group_uuid=group_id,
                )
            noms = ', '.join(s.nom for s in salles_sel)
            messages.success(request, f"{len(salles_sel)} salle(s) réservée(s) ({noms}) le {date_r:%d/%m/%Y}.")
            log_evenement('reservation_directe',
                f"Multi-salles : {nom} — {date_r:%d/%m/%Y} {hd:%H:%M}–{hf:%H:%M} — {noms}",
                request=request)
            return redirect('administration:tableau_de_bord')

        for err in errors:
            messages.error(request, err)

    return render(request, 'administration/reserver_multi_salles.html', {
        'loges': loges,
        'salles_par_type': salles_par_type,
    })


# ── Validation demandes récurrence salle (portail loge) ──────────────────────

@staff_required
def valider_demande_recurrence_salle(request, pk):
    """L'admin valide ou refuse une demande de règle de récurrence salle soumise depuis le portail."""
    from temple_project.apps.administration.email_utils import send_mail_kellermann, get_email_admin
    demande = get_object_or_404(DemandeRegleRecurrenceSalle, pk=pk)

    if request.method == 'POST':
        action            = request.POST.get('action')
        commentaire_admin = request.POST.get('commentaire_admin', '').strip()

        if action not in ('valider', 'refuser'):
            messages.error(request, "Action invalide.")
            return redirect('administration:tableau_de_bord')

        demande.statut           = 'validee' if action == 'valider' else 'refusee'
        demande.commentaire_admin = commentaire_admin
        demande.save()

        if action == 'valider':
            # Créer la RegleRecurrenceSalle correspondante
            regle = RegleRecurrenceSalle.objects.create(
                loge=demande.loge,
                jour_semaine=demande.jour_semaine,
                numero_semaine=demande.numero_semaine,
                heure_debut=demande.heure_debut,
                heure_fin=demande.heure_fin,
                mois_actifs=demande.mois_actifs,
                objet=demande.objet,
                nombre_participants=demande.nombre_participants,
                actif=True,
            )
            regle.salles.set(demande.salles.all())
            demande.regle_creee = regle
            demande.save(update_fields=['regle_creee'])

            noms_salles = ', '.join(s.nom for s in demande.salles.all())
            send_mail_kellermann(
                subject="[Kellermann] Votre demande de récurrence salle a été validée",
                message=(
                    f"Bonjour {demande.nom_demandeur},\n\n"
                    f"Votre demande de règle de récurrence salle a été validée.\n\n"
                    f"Salle(s)  : {noms_salles}\n"
                    f"Fréquence : {regle.get_numero_semaine_display()} {regle.get_jour_semaine_display()}\n"
                    f"Horaires  : {regle.heure_debut:%H:%M} – {regle.heure_fin:%H:%M}\n"
                    f"Objet     : {regle.objet}\n"
                    + (f"\nCommentaire : {commentaire_admin}\n" if commentaire_admin else "")
                    + f"\nLa règle sera appliquée lors de la prochaine génération des réservations.\n\n"
                    f"Fraternellement,\nL'administration des Temples Kellermann"
                ),
                recipient_list=[demande.email_demandeur],
            )
            log_evenement('validation_regle_salle',
                f"Règle récurrence salle validée : {demande.loge} — {regle.get_numero_semaine_display()} {regle.get_jour_semaine_display()} — {noms_salles}",
                request=request, objet=regle)
            messages.success(request, f"Demande validée — règle de récurrence créée pour {demande.loge}.")
        else:
            send_mail_kellermann(
                subject="[Kellermann] Votre demande de récurrence salle",
                message=(
                    f"Bonjour {demande.nom_demandeur},\n\n"
                    f"Votre demande de règle de récurrence salle n'a pas pu être accordée.\n\n"
                    + (f"Motif : {commentaire_admin}\n\n" if commentaire_admin else "")
                    + f"Pour toute question, contactez l'administration.\n\n"
                    f"Fraternellement,\nL'administration des Temples Kellermann"
                ),
                recipient_list=[demande.email_demandeur],
            )
            messages.warning(request, f"Demande refusée — email envoyé à {demande.email_demandeur}.")

        return redirect('administration:tableau_de_bord')

    return render(request, 'administration/valider_demande_recurrence_salle.html', {
        'demande': demande,
    })
