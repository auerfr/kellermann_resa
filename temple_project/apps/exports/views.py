import csv
import io
import json
from datetime import date
from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.utils import timezone
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import timedelta
from temple_project.apps.reservations.models import (
    Reservation, Temple, DemandeAccesPortail, ValidationSaison, ReservationSalle,
)


@login_required
def export_csv(request):
    """Export CSV des réservations selon filtres"""
    qs = _get_queryset_from_request(request)
    response = HttpResponse(content_type="text/csv; charset=utf-8-sig")
    response["Content-Disposition"] = 'attachment; filename="reservations.csv"'

    writer = csv.writer(response)
    writer.writerow([
        "Date", "Heure début", "Heure fin", "Loge", "Obédience",
        "Temple", "Type", "Sous-type", "Statut", "Agapes", "Nb repas",
        "Demandeur", "Email"
    ])
    for r in qs:
        writer.writerow([
            r.date, r.heure_debut, r.heure_fin,
            r.loge.nom if r.loge else (r.nom_organisation or r.nom_demandeur or ''),
            r.loge.obedience.nom if r.loge and r.loge.obedience else '',
            str(r.temple) if r.temple else '',
            r.get_type_reservation_display(), r.get_sous_type_display(),
            r.get_statut_display(), "Oui" if r.besoin_agapes else "Non",
            r.nombre_repas, r.nom_demandeur, r.email_demandeur,
        ])
    return response


@login_required
def export_excel(request):
    """Export Excel avec mise en forme"""
    qs = _get_queryset_from_request(request)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Réservations"

    headers = [
        "Date", "Heure début", "Heure fin", "Loge", "Obédience",
        "Temple", "Type", "Sous-type", "Statut", "Agapes", "Nb repas",
    ]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2C3E50")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    STATUT_COLORS = {
        "validee": "C8E6C9",
        "attente": "FFF9C4",
        "refusee": "FFCDD2",
    }

    for row_idx, r in enumerate(qs, 2):
        data = [
            r.date, str(r.heure_debut), str(r.heure_fin),
            r.loge.nom if r.loge else (r.nom_organisation or r.nom_demandeur or ''),
            r.loge.obedience.nom if r.loge and r.loge.obedience else '',
            str(r.temple) if r.temple else '',
            r.get_type_reservation_display(), r.get_sous_type_display(),
            r.get_statut_display(), "Oui" if r.besoin_agapes else "Non",
            r.nombre_repas,
        ]
        fill_color = STATUT_COLORS.get(r.statut, "FFFFFF")
        fill = PatternFill("solid", fgColor=fill_color)
        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = fill

    # Ajustement largeur colonnes
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="reservations.xlsx"'
    return response


def _saison_courante(request):
    current_year = date.today().year
    default_saison = current_year if date.today().month >= 9 else current_year - 1
    try:
        annee = int(request.GET.get("annee", default_saison))
    except (ValueError, TypeError):
        annee = default_saison
    return annee, default_saison


def _compute_stats(annee_saison):
    """Données analytiques d'une saison : loges + réservations + séries pour graphiques.
    Partagé par la page Statistiques et son export."""
    from django.db.models import Count, Sum, Avg, Q
    from temple_project.apps.loges.models import Loge
    from temple_project.apps.reservations.models import RegleRecurrence

    d1, d2 = date(annee_saison, 9, 1), date(annee_saison + 1, 6, 30)
    reservations = Reservation.objects.filter(date__gte=d1, date__lte=d2)

    total    = reservations.count()
    validees = reservations.filter(statut="validee").count()
    attente  = reservations.filter(statut="attente").count()
    refusees = reservations.filter(statut="refusee").count()
    total_repas = reservations.filter(besoin_agapes=True, statut="validee").aggregate(
        s=Sum("nombre_repas"))["s"] or 0
    stats = {
        "total": total, "validees": validees, "attente": attente, "refusees": refusees,
        "total_repas": total_repas,
        "taux_validation": round(validees / total * 100, 1) if total else 0,
    }

    resa_par_obedience = list(reservations.values('loge__obedience__nom')
        .annotate(nb_reservations=Count('id')).order_by('-nb_reservations')[:15])

    resa_par_mois = []
    for m in [9, 10, 11, 12, 1, 2, 3, 4, 5, 6]:
        ya = annee_saison if m >= 9 else annee_saison + 1
        resa_par_mois.append({'mois': f'{ya}-{m:02d}',
                              'count': reservations.filter(date__year=ya, date__month=m).count()})

    resa_par_temple = list(reservations.values('temple__nom')
        .annotate(nb_reservations=Count('id')).order_by('-nb_reservations'))

    actives = Loge.objects.exclude(statut='inactive')
    loges_par_obedience = list(actives.values('obedience__nom').annotate(
        n=Count('id'),
        loges=Count('id', filter=Q(type_loge='loge')),
        hg=Count('id', filter=Q(type_loge='haut_grade')),
        eff=Sum('effectif_total'),
    ).order_by('-n'))
    loges = {
        'nb_total': actives.count(),
        'nb_loges': actives.filter(type_loge='loge').count(),
        'nb_hg': actives.filter(type_loge='haut_grade').count(),
        'nb_active': actives.filter(statut='active').count(),
        'nb_reconf': actives.filter(statut='a_reconfirmer').count(),
        'nb_inactive': Loge.objects.filter(statut='inactive').count(),
        'effectif_total': actives.aggregate(s=Sum('effectif_total'))['s'] or 0,
        'agapes_moy': round(actives.filter(effectif_moyen_agapes__gt=0)
                            .aggregate(a=Avg('effectif_moyen_agapes'))['a'] or 0),
        'nb_avec_effectif': actives.filter(effectif_total__gt=0).count(),
        'nb_regles': RegleRecurrence.objects.filter(actif=True).count(),
    }
    return {
        'stats': stats, 'resa_par_obedience': resa_par_obedience,
        'resa_par_mois': resa_par_mois, 'resa_par_temple': resa_par_temple,
        'loges': loges, 'loges_par_obedience': loges_par_obedience,
    }


# Sections proposées à l'export (clé, libellé)
SECTIONS_STATS = [
    ('synthese',        'Chiffres clés'),
    ('effectifs',       'Effectifs'),
    ('loges_obedience', 'Loges par obédience'),
    ('resa_statut',     'Réservations par statut'),
    ('resa_mois',       'Réservations par mois'),
    ('resa_temple',     'Réservations par temple'),
    ('resa_obedience',  'Réservations par obédience'),
]
_MOIS_FR = {1:'Janvier',2:'Février',3:'Mars',4:'Avril',5:'Mai',6:'Juin',
            7:'Juillet',8:'Août',9:'Septembre',10:'Octobre',11:'Novembre',12:'Décembre'}


def _section_rows(key, d):
    """(titre, entêtes, lignes) pour une section, ou None si clé inconnue."""
    s, L = d['stats'], d['loges']
    if key == 'synthese':
        return ('Chiffres clés', ['Indicateur', 'Valeur'], [
            ['Structures actives', L['nb_total']],
            ['Loges bleues', L['nb_loges']],
            ['Hauts grades', L['nb_hg']],
            ['Actives confirmées', L['nb_active']],
            ['À reconfirmer', L['nb_reconf']],
            ['Inactives', L['nb_inactive']],
            ['Règles de récurrence', L['nb_regles']],
            ['Réservations (saison)', s['total']],
            ['— validées', s['validees']],
            ['— en attente', s['attente']],
            ['— refusées', s['refusees']],
            ['Taux de validation', f"{s['taux_validation']} %"],
            ['Repas agapes (validés)', s['total_repas']],
        ])
    if key == 'effectifs':
        return ('Effectifs', ['Indicateur', 'Valeur'], [
            ['Effectif total (renseigné)', L['effectif_total']],
            ['Loges avec effectif', L['nb_avec_effectif']],
            ['Moyenne agapes / loge', L['agapes_moy']],
        ])
    if key == 'loges_obedience':
        rows = [[o['obedience__nom'] or '—', o['n'], o['loges'], o['hg'], o['eff'] or 0]
                for o in d['loges_par_obedience']]
        return ('Loges par obédience', ['Obédience', 'Total', 'Loges', 'Hauts grades', 'Effectif'], rows)
    if key == 'resa_statut':
        return ('Réservations par statut', ['Statut', 'Nombre'], [
            ['Validées', s['validees']], ['En attente', s['attente']],
            ['Refusées', s['refusees']], ['Total', s['total']],
        ])
    if key == 'resa_mois':
        rows = []
        for it in d['resa_par_mois']:
            y, m = it['mois'].split('-')
            rows.append([f"{_MOIS_FR[int(m)]} {y}", it['count']])
        return ('Réservations par mois', ['Mois', 'Nombre'], rows)
    if key == 'resa_temple':
        rows = [[t['temple__nom'] or 'Non renseigné', t['nb_reservations']] for t in d['resa_par_temple']]
        return ('Réservations par temple', ['Temple', 'Nombre'], rows)
    if key == 'resa_obedience':
        rows = [[o['loge__obedience__nom'] or 'Non renseignée', o['nb_reservations']]
                for o in d['resa_par_obedience']]
        return ('Réservations par obédience', ['Obédience', 'Réservations'], rows)
    return None


@login_required
def reporting(request):
    """Page unique Statistiques : synthèse des loges + activité (réservations) + graphiques."""
    annee_saison, default_saison = _saison_courante(request)
    d = _compute_stats(annee_saison)
    context = {
        "annee": annee_saison,
        "annee_courante": default_saison,
        "saison_label": f"{annee_saison}–{annee_saison + 1}",
        "stats": d['stats'],
        "reservations_par_obedience": d['resa_par_obedience'],
        "reservations_par_mois": json.dumps(d['resa_par_mois']),
        "reservations_par_temple": json.dumps([
            {'nom': t['temple__nom'] or 'Non renseigné', 'nb_reservations': t['nb_reservations']}
            for t in d['resa_par_temple']]),
        "temples": Temple.objects.all().order_by('nom'),
        "loges": d['loges'],
        "loges_par_obedience": d['loges_par_obedience'],
        "loges_tiles": [
            (d['loges']['nb_loges'],        "Loges bleues"),
            (d['loges']['nb_hg'],           "Hauts grades"),
            (d['loges']['nb_regles'],       "Règles actives"),
            (d['loges']['effectif_total'],  "Effectif total"),
        ],
        "sections_stats": SECTIONS_STATS,
    }
    return render(request, "exports/reporting.html", context)


@login_required
def statistiques_export(request):
    """Export sélectionnable (PDF ou Excel) des sections cochées, pour la saison."""
    annee, _ = _saison_courante(request)
    fmt = request.GET.get('format', 'pdf')
    sections = request.GET.getlist('sections') or [k for k, _ in SECTIONS_STATS]
    d = _compute_stats(annee)
    blocs = [b for b in (_section_rows(k, d) for k in sections) if b]
    label = f"{annee}-{annee + 1}"
    if fmt == 'excel':
        return _stats_export_excel(blocs, label)
    return _stats_export_pdf(blocs, label, d)


def _stats_export_excel(blocs, label):
    from openpyxl.utils import get_column_letter
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Statistiques"
    hf = Font(bold=True, color="FFFFFF"); hfill = PatternFill("solid", fgColor="0F2137")
    r = 1
    ws.cell(r, 1, f"Statistiques — saison {label}").font = Font(bold=True, color="0F2137", size=14)
    r += 2
    for titre, headers, rows in blocs:
        ws.cell(r, 1, titre).font = Font(bold=True, color="0F2137", size=11); r += 1
        for c, h in enumerate(headers, 1):
            cell = ws.cell(r, c, h); cell.font = hf; cell.fill = hfill
        r += 1
        for row in rows:
            for c, v in enumerate(row, 1):
                ws.cell(r, c, v)
            r += 1
        r += 1
    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 26
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    resp = HttpResponse(buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp['Content-Disposition'] = f'attachment; filename="statistiques_kellermann_{label}.xlsx"'
    return resp


_PDF_PALETTE = ['#0F2137', '#0E7C7B', '#C8A84B', '#7C5CBF', '#B4531F', '#2563EB', '#0E9F6E', '#9CA3AF']


def _pdf_table_style(navy, gold):
    from reportlab.lib import colors
    from reportlab.platypus import TableStyle
    return TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), navy),
        ('TEXTCOLOR', (0, 0), (-1, 0), gold),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 6), ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 4), ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ])


def _stats_export_pdf(blocs, label, d):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.graphics.shapes import Drawing
    from reportlab.graphics.charts.barcharts import VerticalBarChart
    from reportlab.graphics.charts.piecharts import Pie
    from reportlab.graphics.charts.legends import Legend

    navy = colors.HexColor('#0F2137'); gold = colors.HexColor('#C8A84B')
    pal = [colors.HexColor(c) for c in _PDF_PALETTE]
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle('h1', parent=styles['Title'], textColor=navy, fontSize=18, alignment=0)
    h2 = ParagraphStyle('h2', parent=styles['Heading2'], textColor=navy, fontSize=12)
    small = ParagraphStyle('sm', parent=styles['Normal'], textColor=colors.HexColor('#64748B'), fontSize=9)

    def _chart_mois():
        dr = Drawing(460, 155)
        bc = VerticalBarChart()
        bc.x, bc.y, bc.width, bc.height = 22, 24, 430, 116
        bc.data = [[it['count'] for it in d['resa_par_mois']]]
        bc.categoryAxis.categoryNames = [_MOIS_FR[int(it['mois'].split('-')[1])][:4].lower() for it in d['resa_par_mois']]
        bc.categoryAxis.labels.fontSize = 8
        bc.categoryAxis.labels.fillColor = colors.HexColor('#64748B')
        bc.valueAxis.valueMin = 0
        bc.valueAxis.labels.fontSize = 8
        bc.valueAxis.labels.fillColor = colors.HexColor('#94A3B8')
        bc.valueAxis.strokeColor = colors.HexColor('#E2E8F0')
        bc.categoryAxis.strokeColor = colors.HexColor('#E2E8F0')
        bc.bars[0].fillColor = gold
        bc.bars[0].strokeColor = None
        bc.barWidth = 6
        dr.add(bc)
        return dr

    def _chart_temples():
        vals = [t['nb_reservations'] for t in d['resa_par_temple']]
        labs = [(t['temple__nom'] or '—').replace('Temple ', '') for t in d['resa_par_temple']]
        dr = Drawing(460, 150)
        pie = Pie()
        pie.x, pie.y, pie.width, pie.height = 15, 15, 120, 120
        pie.data = vals or [1]
        pie.labels = None
        pie.slices.strokeColor = colors.white
        pie.slices.strokeWidth = 1.5
        for i in range(len(vals)):
            pie.slices[i].fillColor = pal[i % len(pal)]
        dr.add(pie)
        leg = Legend()
        leg.x, leg.y = 175, 120
        leg.dx, leg.dy, leg.dxTextSpace = 8, 8, 6
        leg.fontName, leg.fontSize, leg.deltay = 'Helvetica', 9, 15
        leg.colorNamePairs = [(pal[i % len(pal)], f"{labs[i]}  ({vals[i]})") for i in range(len(vals))]
        dr.add(leg)
        return dr

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=1.4 * cm, bottomMargin=1.4 * cm,
                            leftMargin=1.5 * cm, rightMargin=1.5 * cm, title=f"Statistiques {label}")
    elems = [Paragraph(f"Statistiques — saison {label}", h1),
             Paragraph(f"Édité le {timezone.localtime().strftime('%d/%m/%Y à %H:%M')}", small),
             Spacer(1, 0.4 * cm)]

    # Graphiques en tête (option 1)
    if any(it['count'] for it in d['resa_par_mois']):
        elems += [Paragraph("Réservations par mois", h2), _chart_mois(), Spacer(1, 0.3 * cm)]
    if d['resa_par_temple']:
        elems += [Paragraph("Répartition par temple", h2), _chart_temples(), Spacer(1, 0.35 * cm)]

    table_style = _pdf_table_style(navy, gold)
    for titre, headers, rows in blocs:
        elems.append(Paragraph(titre, h2)); elems.append(Spacer(1, 0.12 * cm))
        data = [headers] + [[str(x) for x in row] for row in rows]
        t = Table(data, repeatRows=1, hAlign='LEFT')
        t.setStyle(table_style)
        elems.append(t); elems.append(Spacer(1, 0.4 * cm))
    doc.build(elems)
    buf.seek(0)
    resp = HttpResponse(buf.read(), content_type='application/pdf')
    resp['Content-Disposition'] = f'attachment; filename="statistiques_kellermann_{label}.pdf"'
    return resp


def _get_queryset_from_request(request):
    qs = Reservation.objects.select_related("loge", "loge__obedience", "temple")
    if request.GET.get("mois"):
        qs = qs.filter(date__month=request.GET["mois"])
    if request.GET.get("annee"):
        qs = qs.filter(date__year=request.GET["annee"])
    if request.GET.get("temple"):
        qs = qs.filter(temple_id=request.GET["temple"])
    if request.GET.get("loge"):
        qs = qs.filter(loge_id=request.GET["loge"])
    return qs.order_by("date", "heure_debut")


@login_required
def planning_pdf(request):
    """Export PDF du planning des tenues, par temple — modes mois / saison / perso."""
    from io import BytesIO
    import calendar as _cal
    from django.db.models import Sum
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle,
        Paragraph, Spacer, HRFlowable, PageBreak,
    )
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    today    = date.today()
    temple_p = request.GET.get('temple') or None
    mode     = request.GET.get('mode', 'mois')

    MOIS_NOMS = {1: 'Janvier', 2: 'Février', 3: 'Mars', 4: 'Avril', 5: 'Mai',
                 6: 'Juin', 7: 'Juillet', 8: 'Août', 9: 'Septembre',
                 10: 'Octobre', 11: 'Novembre', 12: 'Décembre'}
    JOURS_FR  = ['Lun', 'Mar', 'Mer', 'Jeu', 'Ven', 'Sam', 'Dim']

    # ── Période selon le mode ──────────────────────────────────────────────────
    if mode == 'saison':
        default_annee = today.year if today.month >= 9 else today.year - 1
        annee_p       = int(request.GET.get('annee', default_annee))
        debut         = date(annee_p, 9, 1)
        fin           = date(annee_p + 1, 6, 30)
        titre_periode = f"Saison {annee_p}\u2013{annee_p + 1}"
        nom_fichier   = f"planning_saison_{annee_p}-{annee_p + 1}"
    elif mode == 'perso':
        try:
            debut = date.fromisoformat(request.GET.get('date_debut', str(today)))
            fin   = date.fromisoformat(request.GET.get('date_fin',   str(today)))
        except ValueError:
            debut = fin = today
        annee_p       = debut.year
        titre_periode = (f"Du {debut.strftime('%d/%m/%Y')} "
                         f"au {fin.strftime('%d/%m/%Y')}")
        nom_fichier   = f"planning_{debut}_{fin}"
    else:  # mois (défaut)
        mois_p  = int(request.GET.get('mois',  today.month))
        annee_p = int(request.GET.get('annee', today.year))
        debut   = date(annee_p, mois_p, 1)
        fin     = date(annee_p, mois_p, _cal.monthrange(annee_p, mois_p)[1])
        titre_periode = f"{MOIS_NOMS[mois_p]} {annee_p}"
        nom_fichier   = f"planning_{annee_p}_{mois_p:02d}"

    if temple_p:
        temples = list(Temple.objects.filter(pk=temple_p))
    else:
        temples = list(Temple.objects.all().order_by('nom'))

    C_NAVY  = colors.HexColor('#0F2137')
    C_GOLD  = colors.HexColor('#C8A84B')
    C_LIGHT = colors.HexColor('#F8FAFC')

    # Couleurs de badge pour la colonne Type
    TYPE_BG = {
        'reguliere':      colors.HexColor('#DBEAFE'),
        'exceptionnelle': colors.HexColor('#FEF3C7'),
        'congres':        colors.HexColor('#FCE7F3'),
    }

    sty_h1   = ParagraphStyle('pg_h1',  fontName='Helvetica-Bold', fontSize=14,
                               textColor=C_NAVY, spaceAfter=2)
    sty_h2   = ParagraphStyle('pg_h2',  fontName='Helvetica-Bold', fontSize=11,
                               textColor=C_GOLD, spaceAfter=6)
    sty_sub  = ParagraphStyle('pg_sub', fontName='Helvetica',      fontSize=9,
                               textColor=colors.grey, spaceAfter=10)
    sty_vide = ParagraphStyle('pg_vide', fontName='Helvetica-Oblique', fontSize=9,
                               textColor=colors.grey, alignment=TA_CENTER)

    buf = BytesIO()

    def _footer(canvas, doc):
        if doc.page == 1:
            return
        canvas.saveState()
        canvas.setFont('Helvetica', 7)
        canvas.setFillColorRGB(0.55, 0.55, 0.55)
        page_num = doc.page - 1
        canvas.drawCentredString(
            A4[0] / 2, 1.2 * cm,
            f"Kellermann \u00b7 Planning des tenues \u00b7 "
            f"{titre_periode} \u00b7 Page {page_num}",
        )
        canvas.restoreState()

    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=1.5 * cm, rightMargin=1.5 * cm,
        topMargin=1.5 * cm,  bottomMargin=2.2 * cm,
    )

    col_widths = [2.2 * cm, 1.4 * cm, 4.8 * cm, 2.8 * cm, 2.6 * cm, 2.0 * cm]
    headers    = ['Date', 'Jour', 'Loge', 'Horaires', 'Type', 'Agapes']

    def _table_temple(temple):
        tenues = (
            Reservation.objects
            .select_related('loge', 'loge__obedience', 'temple')
            .filter(temple=temple, statut='validee',
                    date__gte=debut, date__lte=fin)
            .order_by('date', 'heure_debut')
        )
        rows        = [headers]
        type_colors = []
        for i, t in enumerate(tenues, 1):
            loge_nom = (t.loge.nom if t.loge
                        else (t.nom_organisation
                              if hasattr(t, 'nom_organisation') else '\u2014'))
            agapes = (f"\u2713 {t.nombre_repas} cvts"
                      if t.besoin_agapes else '\u2014')
            rows.append([
                t.date.strftime('%d/%m/%Y'),
                JOURS_FR[t.date.weekday()],
                loge_nom,
                f"{t.heure_debut:%H:%M} \u2013 {t.heure_fin:%H:%M}",
                t.get_type_reservation_display(),
                agapes,
            ])
            if t.type_reservation in TYPE_BG:
                type_colors.append((i, TYPE_BG[t.type_reservation]))
        return rows, tenues.count(), type_colors

    story = []

    # ── Page de garde ──────────────────────────────────────────────────────────
    story.append(Spacer(1, 3 * cm))
    story.append(Paragraph('\u2692', ParagraphStyle(
        'gd_ico', fontName='Helvetica-Bold', fontSize=38,
        textColor=C_NAVY, alignment=TA_CENTER, spaceAfter=4)))
    story.append(Paragraph('Temples Kellermann', ParagraphStyle(
        'gd_brand', fontName='Helvetica-Bold', fontSize=13,
        textColor=C_GOLD, alignment=TA_CENTER, spaceAfter=16)))
    story.append(HRFlowable(width='50%', thickness=0.8, color=C_GOLD, hAlign='CENTER'))
    story.append(Spacer(1, 1 * cm))
    story.append(Paragraph('Planning des tenues', ParagraphStyle(
        'gd_titre', fontName='Helvetica-Bold', fontSize=22,
        textColor=C_NAVY, alignment=TA_CENTER, spaceAfter=8)))
    story.append(Paragraph(titre_periode, ParagraphStyle(
        'gd_periode', fontName='Helvetica', fontSize=14,
        textColor=colors.HexColor('#555555'), alignment=TA_CENTER, spaceAfter=6)))
    story.append(Paragraph(
        f'G\u00e9n\u00e9r\u00e9 le {today.strftime("%d/%m/%Y")}',
        ParagraphStyle('gd_gen', fontName='Helvetica', fontSize=9,
                       textColor=colors.grey, alignment=TA_CENTER)))
    story.append(PageBreak())

    # ── En-tête de contenu ─────────────────────────────────────────────────────
    story.append(Paragraph("Kellermann \u2014 Planning des tenues", sty_h1))
    story.append(Paragraph(titre_periode, sty_h2))
    story.append(Spacer(1, 0.2 * cm))

    # ── Tables par temple ──────────────────────────────────────────────────────
    for i, temple in enumerate(temples):
        rows, nb, type_colors = _table_temple(temple)

        story.append(Paragraph(str(temple), ParagraphStyle(
            f'tnom_{i}', fontName='Helvetica-Bold', fontSize=10,
            textColor=C_NAVY, spaceAfter=3,
        )))
        story.append(Paragraph(
            f"{nb} tenue{'s' if nb != 1 else ''}", sty_sub))

        if nb == 0:
            story.append(Paragraph("Aucune tenue sur cette p\u00e9riode.", sty_vide))
        else:
            n = len(rows)
            style_cmds = [
                ('BACKGROUND',    (0, 0),  (-1, 0),   C_NAVY),
                ('TEXTCOLOR',     (0, 0),  (-1, 0),   C_GOLD),
                ('FONTNAME',      (0, 0),  (-1, 0),   'Helvetica-Bold'),
                ('FONTSIZE',      (0, 0),  (-1, 0),   8),
                ('ALIGN',         (0, 0),  (-1, 0),   'CENTER'),
                ('TOPPADDING',    (0, 0),  (-1, 0),   5),
                ('BOTTOMPADDING', (0, 0),  (-1, 0),   5),
                ('FONTNAME',      (0, 1),  (-1, n-1), 'Helvetica'),
                ('FONTSIZE',      (0, 1),  (-1, n-1), 8),
                ('ROWBACKGROUNDS',(0, 1),  (-1, n-1), [colors.white, C_LIGHT]),
                ('TOPPADDING',    (0, 1),  (-1, n-1), 4),
                ('BOTTOMPADDING', (0, 1),  (-1, n-1), 4),
                ('ALIGN',         (1, 1),  (1, n-1),  'CENTER'),
                ('ALIGN',         (3, 1),  (3, n-1),  'CENTER'),
                ('ALIGN',         (5, 1),  (5, n-1),  'CENTER'),
                ('GRID',          (0, 0),  (-1, -1),  0.4, colors.HexColor('#CBD5E1')),
                ('BOX',           (0, 0),  (-1, -1),  1,   C_NAVY),
            ]
            # Badges colorés par type (colonne 4)
            for ri, bg in type_colors:
                style_cmds.append(('BACKGROUND', (4, ri), (4, ri), bg))

            tbl = Table(rows, colWidths=col_widths, repeatRows=1)
            tbl.setStyle(TableStyle(style_cmds))
            story.append(tbl)

        if i < len(temples) - 1:
            story.append(Spacer(1, 0.5 * cm))
            story.append(HRFlowable(width='100%', thickness=0.5,
                                    color=colors.HexColor('#E2E8F0')))
            story.append(Spacer(1, 0.3 * cm))

    # ── Résumé global ──────────────────────────────────────────────────────────
    story.append(Spacer(1, 0.6 * cm))
    story.append(HRFlowable(width='100%', thickness=0.5,
                             color=colors.HexColor('#E2E8F0')))
    story.append(Spacer(1, 0.3 * cm))

    filter_kw = dict(statut='validee', date__gte=debut, date__lte=fin)
    if temple_p:
        filter_kw['temple__pk'] = temple_p
    summary_qs   = Reservation.objects.filter(**filter_kw)
    total_tenues = summary_qs.count()
    total_cvts   = (summary_qs.filter(besoin_agapes=True)
                               .aggregate(s=Sum('nombre_repas'))['s'] or 0)

    for t in temples:
        nb_t = summary_qs.filter(temple=t).count()
        if nb_t:
            story.append(Paragraph(
                f"{t}\u00a0: {nb_t} tenue(s)",
                ParagraphStyle(f'sl_{t.pk}', fontName='Helvetica', fontSize=8,
                               textColor=colors.HexColor('#555555'), spaceAfter=2),
            ))
    story.append(Spacer(1, 0.2 * cm))
    story.append(Paragraph(
        f"Total\u00a0: {total_tenues} tenue(s) \u00b7 {total_cvts} couverts agapes",
        ParagraphStyle('pg_total', fontName='Helvetica-Bold', fontSize=8,
                       textColor=C_NAVY, spaceAfter=0),
    ))

    doc.build(story, onFirstPage=_footer, onLaterPages=_footer)
    buf.seek(0)
    if temple_p and temples:
        nom_fichier += f"_{temples[0].nom.replace(' ', '_')}"
    nom_fichier += ".pdf"
    response = HttpResponse(buf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{nom_fichier}"'
    return response


# ═══════════════════════════════════════════════════════════════════════════════
# ── Helpers partagés PDF grille / annuel ──────────────────────────────────────
# ═══════════════════════════════════════════════════════════════════════════════

_MOIS_LONG  = {1:'Janvier',2:'Février',3:'Mars',4:'Avril',5:'Mai',6:'Juin',
               7:'Juillet',8:'Août',9:'Septembre',10:'Octobre',
               11:'Novembre',12:'Décembre'}
_MOIS_COURT = ['Jan','Fév','Mar','Avr','Mai','Jun','Jul','Aoû','Sep','Oct','Nov','Déc']
_JOURS_G    = ['Lun','Mar','Mer','Jeu','Ven','Sam','Dim']
_JOURS_M    = ['L','M','M','J','V','S','D']


def _short_loge(resa):
    if resa.loge and resa.loge.abreviation:
        return resa.loge.abreviation
    if resa.loge:
        return resa.loge.nom[:6]
    return (getattr(resa, 'nom_organisation', None) or '—')[:6]


def _short_heure(t):
    if hasattr(t, 'hour'):
        return f"{t.hour}h" if t.minute == 0 else f"{t.hour}h{t.minute:02d}"
    h, m = str(t)[:5].split(':')
    return f"{int(h)}h" if m == '00' else f"{int(h)}h{m}"


_SALLE_TYPE_SHORT = {
    'banquet':  'Banq.',
    'reunion':  'Réun.',
    'chantier': 'Chant.',
    'conseil':  'Cons.',
}


def _resa_par_jour(annee, mois, temple_pk=None, loge=None):
    # Temple tenues
    qs = Reservation.objects.select_related('loge').filter(
        statut='validee', date__year=annee, date__month=mois,
    )
    if temple_pk:
        qs = qs.filter(temple_id=temple_pk)
    if loge:
        qs = qs.filter(loge=loge)
    res = {}
    for r in qs.order_by('heure_debut'):
        res.setdefault(r.date.day, []).append(
            f"{_short_loge(r)} {_short_heure(r.heure_debut)}"
        )

    # Salle reservations
    qs_s = ReservationSalle.objects.select_related('loge', 'salle').filter(
        statut__in=['validee', 'attente'],
        date__year=annee, date__month=mois,
    )
    if loge:
        qs_s = qs_s.filter(loge=loge)
    for rs in qs_s.order_by('heure_debut'):
        if loge:
            label = _SALLE_TYPE_SHORT.get(rs.type_reunion, 'Salle')
        else:
            label = _short_loge(rs)
        res.setdefault(rs.date.day, []).append(
            f"{label} {_short_heure(rs.heure_debut)}"
        )

    return res


def _semaines(annee, mois):
    import calendar as _cal
    fw, nd = _cal.monthrange(annee, mois)
    weeks, week = [], [0] * fw
    for d in range(1, nd + 1):
        week.append(d)
        if len(week) == 7:
            weeks.append(week); week = []
    if week:
        weeks.append(week + [0] * (7 - len(week)))
    while len(weeks) < 6:
        weeks.append([0] * 7)
    return weeks


def _mini_mois_table(annee, mois, resa_dict, col_w, hdr_h, day_h, row_h, fsize):
    """Mini-calendrier compact (pour formats annuels)."""
    from reportlab.platypus import Table, TableStyle, Paragraph
    from reportlab.lib import colors
    from reportlab.lib.styles import ParagraphStyle

    C_BDR = colors.HexColor('#444444')
    C_MHD = colors.HexColor('#CCCCCC')
    C_DHD = colors.HexColor('#E0E0E0')
    C_ALT = colors.HexColor('#F5F5F5')

    def _sty(name, bold=False, sz=None, align=1):
        return ParagraphStyle(name,
            fontName='Helvetica-Bold' if bold else 'Helvetica',
            fontSize=sz or fsize, leading=(sz or fsize) + 1.5,
            alignment=align, spaceAfter=0, spaceBefore=0)

    s_hdr  = _sty('mh', bold=True,  sz=fsize + 0.5)
    s_jour = _sty('mj', bold=True,  sz=fsize - 0.5)
    s_cell = _sty('mc', bold=False, sz=fsize - 0.5, align=0)

    semaines = _semaines(annee, mois)
    label    = f"{_MOIS_COURT[mois - 1]} {annee}"

    data = [[Paragraph(label, s_hdr)] + [''] * 6]
    data.append([Paragraph(j, s_jour) for j in _JOURS_M])
    for sem in semaines:
        row = []
        for day in sem:
            if day == 0:
                row.append('')
            else:
                entries = resa_dict.get(day, [])
                html = f"<b>{day}</b>" + ''.join(f'<br/>{e}' for e in entries)
                row.append(Paragraph(html, s_cell))
        data.append(row)

    tbl = Table(data, colWidths=[col_w] * 7,
                rowHeights=[hdr_h, day_h] + [row_h] * 6)
    tbl.setStyle(TableStyle([
        ('SPAN',          (0, 0), (6, 0)),
        ('BACKGROUND',    (0, 0), (6, 0), C_MHD),
        ('BACKGROUND',    (0, 1), (6, 1), C_DHD),
        ('ALIGN',         (0, 0), (6, 1), 'CENTER'),
        ('VALIGN',        (0, 0), (6, 1), 'MIDDLE'),
        ('VALIGN',        (0, 2), (-1, -1), 'TOP'),
        ('ROWBACKGROUNDS',(0, 2), (-1, -1), [colors.white, C_ALT]),
        ('GRID',          (0, 0), (-1, -1), 0.25, C_BDR),
        ('BOX',           (0, 0), (-1, -1), 0.5,  C_BDR),
        ('TOPPADDING',    (0, 0), (-1, -1), 1),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
        ('LEFTPADDING',   (0, 0), (-1, -1), 1),
        ('RIGHTPADDING',  (0, 0), (-1, -1), 0),
    ]))
    return tbl


def _pdf_grille_mensuelle(annee, mois, titre, nom_fichier, temple_pk=None, loge=None):
    """Génère le PDF grille mensuelle A4 paysage."""
    from io import BytesIO
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    resa = _resa_par_jour(annee, mois, temple_pk=temple_pk, loge=loge)

    PAGE  = landscape(A4)
    MG    = 0.8 * cm
    UW    = PAGE[0] - 2 * MG
    UH    = PAGE[1] - 2 * MG

    TITLE_H = 0.6  * cm
    SPACE_H = 0.15 * cm
    COL_W   = UW / 7
    DAY_H   = 0.45 * cm
    ROW_H   = (UH - TITLE_H - SPACE_H - DAY_H) / 6

    C_BDR = colors.HexColor('#333333')
    C_DHD = colors.HexColor('#DDDDDD')
    C_ALT = colors.HexColor('#F8F8F8')

    sty_title = ParagraphStyle('gt', fontName='Helvetica-Bold', fontSize=10,
                                textColor=colors.HexColor('#222222'),
                                alignment=TA_CENTER, spaceAfter=0)
    sty_jour  = ParagraphStyle('gj', fontName='Helvetica-Bold', fontSize=8,
                                leading=10, alignment=TA_CENTER)
    sty_cell  = ParagraphStyle('gc', fontName='Helvetica', fontSize=7.5,
                                leading=9.5, alignment=TA_LEFT)

    semaines = _semaines(annee, mois)

    data = [[Paragraph(j, sty_jour) for j in _JOURS_G]]
    for sem in semaines:
        row = []
        for day in sem:
            if day == 0:
                row.append('')
            else:
                entries = resa.get(day, [])
                html = f"<b>{day}</b>" + ''.join(f'<br/>{e}' for e in entries)
                row.append(Paragraph(html, sty_cell))
        data.append(row)

    tbl = Table(data, colWidths=[COL_W] * 7,
                rowHeights=[DAY_H] + [ROW_H] * 6)
    tbl.setStyle(TableStyle([
        ('BACKGROUND',    (0, 0), (-1, 0), C_DHD),
        ('ALIGN',         (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN',        (0, 0), (-1, 0), 'MIDDLE'),
        ('VALIGN',        (0, 1), (-1, -1), 'TOP'),
        ('ROWBACKGROUNDS',(0, 1), (-1, -1), [colors.white, C_ALT]),
        ('GRID',          (0, 0), (-1, -1), 0.4, C_BDR),
        ('BOX',           (0, 0), (-1, -1), 0.8, C_BDR),
        ('TOPPADDING',    (0, 0), (-1,  0), 4),
        ('BOTTOMPADDING', (0, 0), (-1,  0), 4),
        ('TOPPADDING',    (0, 1), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 2),
        ('LEFTPADDING',   (0, 1), (-1, -1), 3),
        ('RIGHTPADDING',  (0, 1), (-1, -1), 2),
    ]))

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=PAGE,
                            leftMargin=MG, rightMargin=MG,
                            topMargin=MG,  bottomMargin=MG)
    doc.build([Paragraph(titre, sty_title), Spacer(1, SPACE_H), tbl])
    buf.seek(0)
    response = HttpResponse(buf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{nom_fichier}.pdf"'
    return response


def _pdf_annuel(annee, titre, nom_fichier, base_pagesize, ncols, temple_pk=None, loge=None):
    """
    Génère le PDF annuel (saison Sep→Aug).
    ncols=4 → A3 landscape, 1 page (4×3 mois)
    ncols=3 → A4 landscape, 2 pages (3×2 mois par page)
    """
    from io import BytesIO
    from reportlab.lib.pagesizes import landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle,
        Paragraph, Spacer, PageBreak,
    )
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER

    mois_saison = ([(annee, m) for m in range(9, 13)]
                   + [(annee + 1, m) for m in range(1, 9)])

    PAGE = landscape(base_pagesize)
    MG   = 0.8 * cm
    UW   = PAGE[0] - 2 * MG
    UH   = PAGE[1] - 2 * MG

    sty_title = ParagraphStyle('at', fontName='Helvetica-Bold', fontSize=9,
                                textColor=colors.HexColor('#222222'),
                                alignment=TA_CENTER, spaceAfter=0)

    TITLE_H = 16
    SPACE_H = 6

    if ncols == 4:          # A3 : 4×3 sur 1 page
        fsize     = 6.0
        hdr_h     = 12
        day_h     = 8
        nrows_pg  = 3
        row_gap   = 4
        avail     = UH - TITLE_H - SPACE_H - (nrows_pg - 1) * row_gap
        row_h     = (avail - nrows_pg * (hdr_h + day_h)) / (nrows_pg * 6)
    else:                   # A4 : 3×2 sur 2 pages
        fsize     = 7.0
        hdr_h     = 13
        day_h     = 9
        nrows_pg  = 2
        row_gap   = 4
        avail     = UH - TITLE_H - SPACE_H - (nrows_pg - 1) * row_gap
        row_h     = (avail - nrows_pg * (hdr_h + day_h)) / (nrows_pg * 6)

    col_w       = UW / (ncols * 7)
    outer_col_w = UW / ncols
    outer_row_h = hdr_h + day_h + 6 * row_h

    months_per_page = ncols * nrows_pg

    def _build_page_table(page_months):
        outer_data = []
        for r in range(nrows_pg):
            outer_row = []
            for c in range(ncols):
                idx = r * ncols + c
                if idx < len(page_months) and page_months[idx] is not None:
                    a, m = page_months[idx]
                    rd  = _resa_par_jour(a, m, temple_pk=temple_pk, loge=loge)
                    outer_row.append(
                        _mini_mois_table(a, m, rd, col_w, hdr_h, day_h, row_h, fsize)
                    )
                else:
                    outer_row.append('')
            outer_data.append(outer_row)

        outer = Table(outer_data,
                      colWidths=[outer_col_w] * ncols,
                      rowHeights=[outer_row_h] * nrows_pg)
        outer.setStyle(TableStyle([
            ('VALIGN',       (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING',  (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 0),
            ('TOPPADDING',   (0, 0), (-1, -1), 0),
            ('BOTTOMPADDING',(0, 0), (-1, -1), 0),
        ]))
        return outer

    story = []
    for page_idx in range(0, 12, months_per_page):
        if page_idx > 0:
            story.append(PageBreak())
        page_months = mois_saison[page_idx:page_idx + months_per_page]
        story.append(Paragraph(titre, sty_title))
        story.append(Spacer(1, SPACE_H))
        story.append(_build_page_table(page_months))

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=PAGE,
                            leftMargin=MG, rightMargin=MG,
                            topMargin=MG,  bottomMargin=MG)
    doc.build(story)
    buf.seek(0)
    response = HttpResponse(buf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{nom_fichier}.pdf"'
    return response


# ── Vues admin ────────────────────────────────────────────────────────────────

def planning_pdf_grille_mensuelle(request):
    today    = date.today()
    mois_p   = int(request.GET.get('mois',  today.month))
    annee_p  = int(request.GET.get('annee', today.year))
    temple_p = request.GET.get('temple') or None

    titre = f"Kellermann — {_MOIS_LONG[mois_p]} {annee_p}"
    if temple_p:
        t_obj = Temple.objects.filter(pk=temple_p).first()
        if t_obj:
            titre += f" · {t_obj}"

    return _pdf_grille_mensuelle(
        annee_p, mois_p, titre,
        nom_fichier=f"grille_{annee_p}_{mois_p:02d}",
        temple_pk=temple_p,
    )


def planning_pdf_annuel_a3(request):
    from reportlab.lib.pagesizes import A3
    today    = date.today()
    default_annee = today.year if today.month >= 9 else today.year - 1
    annee_p  = int(request.GET.get('annee', default_annee))
    temple_p = request.GET.get('temple') or None

    titre = f"Kellermann — Saison {annee_p}–{annee_p + 1}"
    if temple_p:
        t_obj = Temple.objects.filter(pk=temple_p).first()
        if t_obj:
            titre += f" · {t_obj}"

    return _pdf_annuel(
        annee_p, titre,
        nom_fichier=f"annuel_a3_{annee_p}-{annee_p + 1}",
        base_pagesize=A3, ncols=4,
        temple_pk=temple_p,
    )


def planning_pdf_annuel_a4(request):
    from reportlab.lib.pagesizes import A4
    today    = date.today()
    default_annee = today.year if today.month >= 9 else today.year - 1
    annee_p  = int(request.GET.get('annee', default_annee))
    temple_p = request.GET.get('temple') or None

    titre = f"Kellermann — Saison {annee_p}–{annee_p + 1}"
    if temple_p:
        t_obj = Temple.objects.filter(pk=temple_p).first()
        if t_obj:
            titre += f" · {t_obj}"

    return _pdf_annuel(
        annee_p, titre,
        nom_fichier=f"annuel_a4_{annee_p}-{annee_p + 1}",
        base_pagesize=A4, ncols=3,
        temple_pk=temple_p,
    )


# ── Vues portail (token) ─────────────────────────────────────────────────────

def planning_loge_grille_pdf(request, token):
    demande = get_object_or_404(DemandeAccesPortail, token=token, statut='validee')
    loge    = demande.loge
    if not loge:
        return HttpResponse("Aucune loge associée à ce token.", status=400)

    today   = date.today()
    mois_p  = int(request.GET.get('mois',  today.month))
    annee_p = int(request.GET.get('annee', today.year))

    titre = f"{loge.nom} — {_MOIS_LONG[mois_p]} {annee_p}"
    return _pdf_grille_mensuelle(
        annee_p, mois_p, titre,
        nom_fichier=f"grille_{loge.nom.replace(' ', '_')}_{annee_p}_{mois_p:02d}",
        loge=loge,
    )


def planning_loge_annuel_pdf(request, token):
    from reportlab.lib.pagesizes import A4
    demande = get_object_or_404(DemandeAccesPortail, token=token, statut='validee')
    loge    = demande.loge
    if not loge:
        return HttpResponse("Aucune loge associée à ce token.", status=400)

    today         = date.today()
    annee_default = today.year if today.month >= 9 else today.year - 1
    annee_p       = int(request.GET.get('annee', annee_default))

    titre = f"{loge.nom} — Saison {annee_p}–{annee_p + 1}"
    return _pdf_annuel(
        annee_p, titre,
        nom_fichier=f"annuel_{loge.nom.replace(' ', '_')}_{annee_p}-{annee_p + 1}",
        base_pagesize=A4, ncols=3,
        loge=loge,
    )


def planning_loge_pdf(request, token):
    """Export PDF du planning de saison d'une loge, accessible via token portail."""
    from io import BytesIO
    from django.db.models import Sum
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle,
        Paragraph, Spacer, HRFlowable, PageBreak,
    )
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    demande = get_object_or_404(DemandeAccesPortail, token=token, statut='validee')
    loge    = demande.loge
    if not loge:
        return HttpResponse("Aucune loge associée à ce token.", status=400)

    today = date.today()

    # Priorité : 1) paramètre GET ?annee=  2) ValidationSaison active  3) today
    annee_default = today.year if today.month >= 9 else today.year - 1
    if request.GET.get('annee'):
        try:
            annee = int(request.GET['annee'])
        except ValueError:
            annee = annee_default
    else:
        val = ValidationSaison.objects.filter(
            loge=loge,
            statut__in=['ouverte', 'soumise'],
        ).order_by('-annee').first()
        annee = val.annee if val else annee_default

    debut_saison = date(annee, 9, 1)
    fin_saison   = date(annee + 1, 8, 31)

    tenues = (
        Reservation.objects
        .select_related('temple')
        .filter(loge=loge, statut='validee',
                date__gte=debut_saison, date__lte=fin_saison)
        .order_by('date', 'heure_debut')
    )

    resas_salle = (
        ReservationSalle.objects
        .select_related('salle')
        .filter(loge=loge, statut__in=['validee', 'attente'],
                date__gte=debut_saison, date__lte=fin_saison)
        .order_by('date', 'heure_debut')
    )

    JOURS_FR = ['Lun', 'Mar', 'Mer', 'Jeu', 'Ven', 'Sam', 'Dim']
    C_HEADER = colors.HexColor('#222222')
    C_LIGHT  = colors.HexColor('#F2F2F2')
    C_BORDER = colors.HexColor('#BBBBBB')

    buf = BytesIO()

    def _footer(canvas, doc):
        if doc.page == 1:
            return
        canvas.saveState()
        canvas.setFont('Helvetica', 7)
        canvas.setFillColorRGB(0.55, 0.55, 0.55)
        canvas.drawCentredString(
            A4[0] / 2, 1.2 * cm,
            f"Temples Kellermann \u00b7 {loge.nom} \u00b7 "
            f"Saison {annee}\u2013{annee + 1} \u00b7 Page {doc.page - 1}",
        )
        canvas.restoreState()

    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=1.8 * cm, rightMargin=1.8 * cm,
        topMargin=1.5 * cm,  bottomMargin=2.2 * cm,
    )

    def _sty(name, size, bold=False, color='#0F2137', align=TA_CENTER, after=6):
        return ParagraphStyle(name,
                              fontName='Helvetica-Bold' if bold else 'Helvetica',
                              fontSize=size,
                              textColor=colors.HexColor(color),
                              alignment=align,
                              spaceAfter=after)

    story = []

    # ── Page de garde ──────────────────────────────────────────────────────────
    story.append(Spacer(1, 3 * cm))
    story.append(Paragraph('\u2692', _sty('lp_ico', 38, bold=True)))
    story.append(Paragraph('Temples Kellermann',
                            _sty('lp_brand', 13, bold=True, color='#C8A84B', after=16)))
    story.append(HRFlowable(width='50%', thickness=0.8,
                             color=colors.HexColor('#C8A84B'), hAlign='CENTER'))
    story.append(Spacer(1, 1 * cm))
    story.append(Paragraph('Planning de saison', _sty('lp_titre', 20, bold=True)))
    story.append(Spacer(1, 0.5 * cm))
    story.append(Paragraph(loge.nom, _sty('lp_loge', 14, bold=True, color='#333333')))
    story.append(Spacer(1, 0.3 * cm))
    story.append(Paragraph(f'Saison {annee}\u2013{annee + 1}',
                            _sty('lp_saison', 11, color='#555555')))
    story.append(Paragraph(f'01/09/{annee} \u2192 31/08/{annee + 1}',
                            _sty('lp_periode', 9, color='#888888')))
    story.append(Spacer(1, 2.5 * cm))
    story.append(Paragraph(f'G\u00e9n\u00e9r\u00e9 le {today.strftime("%d/%m/%Y")}',
                            _sty('lp_gen', 8, color='#999999')))
    story.append(PageBreak())

    # ── Tableau planning (tenues + salles) ──────────────────────────────────────────
    C_SALLE = colors.HexColor('#EDE9FF')

    lignes = []
    for t in tenues:
        lignes.append({
            'kind': 'temple',
            'date': t.date,
            'heure_debut': t.heure_debut,
            'heure_fin': t.heure_fin,
            'lieu': str(t.temple) if t.temple else '\u2014',
            'type': t.get_type_reservation_display(),
            'note': (f"\u2713 {t.nombre_repas} cvts" if t.besoin_agapes else ''),
        })
    for rs in resas_salle:
        if rs.type_reunion == 'banquet':
            type_label = 'Banquet'
        elif hasattr(rs, 'get_type_reunion_display'):
            type_label = rs.get_type_reunion_display()
        else:
            type_label = rs.salle.type_salle if rs.salle else ''
        lignes.append({
            'kind': 'salle',
            'date': rs.date,
            'heure_debut': rs.heure_debut,
            'heure_fin': rs.heure_fin,
            'lieu': str(rs.salle) if rs.salle else '\u2014',
            'type': type_label,
            'note': rs.objet or '',
        })
    lignes.sort(key=lambda x: (x['date'], x['heure_debut']))

    nb_tenues = sum(1 for l in lignes if l['kind'] == 'temple')
    nb_salles = sum(1 for l in lignes if l['kind'] == 'salle')
    nb_total  = len(lignes)

    story.append(Paragraph(
        f'Planning saison {annee}\u2013{annee + 1}',
        _sty('lt_h2', 11, bold=True, align=TA_LEFT, after=3),
    ))
    sous_titre = f'{nb_tenues} tenue{"s" if nb_tenues != 1 else ""} au temple'
    if nb_salles:
        sous_titre += (f' \u00b7 {nb_salles} '
                       f'r\u00e9servation{"s" if nb_salles != 1 else ""} de salle')
    story.append(Paragraph(
        sous_titre,
        _sty('lt_sub', 9, color='#888888', align=TA_LEFT, after=8),
    ))

    if nb_total == 0:
        story.append(Paragraph(
            'Aucune r\u00e9servation pour cette saison.',
            _sty('lt_vide', 9, color='#888888'),
        ))
    else:
        col_widths = [2.2 * cm, 1.3 * cm, 3.8 * cm, 2.8 * cm, 3.8 * cm, 2.5 * cm]
        rows = [['Date', 'Jour', 'Lieu', 'Horaires', 'Type', 'Note']]
        row_kinds = []
        for lg in lignes:
            rows.append([
                lg['date'].strftime('%d/%m/%Y'),
                JOURS_FR[lg['date'].weekday()],
                lg['lieu'],
                f"{lg['heure_debut']:%H:%M} \u2013 {lg['heure_fin']:%H:%M}",
                lg['type'],
                lg['note'],
            ])
            row_kinds.append(lg['kind'])
        nr = len(rows)
        tbl = Table(rows, colWidths=col_widths, repeatRows=1)
        tbl_style = [
            ('BACKGROUND',    (0, 0),   (-1, 0),    C_HEADER),
            ('TEXTCOLOR',     (0, 0),   (-1, 0),    colors.white),
            ('FONTNAME',      (0, 0),   (-1, 0),    'Helvetica-Bold'),
            ('FONTSIZE',      (0, 0),   (-1, -1),   8),
            ('ALIGN',         (0, 0),   (-1, 0),    'CENTER'),
            ('TOPPADDING',    (0, 0),   (-1, -1),   4),
            ('BOTTOMPADDING', (0, 0),   (-1, -1),   4),
            ('FONTNAME',      (0, 1),   (-1, nr-1), 'Helvetica'),
            ('ALIGN',         (1, 1),   (1, nr-1),  'CENTER'),
            ('ALIGN',         (3, 1),   (3, nr-1),  'CENTER'),
            ('GRID',          (0, 0),   (-1, -1),   0.3, C_BORDER),
            ('BOX',           (0, 0),   (-1, -1),   0.6, colors.HexColor('#666666')),
        ]
        for i, kind in enumerate(row_kinds):
            ri = i + 1
            if kind == 'salle':
                tbl_style.append(('BACKGROUND', (0, ri), (-1, ri), C_SALLE))
            elif i % 2 == 0:
                tbl_style.append(('BACKGROUND', (0, ri), (-1, ri), colors.white))
            else:
                tbl_style.append(('BACKGROUND', (0, ri), (-1, ri), C_LIGHT))
        tbl.setStyle(TableStyle(tbl_style))
        story.append(tbl)

        nb_agapes  = tenues.filter(besoin_agapes=True).count()
        total_cvts = (tenues.filter(besoin_agapes=True)
                             .aggregate(s=Sum('nombre_repas'))['s'] or 0)
        story.append(Spacer(1, 0.4 * cm))
        resume = f'Total\u00a0: {nb_tenues} tenue(s) au temple'
        if nb_salles:
            resume += f' + {nb_salles} r\u00e9servation(s) de salle'
        if nb_agapes:
            resume += f'  \u00b7  {nb_agapes} avec agapes ({total_cvts} couverts)'
        story.append(Paragraph(
            resume,
            _sty('lt_resume', 8, color='#777777', align=TA_LEFT, after=0),
        ))

    doc.build(story, onFirstPage=_footer, onLaterPages=_footer)
    buf.seek(0)
    nom = f"planning_{loge.nom.replace(' ', '_')}_{annee}-{annee + 1}.pdf"
    response = HttpResponse(buf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{nom}"'
    return response


@login_required
def bilan_saison_excel(request):
    """Bilan complet d'occupation par structure pour une saison :
    créneaux, heures, effectifs, ventilation par type de salle."""
    from collections import defaultdict
    from datetime import datetime
    from openpyxl.styles import Border, Side, numbers
    from openpyxl.utils import get_column_letter
    from temple_project.apps.reservations.models import ReservationSalle
    from temple_project.apps.loges.models import Loge

    today = date.today()
    default_annee = today.year if today.month >= 7 else today.year - 1
    annee = int(request.GET.get('saison', default_annee))
    debut = date(annee, 9, 1)
    fin   = date(annee + 1, 8, 31)
    saison_label = f"{annee}-{annee + 1}"

    def duree_h(r):
        d = datetime.combine(r.date, r.heure_fin) - datetime.combine(r.date, r.heure_debut)
        return round(d.total_seconds() / 3600, 2)

    # ── Collecte réservations temples ────────────────────────────────────────
    resas_temple = (
        Reservation.objects
        .filter(date__gte=debut, date__lte=fin, statut='validee')
        .select_related('loge', 'loge__obedience', 'temple')
        .order_by('loge__nom', 'date')
    )

    # ── Collecte réservations salles (agapes / cabinets / réunion) ──────────
    resas_salle = (
        ReservationSalle.objects
        .filter(date__gte=debut, date__lte=fin, statut='validee')
        .select_related('loge', 'loge__obedience', 'salle')
        .order_by('loge__nom', 'date')
    )

    # ── Agrégation par loge ──────────────────────────────────────────────────
    # clé : loge pk (ou None pour réservations sans loge)
    struct = defaultdict(lambda: {
        'label': '', 'abrev': '', 'obedience': '', 'type_loge': '',
        'rite': '', 'effectif_total': '', 'effectif_agapes': '',
        'temple_creneaux': 0, 'temple_heures': 0.0,
        'agapes_creneaux': 0, 'agapes_heures': 0.0,
        'cabinet_creneaux': 0,
        'reunion_creneaux': 0, 'reunion_heures': 0.0,
        'loge_obj': None,
    })

    for r in resas_temple:
        key = r.loge_id or f'ext_{r.nom_demandeur}'
        s = struct[key]
        if r.loge and not s['label']:
            lg = r.loge
            s['label']          = lg.nom
            s['abrev']          = lg.abreviation or ''
            s['obedience']      = lg.obedience.nom if lg.obedience else ''
            s['type_loge']      = lg.get_type_loge_display() if hasattr(lg, 'get_type_loge_display') else lg.type_loge
            s['rite']           = lg.rite or ''
            s['effectif_total'] = lg.effectif_total or ''
            s['effectif_agapes']= lg.effectif_moyen_agapes or ''
            s['loge_obj']       = lg
        elif not s['label']:
            s['label'] = r.nom_organisation or r.nom_demandeur or '(sans loge)'
        s['temple_creneaux'] += 1
        s['temple_heures']   += duree_h(r)

    for r in resas_salle:
        key = r.loge_id or f'ext_{r.nom_demandeur}'
        s = struct[key]
        if r.loge and not s['label']:
            lg = r.loge
            s['label']          = lg.nom
            s['abrev']          = lg.abreviation or ''
            s['obedience']      = lg.obedience.nom if lg.obedience else ''
            s['type_loge']      = lg.get_type_loge_display() if hasattr(lg, 'get_type_loge_display') else lg.type_loge
            s['rite']           = lg.rite or ''
            s['effectif_total'] = lg.effectif_total or ''
            s['effectif_agapes']= lg.effectif_moyen_agapes or ''
            s['loge_obj']       = lg
        elif not s['label']:
            s['label'] = r.organisation or r.nom_demandeur or '(sans loge)'
        ts = r.salle.type_salle if r.salle else ''
        if ts == 'agapes':
            s['agapes_creneaux'] += 1
            s['agapes_heures']   += duree_h(r)
        elif ts == 'cabinet_reflexion':
            s['cabinet_creneaux'] += 1
        elif ts == 'reunion':
            s['reunion_creneaux'] += 1
            s['reunion_heures']   += duree_h(r)

    lignes = sorted(struct.values(), key=lambda x: x['label'].lower())

    # ── Excel ────────────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()

    # ── Feuille 1 : BILAN PAR STRUCTURE ─────────────────────────────────────
    ws = wb.active
    ws.title = "Bilan par structure"

    BLEU   = "0F2137"
    OR     = "C8A84B"
    GRIS   = "F1F5F9"
    VERT   = "D1FAE5"
    thin   = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _hdr(ws, row, col, val, bg=BLEU, fg=OR, bold=True, center=True):
        c = ws.cell(row=row, column=col, value=val)
        c.font      = Font(bold=bold, color=fg, size=9)
        c.fill      = PatternFill('solid', fgColor=bg)
        c.alignment = Alignment(horizontal='center' if center else 'left',
                                vertical='center', wrap_text=True)
        c.border    = border
        return c

    def _cell(ws, row, col, val, bold=False, align='left', fmt=None, bg=None):
        c = ws.cell(row=row, column=col, value=val)
        c.font      = Font(bold=bold, size=9)
        c.alignment = Alignment(horizontal=align, vertical='center')
        c.border    = border
        if bg:
            c.fill  = PatternFill('solid', fgColor=bg)
        if fmt:
            c.number_format = fmt
        return c

    # Ligne 1 : titre
    ws.merge_cells('A1:Q1')
    t = ws.cell(row=1, column=1, value=f"BILAN D'OCCUPATION DES TEMPLES KELLERMANN — Saison {saison_label}")
    t.font      = Font(bold=True, color=OR, size=12)
    t.fill      = PatternFill('solid', fgColor=BLEU)
    t.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 22

    # Ligne 2 : groupes de colonnes
    ws.merge_cells('A2:G2'); _hdr(ws, 2, 1,  'STRUCTURE')
    ws.merge_cells('H2:I2'); _hdr(ws, 2, 8,  'TEMPLES', bg='1E3A5F')
    ws.merge_cells('J2:K2'); _hdr(ws, 2, 10, 'BANQUET / AGAPES', bg='92400E')
    _hdr(ws, 2, 12, 'CABINETS', bg='4B5563')
    ws.merge_cells('M2:N2'); _hdr(ws, 2, 13, 'SALLES DE RÉUNION', bg='065F46')
    ws.merge_cells('O2:Q2'); _hdr(ws, 2, 15, 'TOTAUX', bg='1E3A5F')

    # Ligne 3 : sous-en-têtes
    cols_hdr = [
        'Loge', 'Abrév.', 'Obédience', 'Type', 'Rite',
        'Effectif total', 'Moy. agapes',
        'Créneaux', 'Heures',
        'Créneaux', 'Heures',
        'Créneaux',
        'Créneaux', 'Heures',
        'Créneaux total', 'Heures total', 'Dont temples',
    ]
    for ci, h in enumerate(cols_hdr, 1):
        _hdr(ws, 3, ci, h, bg='334155', fg='FFFFFF')
    ws.row_dimensions[3].height = 32

    # Données
    for ri, s in enumerate(lignes, 4):
        tot_creneaux = s['temple_creneaux'] + s['agapes_creneaux'] + s['cabinet_creneaux'] + s['reunion_creneaux']
        tot_heures   = round(s['temple_heures'] + s['agapes_heures'] + s['reunion_heures'], 2)
        bg_row = 'FFFFFF' if ri % 2 == 0 else GRIS

        _cell(ws, ri,  1, s['label'],           bold=True,  bg=bg_row)
        _cell(ws, ri,  2, s['abrev'],            align='center', bg=bg_row)
        _cell(ws, ri,  3, s['obedience'],        bg=bg_row)
        _cell(ws, ri,  4, s['type_loge'],        align='center', bg=bg_row)
        _cell(ws, ri,  5, s['rite'],             align='center', bg=bg_row)
        _cell(ws, ri,  6, s['effectif_total'],   align='center', bg=bg_row)
        _cell(ws, ri,  7, s['effectif_agapes'],  align='center', bg=bg_row)
        _cell(ws, ri,  8, s['temple_creneaux'],  align='center', bg=bg_row)
        _cell(ws, ri,  9, round(s['temple_heures'], 2), align='center', fmt='0.0"h"', bg=bg_row)
        _cell(ws, ri, 10, s['agapes_creneaux'],  align='center', bg=bg_row)
        _cell(ws, ri, 11, round(s['agapes_heures'], 2), align='center', fmt='0.0"h"', bg=bg_row)
        _cell(ws, ri, 12, s['cabinet_creneaux'], align='center', bg=bg_row)
        _cell(ws, ri, 13, s['reunion_creneaux'], align='center', bg=bg_row)
        _cell(ws, ri, 14, round(s['reunion_heures'], 2), align='center', fmt='0.0"h"', bg=bg_row)
        _cell(ws, ri, 15, tot_creneaux,          align='center', bold=True, bg=VERT)
        _cell(ws, ri, 16, tot_heures,            align='center', bold=True, fmt='0.0"h"', bg=VERT)
        _cell(ws, ri, 17, s['temple_creneaux'],  align='center', bg=VERT)

    # Ligne TOTAL
    ri_tot = len(lignes) + 4
    ws.merge_cells(f'A{ri_tot}:G{ri_tot}')
    _cell(ws, ri_tot, 1, 'TOTAL SAISON', bold=True, bg='0F2137')
    ws.cell(row=ri_tot, column=1).font = Font(bold=True, color=OR, size=9)

    for ci, key in enumerate([
        'temple_creneaux', 'temple_heures', 'agapes_creneaux', 'agapes_heures',
        'cabinet_creneaux', 'reunion_creneaux', 'reunion_heures',
    ], 8):
        is_h = 'heures' in key
        v = round(sum(s[key] for s in lignes), 2) if is_h else sum(s[key] for s in lignes)
        _cell(ws, ri_tot, ci, v, bold=True, align='center',
              fmt='0.0"h"' if is_h else None, bg='E2E8F0')

    _cell(ws, ri_tot, 15, sum(
        s['temple_creneaux']+s['agapes_creneaux']+s['cabinet_creneaux']+s['reunion_creneaux']
        for s in lignes), bold=True, align='center', bg=VERT)
    _cell(ws, ri_tot, 16, round(sum(
        s['temple_heures']+s['agapes_heures']+s['reunion_heures']
        for s in lignes), 2), bold=True, align='center', fmt='0.0"h"', bg=VERT)
    _cell(ws, ri_tot, 17, sum(s['temple_creneaux'] for s in lignes),
          bold=True, align='center', bg=VERT)

    # Largeurs colonnes
    largeurs = [38, 10, 22, 14, 12, 14, 12, 10, 9, 10, 9, 10, 10, 9, 14, 12, 12]
    for i, w in enumerate(largeurs, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A4'

    # ── Feuille 2 : DÉTAIL TENUES TEMPLE ────────────────────────────────────
    ws2 = wb.create_sheet("Détail tenues temples")
    hdrs2 = ['Date', 'Jour', 'Loge', 'Obédience', 'Temple',
             'Type', 'Heure début', 'Heure fin', 'Durée (h)',
             'Agapes', 'Nb repas', 'Statut', 'Tarif €']
    for ci, h in enumerate(hdrs2, 1):
        _hdr(ws2, 1, ci, h, bg=BLEU, fg=OR)
    ws2.freeze_panes = 'A2'

    JOURS = ['Lun', 'Mar', 'Mer', 'Jeu', 'Ven', 'Sam', 'Dim']
    for ri, r in enumerate(resas_temple, 2):
        bg = GRIS if ri % 2 == 0 else 'FFFFFF'
        _cell(ws2, ri,  1, r.date,         fmt='DD/MM/YYYY', bg=bg)
        _cell(ws2, ri,  2, JOURS[r.date.weekday()], align='center', bg=bg)
        _cell(ws2, ri,  3, r.loge.nom if r.loge else (r.nom_organisation or r.nom_demandeur or ''), bg=bg)
        _cell(ws2, ri,  4, r.loge.obedience.nom if r.loge and r.loge.obedience else '', bg=bg)
        _cell(ws2, ri,  5, str(r.temple) if r.temple else '', bg=bg)
        _cell(ws2, ri,  6, r.get_type_reservation_display(), align='center', bg=bg)
        _cell(ws2, ri,  7, str(r.heure_debut)[:5], align='center', bg=bg)
        _cell(ws2, ri,  8, str(r.heure_fin)[:5],   align='center', bg=bg)
        _cell(ws2, ri,  9, duree_h(r), align='center', fmt='0.0"h"', bg=bg)
        _cell(ws2, ri, 10, 'Oui' if r.besoin_agapes else 'Non', align='center', bg=bg)
        _cell(ws2, ri, 11, r.nombre_repas or '', align='center', bg=bg)
        _cell(ws2, ri, 12, r.get_statut_display(), align='center', bg=bg)
        _cell(ws2, ri, 13, float(r.tarif) if r.tarif else 0, align='center', fmt='#,##0.00 €', bg=bg)

    for ci, w in enumerate([12,6,32,20,20,14,9,9,9,6,8,10,10], 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w

    # ── Feuille 3 : DÉTAIL SALLES ────────────────────────────────────────────
    ws3 = wb.create_sheet("Détail salles et cabinets")
    hdrs3 = ['Date', 'Jour', 'Loge', 'Obédience', 'Type salle', 'Salle',
             'Heure début', 'Heure fin', 'Durée (h)', 'Participants', 'Statut']
    for ci, h in enumerate(hdrs3, 1):
        _hdr(ws3, 1, ci, h, bg=BLEU, fg=OR)
    ws3.freeze_panes = 'A2'

    TYPE_SALLE_LABEL = {'agapes': 'Banquet/Agapes', 'cabinet_reflexion': 'Cabinet', 'reunion': 'Salle réunion'}
    for ri, r in enumerate(resas_salle, 2):
        bg = GRIS if ri % 2 == 0 else 'FFFFFF'
        ts = r.salle.type_salle if r.salle else ''
        _cell(ws3, ri,  1, r.date,         fmt='DD/MM/YYYY', bg=bg)
        _cell(ws3, ri,  2, JOURS[r.date.weekday()], align='center', bg=bg)
        _cell(ws3, ri,  3, r.loge.nom if r.loge else (r.organisation or r.nom_demandeur or ''), bg=bg)
        _cell(ws3, ri,  4, r.loge.obedience.nom if r.loge and r.loge.obedience else '', bg=bg)
        _cell(ws3, ri,  5, TYPE_SALLE_LABEL.get(ts, ts), align='center', bg=bg)
        _cell(ws3, ri,  6, str(r.salle) if r.salle else '', bg=bg)
        _cell(ws3, ri,  7, str(r.heure_debut)[:5], align='center', bg=bg)
        _cell(ws3, ri,  8, str(r.heure_fin)[:5],   align='center', bg=bg)
        _cell(ws3, ri,  9, duree_h(r), align='center', fmt='0.0"h"', bg=bg)
        _cell(ws3, ri, 10, r.nombre_participants or '', align='center', bg=bg)
        _cell(ws3, ri, 11, r.get_statut_display(), align='center', bg=bg)

    for ci, w in enumerate([12,6,32,20,18,20,9,9,9,12,10], 1):
        ws3.column_dimensions[get_column_letter(ci)].width = w

    # ── Réponse ──────────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"bilan_occupation_kellermann_{saison_label}.xlsx"
    response = HttpResponse(buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
